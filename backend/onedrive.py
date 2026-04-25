"""
Microsoft Graph API integration for OneDrive Excel sync and PDF upload.
Uses MSAL + requests for lightweight implementation.
"""
import os
import io
import logging
import tempfile
from datetime import datetime
from typing import Optional
import msal
import requests
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

logger = logging.getLogger(__name__)

GRAPH_API = "https://graph.microsoft.com/v1.0"


def get_graph_token():
    """Get access token for Microsoft Graph API using client credentials."""
    tenant_id = os.environ.get("AZURE_TENANT_ID")
    client_id = os.environ.get("AZURE_CLIENT_ID")
    client_secret = os.environ.get("AZURE_CLIENT_SECRET")

    if not all([tenant_id, client_id, client_secret]):
        raise ValueError("Azure credentials not configured")

    authority = f"https://login.microsoftonline.com/{tenant_id}"
    app = msal.ConfidentialClientApplication(
        client_id,
        authority=authority,
        client_credential=client_secret
    )

    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])

    if "access_token" not in result:
        error = result.get("error_description", result.get("error", "Unknown error"))
        raise Exception(f"Failed to get token: {error}")

    return result["access_token"]


def graph_headers():
    """Get authorization headers for Graph API calls."""
    token = get_graph_token()
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def find_drive_and_file(file_name: str):
    """
    Find the Excel file across accessible drives.
    Tries path-based access first (ONEDRIVE_FILE_PATH), then search by name.
    Returns (drive_id, item_id) or raises exception.
    """
    headers = graph_headers()
    file_path = os.environ.get("ONEDRIVE_FILE_PATH", "")

    # Strategy 1: Try path-based access on each user's drive
    if file_path:
        logger.info(f"Trying path-based access: {file_path}")
        users_resp = requests.get(f"{GRAPH_API}/users", headers=headers)
        if users_resp.status_code == 200:
            for user in users_resp.json().get("value", []):
                user_id = user["id"]
                drive_resp = requests.get(
                    f"{GRAPH_API}/users/{user_id}/drive",
                    headers={"Authorization": headers["Authorization"]}
                )
                if drive_resp.status_code == 200:
                    drive_id = drive_resp.json().get("id")
                    if drive_id:
                        path_url = f"{GRAPH_API}/drives/{drive_id}/root:/{file_path}"
                        path_resp = requests.get(path_url, headers=headers)
                        if path_resp.status_code == 200:
                            item = path_resp.json()
                            logger.info(f"Found file by path on user {user.get('displayName')}")
                            return drive_id, item["id"]

    # Strategy 2: Search by file name across all user drives
    logger.info(f"Path search failed, trying name search: {file_name}")
    users_resp = requests.get(f"{GRAPH_API}/users", headers=headers)
    if users_resp.status_code == 200:
        for user in users_resp.json().get("value", []):
            user_id = user["id"]
            drive_resp = requests.get(
                f"{GRAPH_API}/users/{user_id}/drive",
                headers={"Authorization": headers["Authorization"]}
            )
            if drive_resp.status_code == 200:
                drive_id = drive_resp.json().get("id")
                if drive_id:
                    search_resp = requests.get(
                        f"{GRAPH_API}/drives/{drive_id}/root/search(q='{file_name}')",
                        headers=headers
                    )
                    if search_resp.status_code == 200:
                        for item in search_resp.json().get("value", []):
                            if file_name.lower() in item.get("name", "").lower():
                                logger.info(f"Found file by search: {item['name']}")
                                return drive_id, item["id"]

    # Strategy 3: Search SharePoint sites
    sites_resp = requests.get(f"{GRAPH_API}/sites?search=*", headers=headers)
    if sites_resp.status_code == 200:
        for site in sites_resp.json().get("value", []):
            site_id = site["id"]
            drives_resp = requests.get(f"{GRAPH_API}/sites/{site_id}/drives", headers=headers)
            if drives_resp.status_code == 200:
                for drive in drives_resp.json().get("value", []):
                    drive_id = drive["id"]
                    if file_path:
                        path_url = f"{GRAPH_API}/drives/{drive_id}/root:/{file_path}"
                        path_resp = requests.get(path_url, headers=headers)
                        if path_resp.status_code == 200:
                            return drive_id, path_resp.json()["id"]
                    search_resp = requests.get(
                        f"{GRAPH_API}/drives/{drive_id}/root/search(q='{file_name}')",
                        headers=headers
                    )
                    if search_resp.status_code == 200:
                        for item in search_resp.json().get("value", []):
                            if file_name.lower() in item.get("name", "").lower():
                                return drive_id, item["id"]

    raise FileNotFoundError(f"File '{file_name}' not found on any accessible OneDrive (path: {file_path})")


def read_excel_employees(file_name: str, sheet_name: str, start_row: int = 17):
    """
    Read employee names (col C) and phones (col B) from Excel on OneDrive.
    Returns list of {name, phone}.
    """
    headers = graph_headers()
    drive_id, item_id = find_drive_and_file(file_name)

    # Download the file content
    download_resp = requests.get(
        f"{GRAPH_API}/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": headers["Authorization"]}
    )

    if download_resp.status_code != 200:
        raise Exception(f"Failed to download file: {download_resp.status_code}")

    # Parse with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(download_resp.content), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        # Try partial match
        matched = [s for s in wb.sheetnames if sheet_name.lower() in s.lower()]
        if matched:
            sheet_name = matched[0]
        else:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    employees = []

    for row in ws.iter_rows(min_row=start_row, max_col=3):
        phone = str(row[1].value).strip() if row[1].value else None
        name = str(row[2].value).strip() if row[2].value else None

        if name and name.lower() != "none" and len(name) > 1:
            employees.append({"name": name, "phone": phone})

    wb.close()
    return employees


SKIP_HEADERS = {"zus", "suma", "kwota", "inne", "murarze", "kierowcy", "dodatki",
                 "kary", "mieszkanie", "zaliczki", "stawka", "wypłaty", "zwroty",
                 "premia", "fegrro"}


def _is_site_column(value):
    """Check if a row-16 cell value is a site identifier (not a financial header)."""
    if value is None:
        return False
    s = str(value).strip().lower()
    if not s or s == "none":
        return False
    for skip in SKIP_HEADERS:
        if skip in s:
            return False
    return True


def _resolve_external_refs(wb, drive_id, auth_header):
    """
    Resolve external workbook references in DANE sheet.
    DANE!B7 → [1]Kody!G15, etc.
    Returns dict mapping DANE cell refs to resolved values.
    """
    import openpyxl
    import re

    if 'DANE' not in wb.sheetnames:
        return {}

    ws_dane = wb['DANE']
    resolved = {}
    ext_cache = {}

    for row in range(7, 25):  # B7-B24 covers sites 1-18
        cell = ws_dane.cell(row=row, column=2)
        formula = cell.value
        if not formula or not isinstance(formula, str) or not formula.startswith('='):
            if formula is not None:
                resolved[f"DANE!B{row}"] = formula
            continue

        # Parse external ref like =[1]Kody!G15
        match = re.match(r"=\[(\d+)\](.+)!([A-Z]+)(\d+)", formula)
        if not match:
            continue

        ext_idx, ext_sheet, ext_col_letter, ext_row = match.groups()
        ext_row = int(ext_row)
        ext_col = openpyxl.utils.column_index_from_string(ext_col_letter)

        if ext_idx not in ext_cache:
            # Find and download external workbook
            file_path = os.environ.get("ONEDRIVE_FILE_PATH", "")
            folder_path = "/".join(file_path.split("/")[:-1]) if "/" in file_path else ""

            if folder_path:
                folder_resp = requests.get(
                    f"{GRAPH_API}/drives/{drive_id}/root:/{folder_path}:/children",
                    headers={"Authorization": auth_header}
                )
                if folder_resp.status_code == 200:
                    for item in folder_resp.json().get("value", []):
                        if item.get("name", "").endswith(".xlsx") and item["name"] != os.environ.get("ONEDRIVE_EXCEL_FILE"):
                            item_resp = requests.get(
                                f"{GRAPH_API}/drives/{drive_id}/items/{item['id']}/content",
                                headers={"Authorization": auth_header}
                            )
                            if item_resp.status_code == 200:
                                try:
                                    ext_wb = openpyxl.load_workbook(
                                        io.BytesIO(item_resp.content), read_only=True, data_only=True
                                    )
                                    if ext_sheet in ext_wb.sheetnames:
                                        ext_cache[ext_idx] = ext_wb
                                        logger.info(f"Resolved external ref [{ext_idx}] → {item['name']}")
                                        break
                                    ext_wb.close()
                                except Exception:
                                    pass

        ext_wb = ext_cache.get(ext_idx)
        if ext_wb and ext_sheet in ext_wb.sheetnames:
            val = ext_wb[ext_sheet].cell(row=ext_row, column=ext_col).value
            resolved[f"DANE!B{row}"] = val

    for wb_ref in ext_cache.values():
        try:
            wb_ref.close()
        except Exception:
            pass

    return resolved


def read_excel_sites(file_name: str, sheet_name: str):
    """
    Read site names from Excel row 16, scanning columns S-Z.
    Resolves formulas that reference DANE sheet and external workbooks.
    Returns list of {col_index, col_letter, name}.
    """
    import openpyxl
    import re

    headers = graph_headers()
    drive_id, item_id = find_drive_and_file(file_name)
    auth_header = headers["Authorization"]

    download_resp = requests.get(
        f"{GRAPH_API}/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": auth_header}
    )
    if download_resp.status_code != 200:
        raise Exception(f"Failed to download file: {download_resp.status_code}")

    file_bytes = download_resp.content

    # First try with data_only=True (cached values)
    wb_data = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    # Also load without data_only to see formulas
    wb_formula = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)

    if sheet_name not in wb_data.sheetnames:
        matched = [s for s in wb_data.sheetnames if sheet_name.lower() in s.lower()]
        if matched:
            sheet_name = matched[0]
        else:
            wb_data.close()
            wb_formula.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb_data.sheetnames}")

    ws_data = wb_data[sheet_name]
    ws_formula = wb_formula[sheet_name]
    sites = []

    # Try to resolve external refs if needed
    ext_resolved = None

    row_16 = list(ws_data.iter_rows(min_row=16, max_row=16, min_col=19, max_col=27))[0]
    row_16_formulas = list(ws_formula.iter_rows(min_row=16, max_row=16, min_col=19, max_col=27))[0]

    for i, (data_cell, formula_cell) in enumerate(zip(row_16, row_16_formulas)):
        col_num = 19 + i
        col_letter = openpyxl.utils.get_column_letter(col_num)

        # Try cached value first
        value = data_cell.value
        formula = formula_cell.value

        # If no cached value but has formula, try to resolve
        if value is None and formula and isinstance(formula, str) and formula.startswith("="):
            # Check if it references DANE sheet
            dane_match = re.match(r"=DANE!([A-Z]+)(\d+)", formula)
            if dane_match:
                if ext_resolved is None:
                    ext_resolved = _resolve_external_refs(wb_formula, drive_id, auth_header)
                dane_ref = f"DANE!{dane_match.group(1)}{dane_match.group(2)}"
                value = ext_resolved.get(dane_ref)

        if value is not None:
            value_str = str(value).strip()
            if _is_site_column(value_str):
                sites.append({
                    "col_index": col_num - 1,
                    "col_letter": col_letter,
                    "name": value_str
                })

    wb_data.close()
    wb_formula.close()
    logger.info(f"Sites found in {sheet_name}: {[s['name'] for s in sites]} (cols {[s['col_letter'] for s in sites]})")
    return sites


def read_excel_employee_rows(file_name: str, sheet_name: str, start_row: int = 17):
    """
    Read employee names and their row numbers from Excel.
    Returns dict {employee_name: row_number}.
    """
    headers = graph_headers()
    drive_id, item_id = find_drive_and_file(file_name)

    download_resp = requests.get(
        f"{GRAPH_API}/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": headers["Authorization"]}
    )
    if download_resp.status_code != 200:
        raise Exception(f"Failed to download file: {download_resp.status_code}")

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(download_resp.content), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        matched = [s for s in wb.sheetnames if sheet_name.lower() in s.lower()]
        if matched:
            sheet_name = matched[0]
        else:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    employee_rows = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_col=3), start=start_row):
        name = str(row[2].value).strip() if row[2].value else None
        if name and name.lower() != "none" and len(name) > 1:
            employee_rows[name] = row_idx

    wb.close()
    return employee_rows


def write_hours_to_excel(file_name: str, sheet_name: str, updates: list):
    """
    Write hour sums back to Excel on OneDrive.
    updates = list of {employee_name, site_name, hours_sum}

    Process:
    1. Download Excel from OneDrive
    2. Find employee row (col C, from row 17) and site column (S-Z, row 16)
    3. Write hours_sum into the intersection cell
    4. Upload modified Excel back to OneDrive
    """
    headers = graph_headers()
    drive_id, item_id = find_drive_and_file(file_name)

    # Download
    download_resp = requests.get(
        f"{GRAPH_API}/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": headers["Authorization"]}
    )
    if download_resp.status_code != 200:
        raise Exception(f"Failed to download file: {download_resp.status_code}")

    import openpyxl
    import re

    file_bytes = download_resp.content
    
    # Validate that we got an actual Excel file (zip format)
    if len(file_bytes) < 100 or file_bytes[:2] != b'PK':
        content_type = download_resp.headers.get('Content-Type', 'unknown')
        preview = file_bytes[:200].decode('utf-8', errors='replace')
        logger.error(f"Downloaded file is not a valid Excel file. Content-Type: {content_type}, Size: {len(file_bytes)}, Preview: {preview}")
        raise Exception(f"Plik Excel jest prawdopodobnie otwarty na innym komputerze. Zamknij plik i spróbuj ponownie. (Content-Type: {content_type})")
    
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

    if sheet_name not in wb.sheetnames:
        matched = [s for s in wb.sheetnames if sheet_name.lower() in s.lower()]
        if matched:
            sheet_name = matched[0]
        else:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Resolve external references for site columns (formulas like =DANE!B7)
    ext_resolved = _resolve_external_refs(wb, drive_id, headers["Authorization"])

    # Build site column mapping from row 16, resolve formulas
    site_columns = {}
    for col in range(19, 28):  # S=19 to AA=27
        cell_val = ws.cell(row=16, column=col).value
        resolved_val = cell_val

        # If it's a formula referencing DANE sheet, resolve it
        if isinstance(cell_val, str) and cell_val.startswith("="):
            dane_match = re.match(r"=DANE!([A-Z]+)(\d+)", cell_val)
            if dane_match:
                dane_ref = f"DANE!{dane_match.group(1)}{dane_match.group(2)}"
                resolved_val = ext_resolved.get(dane_ref, cell_val)

        if _is_site_column(resolved_val):
            site_columns[str(resolved_val).strip()] = col

    # Build employee row mapping from col C (col 3), starting row 17
    employee_rows = {}
    for row in range(17, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=3).value
        if cell_val:
            employee_rows[str(cell_val).strip()] = row

    # First, clear ALL hour cells in site columns so deleted hours don't persist
    cleared = 0
    for row_num in employee_rows.values():
        for col_num in site_columns.values():
            ws.cell(row=row_num, column=col_num, value=0)
            cleared += 1
    logger.info(f"Cleared {cleared} hour cells before writing")

    # Build column letter -> col number mapping
    import openpyxl as _openpyxl
    col_letter_to_num = {}
    for col in range(19, 28):
        col_letter_to_num[_openpyxl.utils.get_column_letter(col)] = col

    # Build case-insensitive lookup maps
    emp_rows_lower = {name.upper(): (name, row) for name, row in employee_rows.items()}
    site_cols_lower = {name.upper(): (name, col) for name, col in site_columns.items()}

    # Write updates
    written = 0
    skipped = []
    for update in updates:
        emp_name = update["employee_name"]
        site_key = update.get("site_key") or update.get("site_name", "")
        hours = update["hours_sum"]

        # Exact match first, then case-insensitive
        row_num = employee_rows.get(emp_name)
        if not row_num:
            match = emp_rows_lower.get(emp_name.upper())
            if match:
                row_num = match[1]

        # Try site_key as column letter first (e.g. "T", "W"), then as site name
        col_num = col_letter_to_num.get(site_key)
        if not col_num:
            col_num = site_columns.get(site_key)
        if not col_num:
            match = site_cols_lower.get(site_key.upper())
            if match:
                col_num = match[1]

        if row_num and col_num:
            ws.cell(row=row_num, column=col_num, value=hours)
            written += 1
        else:
            skipped.append({
                "employee": emp_name,
                "site": site_key,
                "row_found": row_num is not None,
                "col_found": col_num is not None
            })

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    excel_bytes = output.getvalue()
    output.close()

    # Validate output file before uploading
    if excel_bytes[:2] != b'PK' or excel_bytes.rfind(b'PK\x05\x06') == -1:
        raise Exception("Wygenerowany plik Excel jest uszkodzony — zapis anulowany, plik na OneDrive nie został zmieniony.")

    # Upload back to OneDrive
    upload_url = f"{GRAPH_API}/drives/{drive_id}/items/{item_id}/content"
    upload_resp = requests.put(
        upload_url,
        headers={
            "Authorization": headers["Authorization"],
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        },
        data=excel_bytes
    )

    if upload_resp.status_code not in [200, 201]:
        raise Exception(f"Upload failed: {upload_resp.status_code} - {upload_resp.text}")

    return {
        "written": written,
        "skipped": skipped,
        "sheet": sheet_name,
        "sites_found": list(site_columns.keys()),
        "employees_found": len(employee_rows)
    }


def upload_pdf_to_onedrive(pdf_bytes: bytes, file_name: str, folder_name: str = "Archiwizacja"):
    """Upload a PDF file to specified OneDrive folder."""
    headers = graph_headers()

    # Find a drive to upload to (try users' drives)
    users_resp = requests.get(f"{GRAPH_API}/users", headers=headers)
    drive_id = None

    if users_resp.status_code == 200:
        for user in users_resp.json().get("value", []):
            drive_resp = requests.get(
                f"{GRAPH_API}/users/{user['id']}/drive",
                headers={"Authorization": headers["Authorization"]}
            )
            if drive_resp.status_code == 200:
                drive_id = drive_resp.json().get("id")
                if drive_id:
                    break

    if not drive_id:
        raise Exception("No accessible drive found for upload")

    # Upload the file to the archive folder
    upload_url = f"{GRAPH_API}/drives/{drive_id}/root:/{folder_name}/{file_name}:/content"
    upload_resp = requests.put(
        upload_url,
        headers={
            "Authorization": headers["Authorization"],
            "Content-Type": "application/pdf"
        },
        data=pdf_bytes
    )

    if upload_resp.status_code not in [200, 201]:
        raise Exception(f"Upload failed: {upload_resp.status_code} - {upload_resp.text}")

    return upload_resp.json()


def generate_hours_pdf(employees_data: list, month_name: str, year: int, sites: list):
    """
    Generate PDF report with employee hours.
    Returns PDF bytes.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=16, alignment=1, spaceAfter=10
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        fontSize=10, alignment=1, spaceAfter=20
    )

    # Sanitize text for latin-1 (replace Polish chars)
    def sanitize(text):
        replacements = {
            'ą': 'a', 'ć': 'c', 'ę': 'e', 'ł': 'l', 'ń': 'n',
            'ó': 'o', 'ś': 's', 'ź': 'z', 'ż': 'z',
            'Ą': 'A', 'Ć': 'C', 'Ę': 'E', 'Ł': 'L', 'Ń': 'N',
            'Ó': 'O', 'Ś': 'S', 'Ź': 'Z', 'Ż': 'Z'
        }
        for pl_char, ascii_char in replacements.items():
            text = text.replace(pl_char, ascii_char)
        return text

    elements = []
    elements.append(Paragraph(f"FeGrro - Raport godzin pracy", title_style))
    elements.append(Paragraph(sanitize(f"{month_name} {year}"), subtitle_style))
    elements.append(Spacer(1, 5 * mm))

    # Build table data
    header = ["Pracownik"] + [sanitize(s["name"]) for s in sites] + ["SUMA"]
    table_data = [header]

    for emp in employees_data:
        row = [sanitize(emp["name"])]
        total = 0
        for site in sites:
            h = emp.get("site_hours", {}).get(site["id"], 0)
            row.append(str(h) if h else "0")
            total += h
        row.append(str(total))
        table_data.append(row)

    col_widths = [60 * mm] + [30 * mm] * len(sites) + [25 * mm]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2A384C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#334155')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#1E293B'), colors.HexColor('#2A384C')]),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#CBD5E1')),
        ('TEXTCOLOR', (-1, 1), (-1, -1), colors.HexColor('#5F7151')),
        ('FONTNAME', (-1, 0), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 10 * mm))
    elements.append(Paragraph(
        f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        styles['Normal']
    ))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def write_advances_to_excel(file_name: str, sheet_name: str, updates: list):
    """
    Write advance sums to Excel column G (col 7), rows 17-66.
    updates = list of {employee_name, amount}
    """
    headers = graph_headers()
    drive_id, item_id = find_drive_and_file(file_name)

    download_resp = requests.get(
        f"{GRAPH_API}/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": headers["Authorization"]}
    )
    if download_resp.status_code != 200:
        raise Exception(f"Failed to download file: {download_resp.status_code}")

    import openpyxl

    adv_file_bytes = download_resp.content
    if len(adv_file_bytes) < 100 or adv_file_bytes[:2] != b'PK':
        raise Exception("Plik Excel jest prawdopodobnie otwarty na innym komputerze. Zamknij plik i spróbuj ponownie.")

    wb = openpyxl.load_workbook(io.BytesIO(adv_file_bytes))

    if sheet_name not in wb.sheetnames:
        matched = [s for s in wb.sheetnames if sheet_name.lower() in s.lower()]
        if matched:
            sheet_name = matched[0]
        else:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    employee_rows = {}
    for row in range(17, 67):
        cell_val = ws.cell(row=row, column=3).value
        if cell_val:
            employee_rows[str(cell_val).strip()] = row

    # First, clear ALL advance cells (col G) so deleted advances don't persist
    cleared = 0
    for row_num in employee_rows.values():
        ws.cell(row=row_num, column=7, value=0)
        cleared += 1
    logger.info(f"Cleared {cleared} advance cells in column G before writing")

    # Write advances to column G (col 7)
    emp_rows_lower = {name.upper(): (name, row) for name, row in employee_rows.items()}
    written = 0
    skipped = []
    for update in updates:
        emp_name = update["employee_name"]
        amount = update["amount"]
        row_num = employee_rows.get(emp_name)
        if not row_num:
            match = emp_rows_lower.get(emp_name.upper())
            if match:
                row_num = match[1]
        if row_num:
            ws.cell(row=row_num, column=7, value=amount)
            written += 1
        else:
            skipped.append({"employee": emp_name, "amount": amount})

    # Save and upload
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    excel_bytes = output.getvalue()
    output.close()

    if excel_bytes[:2] != b'PK' or excel_bytes.rfind(b'PK\x05\x06') == -1:
        raise Exception("Wygenerowany plik Excel jest uszkodzony — zapis anulowany.")

    upload_url = f"{GRAPH_API}/drives/{drive_id}/items/{item_id}/content"
    upload_resp = requests.put(
        upload_url,
        headers={
            "Authorization": headers["Authorization"],
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        },
        data=excel_bytes
    )

    if upload_resp.status_code not in [200, 201]:
        raise Exception(f"Upload failed: {upload_resp.status_code} - {upload_resp.text}")

    return {
        "written": written,
        "skipped": skipped,
        "sheet": sheet_name,
        "employees_found": len(employee_rows)
    }


def write_penalties_to_excel(file_name: str, sheet_name: str, updates: list):
    """Write penalty sums to Excel column H (col 8), rows 17-66."""
    headers = graph_headers()
    drive_id, item_id = find_drive_and_file(file_name)

    download_resp = requests.get(
        f"{GRAPH_API}/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": headers["Authorization"]}
    )
    if download_resp.status_code != 200:
        raise Exception(f"Failed to download file: {download_resp.status_code}")

    import openpyxl

    pen_file_bytes = download_resp.content
    if len(pen_file_bytes) < 100 or pen_file_bytes[:2] != b'PK':
        raise Exception("Plik Excel jest prawdopodobnie otwarty na innym komputerze. Zamknij plik i spróbuj ponownie.")

    wb = openpyxl.load_workbook(io.BytesIO(pen_file_bytes))

    if sheet_name not in wb.sheetnames:
        matched = [s for s in wb.sheetnames if sheet_name.lower() in s.lower()]
        if matched:
            sheet_name = matched[0]
        else:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    employee_rows = {}
    for row in range(17, 67):
        cell_val = ws.cell(row=row, column=3).value
        if cell_val:
            employee_rows[str(cell_val).strip()] = row

    # First, clear ALL penalty cells (col H) so deleted penalties don't persist
    cleared = 0
    for row_num in employee_rows.values():
        ws.cell(row=row_num, column=8, value=0)
        cleared += 1
    logger.info(f"Cleared {cleared} penalty cells in column H before writing")

    emp_rows_lower = {name.upper(): (name, row) for name, row in employee_rows.items()}
    written = 0
    skipped = []
    for update in updates:
        emp_name = update["employee_name"]
        amount = update["amount"]
        row_num = employee_rows.get(emp_name)
        if not row_num:
            match = emp_rows_lower.get(emp_name.upper())
            if match:
                row_num = match[1]
        if row_num:
            ws.cell(row=row_num, column=8, value=amount)
            written += 1
        else:
            skipped.append({"employee": emp_name, "amount": amount})

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    excel_bytes = output.getvalue()
    output.close()

    if excel_bytes[:2] != b'PK' or excel_bytes.rfind(b'PK\x05\x06') == -1:
        raise Exception("Wygenerowany plik Excel jest uszkodzony — zapis anulowany.")

    upload_url = f"{GRAPH_API}/drives/{drive_id}/items/{item_id}/content"
    upload_resp = requests.put(
        upload_url,
        headers={
            "Authorization": headers["Authorization"],
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        },
        data=excel_bytes
    )

    if upload_resp.status_code not in [200, 201]:
        raise Exception(f"Upload failed: {upload_resp.status_code} - {upload_resp.text}")

    return {
        "written": written,
        "skipped": skipped,
        "sheet": sheet_name,
        "employees_found": len(employee_rows)
    }

