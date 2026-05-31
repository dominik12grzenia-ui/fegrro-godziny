"""Wyceny - standalone quotations/estimates module (iter95k).

Kolekcje MongoDB:
- wyceny             - naglowki wycen (id, name, total_netto)
- wyceny_stages      - etapy wewnatrz wyceny
- wyceny_positions   - pozycje (Etap -> Pozycja)
- wyceny_lines       - linie (Pozycja -> Slot R/M/S, lub child slotu)
- wyceny_price_book  - katalog cen (category: materials/labor/equipment)

Wycena NIE jest powiazana z budowa - to standalone narzedzie ofertowe.
Struktura identyczna jak Budget: stages -> positions -> R/M/S slots -> children.
"""
import uuid
import os
import logging
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr, Field
import io
import base64

from database import db
from auth import get_current_admin, get_current_admin_export

router = APIRouter()
logger = logging.getLogger(__name__)


# =========== MODELE ===========
class WycenaCreate(BaseModel):
    name: str
    notes: Optional[str] = None
    # iter95r: domyslne % dla calej wyceny
    default_gir_pct: Optional[float] = None
    default_dw_pct: Optional[float] = None
    default_koszt_pct: Optional[float] = None
    # iter95s: domyslne narzuty dla subpozycji (materials)
    default_narzut_pct: Optional[float] = None
    default_marza_pct: Optional[float] = None
    # iter95bt: domyslne narzuty dla labor / equipment (rozdzielone)
    default_narzut_labor_pct: Optional[float] = None
    default_narzut_equipment_pct: Optional[float] = None
    # iter95as: dane klienta pre-fill przy tworzeniu
    client_name: Optional[str] = None
    client_nip: Optional[str] = None
    client_address: Optional[str] = None


class WycenaUpdate(BaseModel):
    name: Optional[str] = None
    notes: Optional[str] = None
    default_gir_pct: Optional[float] = None
    default_dw_pct: Optional[float] = None
    default_koszt_pct: Optional[float] = None
    default_narzut_pct: Optional[float] = None
    default_marza_pct: Optional[float] = None
    # iter95bt: rozdzielone narzuty labor / equipment
    default_narzut_labor_pct: Optional[float] = None
    default_narzut_equipment_pct: Optional[float] = None
    # iter95al: dane klienta dla PDF wersji dla klienta
    client_name: Optional[str] = None
    client_nip: Optional[str] = None
    client_address: Optional[str] = None
    # iter95am: powierzchnie budynku (PC = pow. calkowita, PUM = pow. uzytkowa mieszkalna)
    pc_m2: Optional[float] = None
    pum_m2: Optional[float] = None
    # iter95an: podzial PC na podziemie/nadziemie
    pc_podziemie_m2: Optional[float] = None
    pc_nadziemie_m2: Optional[float] = None
    # iter95ar: zapisany szablon emaila zapytania ofertowego (per wycena)
    bom_email_subject: Optional[str] = None
    bom_email_body: Optional[str] = None
    # iter95w: zakres oferty - co obejmuje / czego nie obejmuje
    scope_includes: Optional[str] = None
    scope_excludes: Optional[str] = None
    # iter95ab: tryb zaokraglania finalnej ceny jednostkowej w widoku oferty
    # 'none' (default) | 'natural' | 'up' | 'down'
    rounding_mode: Optional[str] = "natural"
    # iter95bm: tryb negocjacji - gdy True, ceny < price_min sa odrzucane (HTTP 400).
    # Auto-set price_min przy pierwszym wpisie dla materials/equipment.
    negotiation_mode: Optional[bool] = None
    # iter95bn: prog ostrzegawczy marzy % w trybie negocjacji (default 10).
    # Gdy marza po negocjacji spadnie ponizej, panel pokazuje czerwone ostrzezenie.
    min_margin_pct: Optional[float] = None


class StageCreate(BaseModel):
    wycena_id: str
    name: str
    order: int = 0


class PositionCreate(BaseModel):
    wycena_id: str
    stage_id: str
    name: str
    order: int = 0
    # iter95u: ilosc wpisywana recznie na poziomie pozycji glownej
    quantity: Optional[float] = None
    # iter95x: jednostka miary (mb/m2/m3/szt/...)
    unit: Optional[str] = None
    # iter95q: pola dla widoku Excel-style wyceny
    kaucja_gir_pct: Optional[float] = None    # %, domyslnie 2.0
    kaucja_dw_pct: Optional[float] = None     # %, domyslnie 2.0
    koszt_budowy_pct: Optional[float] = None  # %, domyslnie 2.0
    koszt_prognozowany: Optional[float] = None  # kwota wpisywana recznie


class PositionUpdate(BaseModel):
    name: Optional[str] = None
    order: Optional[int] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    kaucja_gir_pct: Optional[float] = None
    kaucja_dw_pct: Optional[float] = None
    koszt_budowy_pct: Optional[float] = None
    koszt_prognozowany: Optional[float] = None
    # iter95am: flagi wliczania budzetu pozycji do wskaznika PC / PUM (zl/m2)
    include_in_pc: Optional[bool] = None
    include_in_pum: Optional[bool] = None
    # iter95an: podzial PC na podziemie / nadziemie
    include_in_pc_podziemie: Optional[bool] = None
    include_in_pc_nadziemie: Optional[bool] = None


class LineCreate(BaseModel):
    wycena_id: str
    stage_id: str
    position_id: str
    parent_id: Optional[str] = None
    type: str = "materials"  # materials | labor | equipment
    name: str
    unit: Optional[str] = None
    quantity: float = 0
    unit_price_netto: float = 0
    order: int = 0
    # iter95s: narzut na zapas + marza (procentowe - mnoza budzet)
    narzut_zapas_pct: Optional[float] = None
    marza_pct: Optional[float] = None
    # iter95ae: formula obliczania ilosci (np. "=100 m² * 0,24 m")
    quantity_formula: Optional[str] = None
    # iter95ab: ceny graniczne (kontrola minimum przy negocjacji)
    price_min: Optional[float] = None
    price_max: Optional[float] = None
    # iter95ab: cennik_id - link do CenniK aby propagowac price_min/max
    price_book_id: Optional[str] = None


class LineUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = None
    unit_price_netto: Optional[float] = None
    type: Optional[str] = None
    order: Optional[int] = None
    narzut_zapas_pct: Optional[float] = None
    marza_pct: Optional[float] = None
    quantity_formula: Optional[str] = None
    # iter95ab: ceny graniczne (override z pozycji wyceny)
    price_min: Optional[float] = None
    price_max: Optional[float] = None
    # iter95ab: czy admin zaakceptowal cene ponizej minimum
    below_min_accepted: Optional[bool] = None
    below_min_reason: Optional[str] = None
    price_book_id: Optional[str] = None


class PriceBookCreate(BaseModel):
    category: str  # materials | labor | equipment
    name: str
    unit: Optional[str] = None
    unit_price_netto: float = 0
    # iter95ab: ceny graniczne uzywane przy negocjacjach (kontrola minimum)
    price_min: Optional[float] = None
    price_max: Optional[float] = None
    notes: Optional[str] = None
    # iter95l: dodatkowe pola dla materials (Excel-style)
    sub_category: Optional[str] = None       # izolacje | betony | stal | murowane | drobnica | pozostale
    oferent: Optional[str] = None            # dostawca / oferent
    opakowanie: Optional[str] = None         # wiaderko / paleta / rolka...
    pkg_qty: Optional[float] = None          # ilosc w opakowaniu (np. 20)
    pkg_unit: Optional[str] = None           # jd opakowania (kg, m2, szt)
    zapotrzebowanie: Optional[float] = None  # norma zuzycia (np. 1.5)
    zap_unit: Optional[str] = None           # jd. do jd. (kg/m2, szt/m2, mb/mb)
    liczba_warstw: Optional[float] = None    # liczba warstw
    koszty_inne_do_jd: Optional[float] = None  # iter95y: koszty inne doliczane do jednostki
    # iter95m: pola dla LABOR (robocizna)
    price_m2: Optional[float] = None         # cena za m2
    price_m3: Optional[float] = None         # cena za m3
    # iter95bo: dodatkowa cena za dowolna jednostke (mb/szt/kpl/godz/dzien/kg/t)
    price_other: Optional[float] = None
    unit_other: Optional[str] = None         # np. "mb", "szt", "kpl", "godz"
    # iter95n: pola dla EQUIPMENT (sprzet)
    price_hour: Optional[float] = None       # koszt za godzine
    price_day: Optional[float] = None        # koszt za dzien
    price_month: Optional[float] = None      # koszt za miesiac
    wynajmujacy: Optional[str] = None        # nazwa firmy wynajmujacej
    extra_cost: Optional[float] = None       # koszty poboczne doliczane do kazdej jednostki
    extra_cost_desc: Optional[str] = None    # opis kosztow pobocznych


class PriceBookUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    unit_price_netto: Optional[float] = None
    # iter95ab: ceny graniczne
    price_min: Optional[float] = None
    price_max: Optional[float] = None
    notes: Optional[str] = None
    # iter95l: dodatkowe pola dla materials
    sub_category: Optional[str] = None
    oferent: Optional[str] = None
    opakowanie: Optional[str] = None
    pkg_qty: Optional[float] = None
    pkg_unit: Optional[str] = None
    zapotrzebowanie: Optional[float] = None
    zap_unit: Optional[str] = None
    liczba_warstw: Optional[float] = None
    koszty_inne_do_jd: Optional[float] = None
    # iter95m: labor
    price_m2: Optional[float] = None
    price_m3: Optional[float] = None
    # iter95bo: trzecia jednostka (mb/szt/kpl/godz/dzien/kg/t)
    price_other: Optional[float] = None
    unit_other: Optional[str] = None
    # iter95n: equipment
    price_hour: Optional[float] = None
    price_day: Optional[float] = None
    price_month: Optional[float] = None
    wynajmujacy: Optional[str] = None
    extra_cost: Optional[float] = None
    extra_cost_desc: Optional[str] = None


VALID_CATEGORIES = {"materials", "labor", "equipment"}


def _ensure_cat(cat: str):
    if cat not in VALID_CATEGORIES:
        raise HTTPException(400, f"Kategoria musi byc jedna z: {sorted(VALID_CATEGORIES)}")


# =========== WYCENY (naglowki) ===========
@router.get("/wyceny")
async def list_wyceny(_user: dict = Depends(get_current_admin)):
    # iter95bj: optymalizacja N+1 query. Wczesniej dla kazdej wyceny robilismy
    # osobny find na wyceny_lines (O(N*M)). Teraz 2 zapytania: 1) wyceny, 2) wszystkie
    # linie naraz przez $in i grupowanie w Pythonie.
    rows = await db.wyceny.find({}, {"_id": 0}).sort([("created_at", -1)]).to_list(length=1000)
    if not rows:
        return {"rows": []}
    wycena_ids = [w["id"] for w in rows]
    all_lines = await db.wyceny_lines.find(
        {"wycena_id": {"$in": wycena_ids}},
        {"_id": 0, "wycena_id": 1, "id": 1, "parent_id": 1, "quantity": 1, "unit_price_netto": 1},
    ).to_list(length=None)
    # Grupuj w Pythonie - jeden przejazd
    lines_by_wycena = {}
    for ln in all_lines:
        lines_by_wycena.setdefault(ln["wycena_id"], []).append(ln)
    for w in rows:
        lines = lines_by_wycena.get(w["id"], [])
        parent_ids = {ln.get("parent_id") for ln in lines if ln.get("parent_id")}
        total = 0.0
        for ln in lines:
            if ln["id"] in parent_ids:
                continue  # ma dzieci - pomijamy (liczymy tylko liscie)
            total += float(ln.get("quantity") or 0) * float(ln.get("unit_price_netto") or 0)
        w["total_netto"] = round(total, 2)
        w["lines_count"] = len(lines)
    return {"rows": rows}


# iter95as: lista unikalnych klientow z istniejacych wycen (do autouzupelniania)
@router.get("/wyceny/clients")
async def list_wycena_clients(_user: dict = Depends(get_current_admin)):
    """Zwroc unikalnych klientow (po nazwie) z poprzednich wycen + ich dane (nip, adres)."""
    rows = await db.wyceny.find(
        {"client_name": {"$exists": True, "$ne": ""}},
        {"_id": 0, "client_name": 1, "client_nip": 1, "client_address": 1, "created_at": 1},
    ).sort("created_at", -1).to_list(length=500)
    seen = {}
    for r in rows:
        cname = (r.get("client_name") or "").strip()
        if not cname:
            continue
        key = cname.lower()
        if key not in seen:
            seen[key] = {
                "name": cname,
                "nip": (r.get("client_nip") or "").strip(),
                "address": (r.get("client_address") or "").strip(),
            }
    return {"rows": sorted(seen.values(), key=lambda x: x["name"].lower())}


@router.post("/wyceny")
async def create_wycena(payload: WycenaCreate, current_user: dict = Depends(get_current_admin)):
    wid = str(uuid.uuid4())
    doc = {
        "id": wid,
        "name": payload.name,
        "notes": payload.notes,
        "default_gir_pct": payload.default_gir_pct if payload.default_gir_pct is not None else 2.0,
        "default_dw_pct": payload.default_dw_pct if payload.default_dw_pct is not None else 2.0,
        "default_koszt_pct": payload.default_koszt_pct if payload.default_koszt_pct is not None else 2.0,
        "default_narzut_pct": payload.default_narzut_pct if payload.default_narzut_pct is not None else 0.0,
        "default_marza_pct": payload.default_marza_pct if payload.default_marza_pct is not None else 0.0,
        # iter95bt: rozdzielone narzuty per typ linii
        "default_narzut_labor_pct": payload.default_narzut_labor_pct if payload.default_narzut_labor_pct is not None else 0.0,
        "default_narzut_equipment_pct": payload.default_narzut_equipment_pct if payload.default_narzut_equipment_pct is not None else 0.0,
        # iter95as: dane klienta pre-fill
        "client_name": payload.client_name or "",
        "client_nip": payload.client_nip or "",
        "client_address": payload.client_address or "",
        "created_at": datetime.now().isoformat(),
        "created_by": current_user["sub"],
    }
    await db.wyceny.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.patch("/wyceny/{wycena_id}")
async def update_wycena(wycena_id: str, payload: WycenaUpdate,
                         current_user: dict = Depends(get_current_admin)):
    raw = payload.dict(exclude_unset=True)
    updates = dict(raw)
    updates["updated_at"] = datetime.now().isoformat()
    res = await db.wyceny.update_one({"id": wycena_id}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "Wycena nie istnieje")
    doc = await db.wyceny.find_one({"id": wycena_id}, {"_id": 0})
    return doc


@router.delete("/wyceny/{wycena_id}")
async def delete_wycena(wycena_id: str, _user: dict = Depends(get_current_admin)):
    res = await db.wyceny.delete_one({"id": wycena_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Wycena nie istnieje")
    # Kaskadowo
    await db.wyceny_stages.delete_many({"wycena_id": wycena_id})
    await db.wyceny_positions.delete_many({"wycena_id": wycena_id})
    await db.wyceny_lines.delete_many({"wycena_id": wycena_id})
    return {"ok": True}


# =========== TEMPLATE (dane do widoku Excel-style) ===========
@router.get("/wyceny/{wycena_id}/template")
async def get_template(wycena_id: str, _user: dict = Depends(get_current_admin)):
    """Zwraca strukture wyceny: stages -> positions -> lines (+ children)."""
    wycena = await db.wyceny.find_one({"id": wycena_id}, {"_id": 0})
    if not wycena:
        raise HTTPException(404, "Wycena nie istnieje")
    stages = await db.wyceny_stages.find({"wycena_id": wycena_id}, {"_id": 0}).sort([("order", 1)]).to_list(length=None)
    positions = await db.wyceny_positions.find({"wycena_id": wycena_id}, {"_id": 0}).sort([("order", 1)]).to_list(length=None)
    lines = await db.wyceny_lines.find({"wycena_id": wycena_id}, {"_id": 0}).sort([("order", 1)]).to_list(length=None)

    pos_by_stage: dict = {}
    for p in positions:
        pos_by_stage.setdefault(p["stage_id"], []).append(p)
    lines_by_pos: dict = {}
    for ln in lines:
        lines_by_pos.setdefault(ln.get("position_id"), []).append(ln)

    # Struktura: stages[].positions[].slots[] (sloty top-level R/M/S) + slot.children
    for st in stages:
        st["positions"] = []
        for p in pos_by_stage.get(st["id"], []):
            slots = []
            all_pos_lines = lines_by_pos.get(p["id"], [])
            top_level = [ln for ln in all_pos_lines if not ln.get("parent_id")]
            children_by_parent: dict = {}
            for ln in all_pos_lines:
                if ln.get("parent_id"):
                    children_by_parent.setdefault(ln["parent_id"], []).append(ln)
            for slot in top_level:
                slot_copy = dict(slot)
                slot_copy["children"] = children_by_parent.get(slot["id"], [])
                slots.append(slot_copy)
            p_copy = dict(p)
            p_copy["slots"] = slots
            st["positions"].append(p_copy)
    return {"wycena": wycena, "stages": stages}


# =========== STAGES ===========
@router.post("/wyceny/stages")
async def create_stage(payload: StageCreate, current_user: dict = Depends(get_current_admin)):
    if not await db.wyceny.find_one({"id": payload.wycena_id}, {"_id": 0, "id": 1}):
        raise HTTPException(400, "Wycena nie istnieje")
    sid = str(uuid.uuid4())
    doc = {
        "id": sid, "wycena_id": payload.wycena_id, "name": payload.name,
        "order": payload.order,
        "created_at": datetime.now().isoformat(),
    }
    await db.wyceny_stages.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.patch("/wyceny/stages/{stage_id}")
async def update_stage(stage_id: str, payload: StageCreate, _user: dict = Depends(get_current_admin)):
    updates = {"name": payload.name, "order": payload.order, "updated_at": datetime.now().isoformat()}
    res = await db.wyceny_stages.update_one({"id": stage_id}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "Etap nie istnieje")
    return {"ok": True}


@router.delete("/wyceny/stages/{stage_id}")
async def delete_stage(stage_id: str, _user: dict = Depends(get_current_admin)):
    res = await db.wyceny_stages.delete_one({"id": stage_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Etap nie istnieje")
    # Kaskadowo
    await db.wyceny_positions.delete_many({"stage_id": stage_id})
    await db.wyceny_lines.delete_many({"stage_id": stage_id})
    return {"ok": True}


# =========== POSITIONS ===========
@router.post("/wyceny/positions")
async def create_position(payload: PositionCreate, _user: dict = Depends(get_current_admin)):
    pid = str(uuid.uuid4())
    doc = {
        "id": pid, "wycena_id": payload.wycena_id, "stage_id": payload.stage_id,
        "name": payload.name, "order": payload.order,
        "quantity": payload.quantity,
        "unit": payload.unit,
        "kaucja_gir_pct": payload.kaucja_gir_pct if payload.kaucja_gir_pct is not None else 2.0,
        "kaucja_dw_pct": payload.kaucja_dw_pct if payload.kaucja_dw_pct is not None else 2.0,
        "koszt_budowy_pct": payload.koszt_budowy_pct if payload.koszt_budowy_pct is not None else 2.0,
        "koszt_prognozowany": payload.koszt_prognozowany,
        "created_at": datetime.now().isoformat(),
    }
    await db.wyceny_positions.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.patch("/wyceny/positions/{position_id}")
async def update_position(position_id: str, payload: PositionUpdate, _user: dict = Depends(get_current_admin)):
    raw = payload.dict(exclude_unset=True)
    updates = dict(raw)
    updates["updated_at"] = datetime.now().isoformat()
    res = await db.wyceny_positions.update_one({"id": position_id}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "Pozycja nie istnieje")
    return {"ok": True}


@router.delete("/wyceny/positions/{position_id}")
async def delete_position(position_id: str, _user: dict = Depends(get_current_admin)):
    res = await db.wyceny_positions.delete_one({"id": position_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Pozycja nie istnieje")
    await db.wyceny_lines.delete_many({"position_id": position_id})
    return {"ok": True}


# iter95ao: bulk apply flagi PC/PC_pod/PC_nad/PUM na wszystkie pozycje w etapie
class StageBulkFlag(BaseModel):
    flag: str  # 'include_in_pc' | 'include_in_pc_podziemie' | 'include_in_pc_nadziemie' | 'include_in_pum'
    value: bool


ALLOWED_BULK_FLAGS = {
    "include_in_pc", "include_in_pc_podziemie", "include_in_pc_nadziemie", "include_in_pum"
}


@router.post("/wyceny/stages/{stage_id}/bulk-flag")
async def stage_bulk_flag(stage_id: str, payload: StageBulkFlag,
                          _user: dict = Depends(get_current_admin)):
    if payload.flag not in ALLOWED_BULK_FLAGS:
        raise HTTPException(400, "Nieprawidlowa flaga")
    res = await db.wyceny_positions.update_many(
        {"stage_id": stage_id},
        {"$set": {payload.flag: payload.value, "updated_at": datetime.now().isoformat()}},
    )
    return {"ok": True, "modified": res.modified_count}


# =========== LINES (slots R/M/S + children) ===========
@router.post("/wyceny/lines")
async def create_line(payload: LineCreate, _user: dict = Depends(get_current_admin)):
    if payload.type not in {"materials", "labor", "equipment"}:
        raise HTTPException(400, "Nieprawidlowy typ")
    lid = str(uuid.uuid4())
    # iter95bm: auto-set price_min przy tworzeniu linii dla materials/equipment.
    # Bierze MAX z (ceny wpisanej, price_min z cennika jezeli wybrano).
    eff_price_min = payload.price_min
    if payload.type in ("materials", "equipment") and (payload.unit_price_netto or 0) > 0:
        candidates = []
        if payload.price_min is not None:
            candidates.append(float(payload.price_min))
        if payload.unit_price_netto:
            candidates.append(float(payload.unit_price_netto))
        if payload.price_book_id:
            pb = await db.wyceny_price_book.find_one(
                {"id": payload.price_book_id}, {"_id": 0, "price_min": 1}
            )
            if pb and pb.get("price_min") is not None:
                candidates.append(float(pb["price_min"]))
        if candidates:
            eff_price_min = max(candidates)
    # iter95bm: dla labor - przy wybraniu z cennika kopiuje price_min/max z cennika
    eff_price_max = payload.price_max
    if payload.type == "labor" and payload.price_book_id:
        pb = await db.wyceny_price_book.find_one(
            {"id": payload.price_book_id}, {"_id": 0, "price_min": 1, "price_max": 1}
        )
        if pb:
            if eff_price_min is None and pb.get("price_min") is not None:
                eff_price_min = float(pb["price_min"])
            if eff_price_max is None and pb.get("price_max") is not None:
                eff_price_max = float(pb["price_max"])
    doc = {
        "id": lid, "wycena_id": payload.wycena_id, "stage_id": payload.stage_id,
        "position_id": payload.position_id, "parent_id": payload.parent_id,
        "type": payload.type, "name": payload.name,
        "unit": payload.unit, "quantity": payload.quantity,
        "unit_price_netto": payload.unit_price_netto,
        "narzut_zapas_pct": payload.narzut_zapas_pct,
        "marza_pct": payload.marza_pct,
        "quantity_formula": payload.quantity_formula,
        "price_min": eff_price_min,
        "price_max": eff_price_max,
        "price_book_id": payload.price_book_id,
        "order": payload.order,
        "created_at": datetime.now().isoformat(),
    }
    await db.wyceny_lines.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.patch("/wyceny/lines/{line_id}")
async def update_line(line_id: str, payload: LineUpdate, current_user: dict = Depends(get_current_admin)):
    # iter95ab: zachowaj historie zmian ceny dla audytu negocjacji
    existing = await db.wyceny_lines.find_one({"id": line_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Linia nie istnieje")

    raw = payload.dict(exclude_unset=True)
    # uzywamy exclude_unset zeby None mogly byc zapisane explicit dla price_min/max
    clearable = {"price_min", "price_max", "below_min_reason"}
    updates = {k: v for k, v in raw.items() if v is not None or k in clearable}

    # iter95bm: gdy zmieniono price_book_id, skopiuj price_min/max z cennika.
    # Materials/Equipment: MAX(istniejacy, book) | Labor: tylko jak brak.
    if "price_book_id" in raw and raw.get("price_book_id"):
        pb = await db.wyceny_price_book.find_one(
            {"id": raw["price_book_id"]}, {"_id": 0, "price_min": 1, "price_max": 1}
        )
        if pb:
            ltype = existing.get("type")
            cur_min = updates.get("price_min", existing.get("price_min"))
            cur_max = updates.get("price_max", existing.get("price_max"))
            book_min = pb.get("price_min")
            book_max = pb.get("price_max")
            if book_min is not None:
                if ltype in ("materials", "equipment"):
                    updates["price_min"] = max(float(cur_min) if cur_min is not None else book_min, float(book_min))
                elif cur_min is None:
                    updates["price_min"] = float(book_min)
            if book_max is not None and cur_max is None:
                updates["price_max"] = float(book_max)

    # iter95ab: zapis historii zmian ceny do price_change_history
    new_price = raw.get("unit_price_netto")
    old_price = existing.get("unit_price_netto")

    # Pobierz effective_min (z linii lub cennika)
    line_min = updates.get("price_min", existing.get("price_min"))
    pb_min = None
    if line_min is None and existing.get("price_book_id"):
        pb_doc = await db.wyceny_price_book.find_one(
            {"id": existing["price_book_id"]}, {"_id": 0, "price_min": 1}
        )
        pb_min = (pb_doc or {}).get("price_min")
    effective_min = line_min if line_min is not None else pb_min

    if new_price is not None and old_price is not None and abs(new_price - old_price) > 1e-9:
        # iter95bm: walidacja trybu negocjacji - blokuje zapis ponizej min
        if not raw.get("below_min_accepted"):
            wycena = await db.wyceny.find_one(
                {"id": existing["wycena_id"]}, {"_id": 0, "negotiation_mode": 1}
            )
            neg_mode = (wycena or {}).get("negotiation_mode", False)
            if neg_mode and effective_min is not None and float(new_price) < float(effective_min) - 1e-9:
                raise HTTPException(
                    400,
                    f"Tryb negocjacji: cena {new_price:.2f} zl < minimum {float(effective_min):.2f} zl. "
                    f"Wylacz tryb negocjacji lub akceptuj ceny ponizej minimum.",
                )

        # iter95bm: auto-set price_min przy pierwszym wpisaniu ceny dla materials/equipment
        # (jezeli price_min jest jeszcze nieustawione)
        if (
            existing.get("type") in ("materials", "equipment")
            and effective_min is None
            and float(new_price) > 0
        ):
            updates["price_min"] = float(new_price)
            effective_min = float(new_price)

        below_min = effective_min is not None and new_price < effective_min

        history_entry = {
            "ts": datetime.now().isoformat(),
            "user_id": current_user.get("sub"),
            "user_email": current_user.get("email"),
            "from_price": float(old_price),
            "to_price": float(new_price),
            "min_price": float(effective_min) if effective_min is not None else None,
            "below_min": below_min,
            "reason": raw.get("below_min_reason"),
        }
        updates.setdefault("price_change_history", existing.get("price_change_history", []) + [history_entry])

    updates["updated_at"] = datetime.now().isoformat()
    res = await db.wyceny_lines.update_one({"id": line_id}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "Linia nie istnieje")
    return {"ok": True}


@router.delete("/wyceny/lines/{line_id}")
async def delete_line(line_id: str, _user: dict = Depends(get_current_admin)):
    # Kaskada: usun tez dzieci
    await db.wyceny_lines.delete_many({"parent_id": line_id})
    res = await db.wyceny_lines.delete_one({"id": line_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Linia nie istnieje")
    return {"ok": True}


# =========== PRICE BOOK (Cennik) ===========
@router.get("/wyceny/cennik")
async def list_price_book(
    category: Optional[str] = Query(None),
    q: Optional[str] = Query(None),  # search by name
    _user: dict = Depends(get_current_admin),
):
    fq: dict = {}
    if category:
        _ensure_cat(category)
        fq["category"] = category
    if q:
        fq["name"] = {"$regex": q, "$options": "i"}
    rows = await db.wyceny_price_book.find(fq, {"_id": 0}).sort([("name", 1)]).to_list(length=5000)
    return {"rows": rows}


@router.post("/wyceny/cennik")
async def create_price_book(payload: PriceBookCreate, current_user: dict = Depends(get_current_admin)):
    _ensure_cat(payload.category)
    pid = str(uuid.uuid4())
    doc = {
        "id": pid, "category": payload.category, "name": payload.name,
        "unit": payload.unit, "unit_price_netto": payload.unit_price_netto,
        "notes": payload.notes,
        # iter95ab: ceny graniczne (negocjacje)
        "price_min": payload.price_min,
        "price_max": payload.price_max,
        # iter95l: extended fields (materials)
        "sub_category": payload.sub_category,
        "oferent": payload.oferent,
        "opakowanie": payload.opakowanie,
        "pkg_qty": payload.pkg_qty,
        "pkg_unit": payload.pkg_unit,
        "zapotrzebowanie": payload.zapotrzebowanie,
        "zap_unit": payload.zap_unit,
        "liczba_warstw": payload.liczba_warstw,
        "koszty_inne_do_jd": payload.koszty_inne_do_jd,
        # iter95m: labor
        "price_m2": payload.price_m2,
        "price_m3": payload.price_m3,
        # iter95bo: trzecia jednostka labor (mb/szt/kpl/godz/dzien/kg/t)
        "price_other": payload.price_other,
        "unit_other": payload.unit_other,
        # iter95n: equipment
        "price_hour": payload.price_hour,
        "price_day": payload.price_day,
        "price_month": payload.price_month,
        "wynajmujacy": payload.wynajmujacy,
        "extra_cost": payload.extra_cost,
        "extra_cost_desc": payload.extra_cost_desc,
        "price_history": [],  # iter95m: lista wpisow {date, field, old, new}
        "created_at": datetime.now().isoformat(),
        "created_by": current_user["sub"],
    }
    await db.wyceny_price_book.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.patch("/wyceny/cennik/{item_id}")
async def update_price_book(item_id: str, payload: PriceBookUpdate, _user: dict = Depends(get_current_admin)):
    existing = await db.wyceny_price_book.find_one({"id": item_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Pozycja cennika nie istnieje")
    raw = payload.dict(exclude_unset=True)
    updates = dict(raw)
    updates["updated_at"] = datetime.now().isoformat()
    # iter95m: dla labor - sledz zmiany price_m2/price_m3/price_other w price_history
    history_entries = []
    if existing.get("category") == "labor":
        for field in ("price_m2", "price_m3", "price_other", "unit_price_netto"):
            if field in raw:
                old_val = existing.get(field)
                new_val = raw[field]
                if old_val != new_val and (old_val is not None or new_val is not None):
                    history_entries.append({
                        "date": datetime.now().isoformat(),
                        "field": field,
                        "old": old_val,
                        "new": new_val,
                    })
    if history_entries:
        # Append entries do price_history
        await db.wyceny_price_book.update_one(
            {"id": item_id},
            {"$set": updates, "$push": {"price_history": {"$each": history_entries}}},
        )
    else:
        await db.wyceny_price_book.update_one({"id": item_id}, {"$set": updates})
    return {"ok": True}


@router.delete("/wyceny/cennik/{item_id}")
async def delete_price_book(item_id: str, _user: dict = Depends(get_current_admin)):
    res = await db.wyceny_price_book.delete_one({"id": item_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Pozycja cennika nie istnieje")
    return {"ok": True}


# iter95af/ag: zestawienie potrzebnych materialow (Bill of Materials / RFQ)
async def _build_bom(wycena_id: str):
    """Zestawienie: grupuj po nazwie, sumuj ilosci, dolicz liczbe opakowan (ceil)
    na bazie cennika (pkg_qty + zapotrzebowanie + zap_unit)."""
    import math
    wycena = await db.wyceny.find_one({"id": wycena_id}, {"_id": 0})
    if not wycena:
        raise HTTPException(404, "Wycena nie istnieje")
    lines = await db.wyceny_lines.find(
        {"wycena_id": wycena_id, "type": "materials"}, {"_id": 0}
    ).to_list(length=None)
    # Pobierz cennik materialow do dopasowania po nazwie
    cennik = await db.wyceny_price_book.find(
        {"category": "materials"}, {"_id": 0}
    ).to_list(length=None)
    cennik_by_name = {(c.get("name") or "").lower().strip(): c for c in cennik}

    grouped: dict = {}
    for ln in lines:
        name = (ln.get("name") or "").strip()
        if not name:
            continue
        unit = (ln.get("unit") or "").strip()
        qty = float(ln.get("quantity") or 0)
        if qty <= 0:
            continue
        # Grupuj po (nazwa, jednostka) - to samo co przedtem
        key = (name.lower(), unit.lower())
        if key not in grouped:
            ce = cennik_by_name.get(name.lower(), {}) or {}
            grouped[key] = {
                "name": name,
                "unit": unit,
                "quantity": 0.0,
                "occurrences": 0,
                "opakowanie": ce.get("opakowanie") or "",
                "pkg_qty": ce.get("pkg_qty"),
                "pkg_unit": ce.get("pkg_unit") or "",
                "zapotrzebowanie": ce.get("zapotrzebowanie"),
                "zap_unit": ce.get("zap_unit") or "",
                # iter95ar: sub_category z cennika do filtrowania zapytan (izolacje/stal/...)
                "sub_category": ce.get("sub_category") or "",
            }
        grouped[key]["quantity"] += qty
        grouped[key]["occurrences"] += 1

    rows = []
    for r in grouped.values():
        # iter95ag: wylicz liczbe opakowan
        qty_in_pkg_unit = None
        num_packages = None
        pkg_qty = r.get("pkg_qty")
        zap = r.get("zapotrzebowanie")
        zap_unit = r.get("zap_unit") or ""
        pkg_unit = r.get("pkg_unit") or ""
        if pkg_qty and pkg_qty > 0:
            # Wariant 1: jednostka linii pasuje do mianownika zap_unit (np. line=m², zap=kg/m²)
            if "/" in zap_unit and zap:
                _num, denom = zap_unit.split("/", 1)
                if denom.strip() == r["unit"]:
                    qty_in_pkg_unit = r["quantity"] * zap
            # Wariant 2: jednostka linii = jd. opakowania (np. line=kg, pkg=kg)
            if qty_in_pkg_unit is None and r["unit"] and r["unit"] == pkg_unit:
                qty_in_pkg_unit = r["quantity"]
            if qty_in_pkg_unit is not None:
                num_packages = math.ceil(qty_in_pkg_unit / pkg_qty)
        rows.append({
            "name": r["name"],
            "unit": r["unit"],
            "quantity": round(r["quantity"], 3),
            "occurrences": r["occurrences"],
            "opakowanie": r["opakowanie"],
            "pkg_qty": r["pkg_qty"],
            "pkg_unit": pkg_unit,
            "qty_in_pkg_unit": round(qty_in_pkg_unit, 3) if qty_in_pkg_unit is not None else None,
            "num_packages": num_packages,
            "sub_category": r.get("sub_category") or "",
        })
    rows.sort(key=lambda r: r["name"].lower())
    return {"wycena_name": wycena.get("name", ""), "rows": rows}


@router.get("/wyceny/{wycena_id}/bom")
async def get_materials_bom(wycena_id: str, _user: dict = Depends(get_current_admin)):
    """JSON: zestawienie materialow do podgladu we frontendzie."""
    return await _build_bom(wycena_id)


@router.get("/wyceny/{wycena_id}/bom.xlsx")
async def export_bom_xlsx(
    wycena_id: str,
    subcategories: Optional[str] = Query(None, description="Lista sub_categories oddzielona przecinkiem"),
    _user: dict = Depends(get_current_admin),
):
    data = await _build_bom(wycena_id)
    if subcategories:
        data = _filter_bom_rows(data, [s.strip() for s in subcategories.split(",") if s.strip()])
    content, filename = _generate_bom_xlsx_bytes(data)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _get_logo_path() -> Optional[str]:
    """iter95av: zwraca pierwszą istniejącą ścieżkę do logo firmy (do PDF/Excel)."""
    import os as _os
    candidates = [
        "/app/frontend/public/icon-512x512.png",
        "/app/frontend/public/icon-192x192.png",
        "/app/frontend/public/apple-touch-icon.png",
    ]
    return next((p for p in candidates if _os.path.exists(p)), None)


def _xlsx_add_logo(ws, anchor: str = "A1", width: int = 110, height: int = 110) -> None:
    """iter95av: wstawia logo w arkuszu XLSX na danej kotwicy. Cicho ignoruje błąd."""
    try:
        path = _get_logo_path()
        if not path:
            return
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(path)
        img.width = width
        img.height = height
        img.anchor = anchor
        ws.add_image(img)
    except Exception:
        pass


def _filter_bom_rows(data: dict, subcategories: Optional[List[str]] = None) -> dict:
    """iter95ar: filtruj BOM po sub_category (kategoriach materialow)."""
    if not subcategories:
        return data
    wanted = {(s or "").lower().strip() for s in subcategories if s}
    if not wanted:
        return data
    rows = [r for r in data.get("rows", []) if (r.get("sub_category") or "").lower().strip() in wanted]
    return {**data, "rows": rows}


def _generate_bom_xlsx_bytes(data: dict):
    """Refaktor iter95ah: generuj XLSX jako bytes - reuzywalne przez endpoint i wysylke maila."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Zestawienie materiałów"
    # iter95av: logo w nagłówku
    _xlsx_add_logo(ws, "A1", width=90, height=90)
    ws["B1"] = f"Zestawienie materiałów: {data['wycena_name']}"
    ws["B1"].font = Font(bold=True, size=14, color="D4AF37")
    ws.merge_cells("B1:E1")
    ws["B2"] = f"Data wygenerowania: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("B2:E2")
    ws.row_dimensions[1].height = 50
    # iter95ar: usunieto kolumne "Termin dostawy" - klient nie chce
    headers = ["L.p.", "Nazwa materiału", "Ilość zużycia", "Jednostka",
               "Opakowanie", "Wielkość opak.", "Liczba opakowań",
               "Cena netto za opak. (PLN)", "Wartość netto (PLN)", "Uwagi"]
    header_fill = PatternFill(start_color="3F5235", end_color="3F5235", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    n_cols = len(headers)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[4].height = 28
    for idx, row in enumerate(data["rows"], start=1):
        r_excel = 4 + idx
        ws.cell(row=r_excel, column=1, value=idx).border = border
        ws.cell(row=r_excel, column=2, value=row["name"]).border = border
        if row.get("qty_in_pkg_unit") is not None:
            ws.cell(row=r_excel, column=3, value=round(row["qty_in_pkg_unit"], 3)).border = border
            ws.cell(row=r_excel, column=4, value=row.get("pkg_unit") or "").border = border
        else:
            ws.cell(row=r_excel, column=3, value=round(row["quantity"], 3)).border = border
            ws.cell(row=r_excel, column=4, value=row["unit"]).border = border
        ws.cell(row=r_excel, column=5, value=row.get("opakowanie") or "—").border = border
        ws.cell(row=r_excel, column=6,
                value=f"{row['pkg_qty']} {row.get('pkg_unit') or ''}" if row.get("pkg_qty") else "—").border = border
        if row.get("num_packages") is not None:
            cell_pkg = ws.cell(row=r_excel, column=7, value=row["num_packages"])
            cell_pkg.font = Font(bold=True, color="D4AF37")
        else:
            ws.cell(row=r_excel, column=7, value="—")
        ws.cell(row=r_excel, column=7).border = border
        # 8=cena, 9=wartosc, 10=uwagi (3 puste do uzupelnienia przez dostawce)
        for col in range(8, n_cols + 1):
            ws.cell(row=r_excel, column=col, value="").border = border
        ws.cell(row=r_excel, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r_excel, column=3).alignment = Alignment(horizontal="right")
        for col in (4, 5, 6, 7):
            ws.cell(row=r_excel, column=col).alignment = Alignment(horizontal="center")
    foot_row = 5 + len(data["rows"]) + 1
    ws.cell(row=foot_row, column=1,
            value="Prosimy o uzupełnienie kolumn: cena netto, wartość netto i uwagi.").font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=n_cols)
    widths = {"A": 14, "B": 40, "C": 12, "D": 11, "E": 14, "F": 14, "G": 12, "H": 18, "I": 16, "J": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf)
    safe_name = (data["wycena_name"] or "wycena").replace("/", "_").replace(" ", "_")[:50]
    filename = f"BOM_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename


@router.get("/wyceny/{wycena_id}/bom.pdf")
async def export_bom_pdf(
    wycena_id: str,
    subcategories: Optional[str] = Query(None),
    _user: dict = Depends(get_current_admin),
):
    data = await _build_bom(wycena_id)
    if subcategories:
        data = _filter_bom_rows(data, [s.strip() for s in subcategories.split(",") if s.strip()])
    content, filename = _generate_bom_pdf_bytes(data)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _generate_bom_pdf_bytes(data: dict):
    """Refaktor iter95ah: generuj PDF jako bytes."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os
    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]
    font_bold_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    ]
    base_font, bold_font = "Helvetica", "Helvetica-Bold"
    for fp in font_paths:
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("DejaVu", fp)); base_font = "DejaVu"; break
            except Exception:
                pass
    for fp in font_bold_paths:
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("DejaVuBold", fp)); bold_font = "DejaVuBold"; break
            except Exception:
                pass
    buf = BytesIO()
    # iter95ar: PDF zapytania ZAWSZE pionowo (A4 portrait), bez kolumny "Termin"
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=12 * mm, leftMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("title", parent=styles["Title"], fontName=bold_font, fontSize=15, textColor=colors.HexColor("#3F5235"))
    sub_st = ParagraphStyle("sub", parent=styles["Normal"], fontName=base_font, fontSize=9, textColor=colors.grey)
    name_st = ParagraphStyle("name", parent=styles["Normal"], fontName=base_font, fontSize=7.5, leading=9)
    # iter95av: styl nagłówka kolumny - umożliwia zawijanie tekstu
    head_st = ParagraphStyle("head", parent=styles["Normal"], fontName=bold_font, fontSize=7.5,
                             leading=9, textColor=colors.white, alignment=1)
    elements = []
    # iter95av: nagłówek z logo + tytuł
    from reportlab.platypus import Image as RLImage
    logo_path = _get_logo_path()
    title_cell = []
    title_cell.append(Paragraph(
        f"<b>Zapytanie ofertowe — Zestawienie materiałów</b><br/>"
        f"<font size=10>{data['wycena_name']}</font><br/>"
        f"<font size=8 color='#666666'>Data: {datetime.now().strftime('%Y-%m-%d %H:%M')}</font>",
        ParagraphStyle("htitle", parent=styles["Normal"], fontName=base_font, fontSize=13,
                       leading=15, textColor=colors.HexColor("#3F5235")),
    ))
    if logo_path:
        try:
            img = RLImage(logo_path, width=22 * mm, height=22 * mm)
            head_tbl = Table([[img, title_cell[0]]], colWidths=[26 * mm, 160 * mm])
        except Exception:
            head_tbl = Table([[title_cell[0]]], colWidths=[186 * mm])
    else:
        head_tbl = Table([[title_cell[0]]], colWidths=[186 * mm])
    head_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(head_tbl)
    elements.append(Spacer(1, 6 * mm))
    table_data = [[
        Paragraph("L.p.", head_st),
        Paragraph("Nazwa materiału", head_st),
        Paragraph("Ilość", head_st),
        Paragraph("Jedn.", head_st),
        Paragraph("Opak.", head_st),
        Paragraph("Wielk.<br/>opak.", head_st),
        Paragraph("Liczba<br/>opak.", head_st),
        Paragraph("Cena netto<br/>za opak.", head_st),
        Paragraph("Uwagi", head_st),
    ]]
    for idx, row in enumerate(data["rows"], start=1):
        if row.get("qty_in_pkg_unit") is not None:
            qty_str = f"{row['qty_in_pkg_unit']:.3f}".replace(".", ",")
            unit_str = row.get("pkg_unit") or ""
        else:
            qty_str = f"{row['quantity']:.3f}".replace(".", ",")
            unit_str = row["unit"]
        pkg_size = f"{row['pkg_qty']} {row.get('pkg_unit') or ''}" if row.get("pkg_qty") else "—"
        num_pkg = str(row["num_packages"]) if row.get("num_packages") is not None else "—"
        table_data.append([
            str(idx), Paragraph(row["name"], name_st), qty_str, unit_str,
            row.get("opakowanie") or "—", pkg_size, num_pkg, "", "",
        ])
    # iter95av: szerokości kolumn zoptymalizowane — nagłówki zawijają w 2 linie aby nie nakładać tekstu
    # Suma = 186mm (A4 portrait - 2*12mm marginesy)
    tbl = Table(table_data, colWidths=[10 * mm, 52 * mm, 14 * mm, 11 * mm, 20 * mm,
                                        20 * mm, 18 * mm, 22 * mm, 19 * mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), base_font, 8),
        ("FONT", (0, 0), (-1, 0), bold_font, 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3F5235")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("ALIGN", (3, 0), (3, -1), "CENTER"),
        ("ALIGN", (4, 0), (6, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
        ("TEXTCOLOR", (6, 1), (6, -1), colors.HexColor("#B8860B")),
        ("FONT", (6, 1), (6, -1), bold_font, 8),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 8 * mm))
    elements.append(Paragraph(
        "Prosimy o uzupełnienie kolumn: <b>cena netto za opakowanie</b> oraz <b>uwagi</b>. "
        "Liczby opakowań zostały zaokrąglone w górę do pełnych jednostek (paleta / wiaderko / rolka).",
        ParagraphStyle("foot", parent=styles["Normal"], fontName=base_font, fontSize=8, textColor=colors.grey)
    ))
    doc.build(elements)
    safe_name = (data["wycena_name"] or "wycena").replace("/", "_").replace(" ", "_")[:50]
    filename = f"BOM_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buf.getvalue(), filename


# ============================================================
# iter95ai: Tabela hurtowni (suppliers) + wysylka BOM emailem
# ============================================================

class SupplierCreate(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None  # iter95aq: numer telefonu
    branze: Optional[str] = None  # branze (np. "izolacje, betony")
    notes: Optional[str] = None


class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    branze: Optional[str] = None
    notes: Optional[str] = None


@router.get("/wyceny/suppliers")
async def list_suppliers(_user: dict = Depends(get_current_admin)):
    rows = await db.wyceny_suppliers.find({}, {"_id": 0}).sort("name", 1).to_list(length=None)
    return {"rows": rows}


@router.post("/wyceny/suppliers")
async def create_supplier(payload: SupplierCreate, _user: dict = Depends(get_current_admin)):
    sid = str(uuid.uuid4())
    doc = {
        "id": sid, "name": payload.name, "email": payload.email,
        "phone": payload.phone or "",
        "branze": payload.branze or "", "notes": payload.notes or "",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.wyceny_suppliers.insert_one(doc)
    return {"id": sid, "ok": True}


@router.patch("/wyceny/suppliers/{sid}")
async def update_supplier(sid: str, payload: SupplierUpdate, _user: dict = Depends(get_current_admin)):
    update = payload.dict(exclude_unset=True)
    if not update:
        return {"ok": True}
    r = await db.wyceny_suppliers.update_one({"id": sid}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(404, "Hurtownia nie istnieje")
    return {"ok": True}


@router.delete("/wyceny/suppliers/{sid}")
async def delete_supplier(sid: str, _user: dict = Depends(get_current_admin)):
    r = await db.wyceny_suppliers.delete_one({"id": sid})
    if r.deleted_count == 0:
        raise HTTPException(404, "Hurtownia nie istnieje")
    return {"ok": True}


class SendBomRequest(BaseModel):
    to_email: EmailStr
    subject: Optional[str] = None
    body: Optional[str] = None
    supplier_id: Optional[str] = None  # zapisz historie
    # iter95ar: filtr kategorii materialow (zapytanie do hurtownika tylko o jego branze)
    subcategories: Optional[List[str]] = None


@router.post("/wyceny/{wycena_id}/bom/send")
async def send_bom_email(wycena_id: str, payload: SendBomRequest, _user: dict = Depends(get_current_admin)):
    """Wyslij zapytanie ofertowe (BOM) na maila hurtownika z zalacznikami PDF + XLSX."""
    import base64
    api_key = os.environ.get("RESEND_API_KEY")
    if not api_key:
        raise HTTPException(503, "Resend nie skonfigurowany (brak RESEND_API_KEY)")
    # Z 4 ustaleń: nadawca = biuro@fegrro.pl
    from_addr = "FeGrro <biuro@fegrro.pl>"
    try:
        import httpx
    except ImportError:
        raise HTTPException(500, "httpx not installed")
    data = await _build_bom(wycena_id)
    # iter95ar: filtr po kategoriach (np. tylko izolacje dla hurtowni izolacji)
    if payload.subcategories:
        data = _filter_bom_rows(data, payload.subcategories)
    if not data.get("rows"):
        raise HTTPException(400, "Wycena nie zawiera materiałów do wysłania")
    xlsx_bytes, xlsx_name = _generate_bom_xlsx_bytes(data)
    pdf_bytes, pdf_name = _generate_bom_pdf_bytes(data)
    # Domyslny szablon
    wycena_name = data["wycena_name"] or "—"
    subject = payload.subject or f"Zapytanie ofertowe — {wycena_name}"
    body_text = payload.body or (
        f"Dzień dobry,\n\n"
        f"W załączeniu przesyłam zestawienie materiałów do wyceny: \u201E{wycena_name}\u201D.\n"
        f"Proszę o przygotowanie oferty cenowej (cena netto za opakowanie, termin dostawy).\n\n"
        f"Termin oferty: 7 dni.\n\n"
        f"Pozdrawiam,\nFeGrro"
    )
    body_html = "<p>" + body_text.replace("\n", "<br>") + "</p>"
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            resp = await client.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={
                    "from": from_addr,
                    "to": [payload.to_email],
                    "reply_to": ["biuro@fegrro.pl"],
                    "subject": subject,
                    "html": body_html,
                    "text": body_text,
                    "attachments": [
                        {"filename": xlsx_name, "content": base64.b64encode(xlsx_bytes).decode("ascii")},
                        {"filename": pdf_name, "content": base64.b64encode(pdf_bytes).decode("ascii")},
                    ],
                },
            )
        if resp.status_code >= 300:
            logger.warning(f"Resend BOM email returned {resp.status_code}: {resp.text}")
            raise HTTPException(502, f"Resend: {resp.status_code} {resp.text}")
        result = resp.json()
    except httpx.HTTPError as e:
        raise HTTPException(502, f"Network error: {e}")
    # Zapisz w historii
    history = {
        "id": str(uuid.uuid4()),
        "wycena_id": wycena_id,
        "to_email": payload.to_email,
        "supplier_id": payload.supplier_id,
        "subject": subject,
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "message_id": result.get("id"),
    }
    await db.wyceny_bom_history.insert_one(history)
    return {"ok": True, "message_id": result.get("id")}


@router.get("/wyceny/{wycena_id}/bom/history")
async def get_bom_history(wycena_id: str, _user: dict = Depends(get_current_admin)):
    rows = await db.wyceny_bom_history.find(
        {"wycena_id": wycena_id}, {"_id": 0}
    ).sort("sent_at", -1).to_list(length=100)
    return {"rows": rows}



# ============================================================
# iter95aj: Eksport pelnej wyceny do PDF/XLSX
# ============================================================

async def _build_wycena_export(wycena_id: str):
    w = await db.wyceny.find_one({"id": wycena_id}, {"_id": 0})
    if not w:
        raise HTTPException(404, "Wycena nie istnieje")
    stages = await db.wyceny_stages.find({"wycena_id": wycena_id}, {"_id": 0}).sort("order", 1).to_list(length=None)
    positions = await db.wyceny_positions.find({"wycena_id": wycena_id}, {"_id": 0}).sort("order", 1).to_list(length=None)
    lines = await db.wyceny_lines.find({"wycena_id": wycena_id}, {"_id": 0}).sort("order", 1).to_list(length=None)
    defaults = {
        "gir": float(w.get("default_gir_pct") or 2.0),
        "dw": float(w.get("default_dw_pct") or 2.0),
        "koszt": float(w.get("default_koszt_pct") or 2.0),
        "narzut": float(w.get("default_narzut_pct") or 0.0),
        "marza": float(w.get("default_marza_pct") or 0.0),
        # iter95bt: rozdzielone narzuty per typ linii
        "narzut_labor": float(w.get("default_narzut_labor_pct") or 0.0),
        "narzut_equipment": float(w.get("default_narzut_equipment_pct") or 0.0),
    }
    enriched_stages = []
    for st in stages:
        pos_list = [p for p in positions if p.get("stage_id") == st["id"]]
        es_positions = []
        for p in pos_list:
            subs = [ln for ln in lines if ln.get("position_id") == p["id"]]
            gir_pct = float(p.get("kaucja_gir_pct") if p.get("kaucja_gir_pct") is not None else defaults["gir"])
            dw_pct = float(p.get("kaucja_dw_pct") if p.get("kaucja_dw_pct") is not None else defaults["dw"])
            koszt_pct = float(p.get("koszt_budowy_pct") if p.get("koszt_budowy_pct") is not None else defaults["koszt"])
            sub_calcs = []
            budzet_zwolniony_pos = 0.0
            for s in subs:
                qty = float(s.get("quantity") or 0)
                cena = float(s.get("unit_price_netto") or 0)
                # iter95bt: dobierz default narzutu wg typu linii
                t_ = s.get("type")
                if t_ == "labor":
                    default_narzut_for_type = defaults["narzut_labor"]
                    default_marza_for_type = 0.0
                elif t_ == "equipment":
                    default_narzut_for_type = defaults["narzut_equipment"]
                    default_marza_for_type = 0.0
                else:
                    # materials lub legacy
                    default_narzut_for_type = defaults["narzut"]
                    default_marza_for_type = defaults["marza"]
                narzut = float(s.get("narzut_zapas_pct") if s.get("narzut_zapas_pct") is not None else default_narzut_for_type)
                # marza_pct tylko dla materials (per-linia override moze ja podac, ale dla labor/equipment ignoruj)
                if t_ in ("materials", None, ""):
                    marza = float(s.get("marza_pct") if s.get("marza_pct") is not None else default_marza_for_type)
                else:
                    marza = 0.0
                zwolniony = qty * cena * (1 + narzut / 100 + marza / 100)
                budzet_zwolniony_pos += zwolniony
                sub_calcs.append({"line": s, "qty": qty, "cena": cena, "narzut": narzut, "marza": marza, "zwolniony": zwolniony})
            kaucja_gir = budzet_zwolniony_pos * gir_pct / 100
            kaucja_dw = budzet_zwolniony_pos * dw_pct / 100
            koszt_budowy = budzet_zwolniony_pos * koszt_pct / 100
            budzet = budzet_zwolniony_pos + kaucja_gir + kaucja_dw + koszt_budowy
            manual_qty = p.get("quantity")
            if manual_qty and float(manual_qty) > 0:
                qty_pos = float(manual_qty)
            else:
                qty_pos = max([sc["qty"] for sc in sub_calcs], default=0)
            cena_pos = budzet / qty_pos if qty_pos > 0 else 0
            # iter95bd: zaokraglenie ceny jednostkowej finalnej (calkowita)
            cena_pos_rounded = round(cena_pos) if cena_pos else 0
            by_type = {"materials": [], "labor": [], "equipment": []}
            for s in subs:
                t = s.get("type")
                nm = (s.get("name") or "").strip()
                if t in by_type and nm:
                    by_type[t].append(nm)
            label_map = {"materials": "Materiały", "labor": "Robocizna", "equipment": "Sprzęt"}
            uwagi_parts = []
            for t in ("materials", "labor", "equipment"):
                if by_type[t]:
                    uwagi_parts.append(f"{label_map[t]}: " + ", ".join(by_type[t]))
            uwagi = " · ".join(uwagi_parts) or "—"
            es_positions.append({
                "position": p, "qty": qty_pos, "cena": cena_pos_rounded,
                "cena_unrounded": cena_pos,
                "kaucja_gir": kaucja_gir, "kaucja_dw": kaucja_dw, "koszt_budowy": koszt_budowy,
                "budzet_zwolniony": budzet_zwolniony_pos, "budzet": budzet,
                "uwagi": uwagi, "subs_calc": sub_calcs,
            })
        enriched_stages.append({"stage": st, "positions": es_positions})
    return {"wycena": w, "stages": enriched_stages, "defaults": defaults}


def _generate_wycena_xlsx_bytes(data: dict, detail: str = "positions"):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Wycena"
    wycena_name = data["wycena"].get("name", "")
    # iter95av: logo w nagłówku
    _xlsx_add_logo(ws, "A1", width=90, height=90)
    ws["B1"] = f"Wycena: {wycena_name}"
    ws["B1"].font = Font(bold=True, size=14, color="3F5235")
    ws.merge_cells("B1:M1")
    ws["B2"] = f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M')} · Tryb: " + (
        "Pozycje główne" if detail == "positions" else "Pełna (z podpozycjami)")
    ws["B2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("B2:M2")
    ws.row_dimensions[1].height = 50
    headers = ["Kod", "Nazwa", "Ilość", "Jedn.", "Cena", "Narzut %", "Marża %",
               "Kaucja GIR", "Kaucja DW", "Koszt budowy", "Budżet zwolniony", "Budżet", "Uwagi"]
    header_fill = PatternFill(start_color="3F5235", end_color="3F5235", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[4].height = 28
    r = 5
    total_budzet = 0.0
    for st_idx, st_data in enumerate(data["stages"], start=1):
        st = st_data["stage"]
        st_cell = ws.cell(row=r, column=1, value=f"ETAP {st_idx}: {st.get('name', '')}")
        st_cell.fill = PatternFill(start_color="C8E4B5", end_color="C8E4B5", fill_type="solid")
        st_cell.font = Font(bold=True, color="3F5235")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        r += 1
        for p_idx, pe in enumerate(st_data["positions"], start=1):
            p = pe["position"]
            code = f"{st_idx}.{p_idx}"
            ws.cell(row=r, column=1, value=code)
            ws.cell(row=r, column=2, value=p.get("name", ""))
            ws.cell(row=r, column=3, value=round(pe["qty"], 3))
            ws.cell(row=r, column=4, value=p.get("unit") or "")
            ws.cell(row=r, column=5, value=round(pe["cena"], 2))
            ws.cell(row=r, column=6, value="—")
            ws.cell(row=r, column=7, value="—")
            ws.cell(row=r, column=8, value=round(pe["kaucja_gir"], 2))
            ws.cell(row=r, column=9, value=round(pe["kaucja_dw"], 2))
            ws.cell(row=r, column=10, value=round(pe["koszt_budowy"], 2))
            ws.cell(row=r, column=11, value=round(pe["budzet_zwolniony"], 2))
            ws.cell(row=r, column=12, value=round(pe["budzet"], 2))
            ws.cell(row=r, column=13, value=pe["uwagi"])
            for c in range(1, 14):
                ws.cell(row=r, column=c).border = border
                ws.cell(row=r, column=c).font = Font(bold=True)
            ws.cell(row=r, column=12).fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
            ws.cell(row=r, column=13).alignment = Alignment(wrap_text=True, vertical="top")
            total_budzet += pe["budzet"]
            r += 1
            if detail == "full":
                for sub_idx, sc in enumerate(pe["subs_calc"], start=1):
                    s = sc["line"]
                    sub_code = f"{code}.{sub_idx}"
                    ratio = sc["zwolniony"] / pe["budzet_zwolniony"] if pe["budzet_zwolniony"] > 0 else 0
                    sub_budzet = pe["budzet"] * ratio
                    type_label = {"materials": "Materiał", "labor": "Robocizna", "equipment": "Sprzęt"}.get(s.get("type"), "")
                    ws.cell(row=r, column=1, value=sub_code)
                    ws.cell(row=r, column=2, value=f"  ↳ {s.get('name', '')}")
                    ws.cell(row=r, column=3, value=round(sc["qty"], 3))
                    ws.cell(row=r, column=4, value=s.get("unit") or "")
                    ws.cell(row=r, column=5, value=round(sc["cena"], 2))
                    ws.cell(row=r, column=6, value=round(sc["narzut"], 1) if sc["narzut"] else "")
                    ws.cell(row=r, column=7, value=round(sc["marza"], 1) if sc["marza"] else "")
                    ws.cell(row=r, column=8, value=round(pe["kaucja_gir"] * ratio, 2))
                    ws.cell(row=r, column=9, value=round(pe["kaucja_dw"] * ratio, 2))
                    ws.cell(row=r, column=10, value=round(pe["koszt_budowy"] * ratio, 2))
                    ws.cell(row=r, column=11, value=round(sc["zwolniony"], 2))
                    ws.cell(row=r, column=12, value=round(sub_budzet, 2))
                    ws.cell(row=r, column=13, value=type_label)
                    for c in range(1, 14):
                        ws.cell(row=r, column=c).border = border
                        ws.cell(row=r, column=c).font = Font(italic=True, color="555555")
                    r += 1
    r += 1
    # iter95bg: scal kolumny 1-11 (A-K) dla "RAZEM:" zeby tekst sie ladnie zmiescil
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    razem_cell_full = ws.cell(row=r, column=1, value="RAZEM:")
    razem_cell_full.font = Font(bold=True, size=12)
    razem_cell_full.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=r, column=12, value=round(total_budzet, 2)).font = Font(bold=True, size=12, color="B8860B")
    widths = {"A": 12, "B": 38, "C": 10, "D": 8, "E": 12, "F": 9, "G": 9,
              "H": 13, "I": 13, "J": 14, "K": 16, "L": 15, "M": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf)
    safe_name = (wycena_name or "wycena").replace("/", "_").replace(" ", "_")[:50]
    filename = f"Wycena_{safe_name}_{'pelna' if detail == 'full' else 'pozycje'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename


def _generate_wycena_pdf_bytes(data: dict, detail: str = "positions"):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os
    font_paths = ["/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"]
    font_bold_paths = ["/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"]
    base_font, bold_font = "Helvetica", "Helvetica-Bold"
    for fp in font_paths:
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("DejaVu", fp)); base_font = "DejaVu"; break
            except Exception: pass
    for fp in font_bold_paths:
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("DejaVuBold", fp)); bold_font = "DejaVuBold"; break
            except Exception: pass
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=8 * mm, leftMargin=8 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("title", parent=styles["Title"], fontName=bold_font, fontSize=14, textColor=colors.HexColor("#3F5235"))
    sub_st = ParagraphStyle("sub", parent=styles["Normal"], fontName=base_font, fontSize=9, textColor=colors.grey)
    cell_st = ParagraphStyle("cell", parent=styles["Normal"], fontName=base_font, fontSize=7, leading=8)
    cell_b = ParagraphStyle("cellb", parent=styles["Normal"], fontName=bold_font, fontSize=7, leading=8)
    elements = []
    wycena_name = data["wycena"].get("name", "")
    # iter95av: nagłówek z logo + tytuł
    from reportlab.platypus import Image as RLImage
    logo_path = _get_logo_path()
    title_para = Paragraph(
        f"<b>Wycena: {wycena_name}</b><br/>"
        f"<font size=9 color='#666666'>Data: {datetime.now().strftime('%Y-%m-%d %H:%M')} · Tryb: "
        + ("Pozycje główne" if detail == "positions" else "Pełna (z podpozycjami)") + "</font>",
        ParagraphStyle("ht", parent=styles["Normal"], fontName=base_font, fontSize=13,
                       leading=15, textColor=colors.HexColor("#3F5235")),
    )
    if logo_path:
        try:
            img = RLImage(logo_path, width=20 * mm, height=20 * mm)
            head_tbl = Table([[img, title_para]], colWidths=[24 * mm, 257 * mm])
        except Exception:
            head_tbl = Table([[title_para]], colWidths=[281 * mm])
    else:
        head_tbl = Table([[title_para]], colWidths=[281 * mm])
    head_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(head_tbl)
    elements.append(Spacer(1, 4 * mm))
    head_cell = ParagraphStyle("headcell", parent=styles["Normal"], fontName=bold_font, fontSize=7,
                               leading=8, textColor=colors.white, alignment=1)
    headers = [
        Paragraph("Kod", head_cell), Paragraph("Nazwa", head_cell),
        Paragraph("Ilość", head_cell), Paragraph("Jedn.", head_cell),
        Paragraph("Cena", head_cell),
        Paragraph("Narzut<br/>%", head_cell), Paragraph("Marża<br/>%", head_cell),
        Paragraph("Kaucja<br/>GIR", head_cell), Paragraph("Kaucja<br/>DW", head_cell),
        Paragraph("Koszt<br/>budowy", head_cell),
        Paragraph("Budżet<br/>zwolniony", head_cell), Paragraph("Budżet", head_cell),
        Paragraph("Uwagi", head_cell),
    ]
    table_data = [headers]
    total_budzet = 0.0
    row_styles = []
    for st_idx, st_data in enumerate(data["stages"], start=1):
        st = st_data["stage"]
        row_styles.append((len(table_data), 'stage'))
        table_data.append([f"ETAP {st_idx}: {st.get('name', '')}", "", "", "", "", "", "", "", "", "", "", "", ""])
        for p_idx, pe in enumerate(st_data["positions"], start=1):
            p = pe["position"]
            code = f"{st_idx}.{p_idx}"
            row_styles.append((len(table_data), 'pos'))
            table_data.append([
                code, Paragraph(p.get("name", ""), cell_b),
                f"{pe['qty']:.2f}".replace(".", ","), p.get("unit") or "",
                f"{pe['cena']:.2f}".replace(".", ","),
                "—", "—",
                f"{pe['kaucja_gir']:.2f}".replace(".", ","),
                f"{pe['kaucja_dw']:.2f}".replace(".", ","),
                f"{pe['koszt_budowy']:.2f}".replace(".", ","),
                f"{pe['budzet_zwolniony']:.2f}".replace(".", ","),
                f"{pe['budzet']:.2f}".replace(".", ","),
                Paragraph(pe["uwagi"], cell_st),
            ])
            total_budzet += pe["budzet"]
            if detail == "full":
                for sub_idx, sc in enumerate(pe["subs_calc"], start=1):
                    s = sc["line"]
                    sub_code = f"{code}.{sub_idx}"
                    ratio = sc["zwolniony"] / pe["budzet_zwolniony"] if pe["budzet_zwolniony"] > 0 else 0
                    sub_budzet = pe["budzet"] * ratio
                    type_label = {"materials": "Materiał", "labor": "Robocizna", "equipment": "Sprzęt"}.get(s.get("type"), "")
                    row_styles.append((len(table_data), 'sub'))
                    table_data.append([
                        sub_code, Paragraph(f"\u21B3 {s.get('name', '')}", cell_st),
                        f"{sc['qty']:.2f}".replace(".", ","), s.get("unit") or "",
                        f"{sc['cena']:.2f}".replace(".", ","),
                        f"{sc['narzut']:.1f}".replace(".", ",") if sc["narzut"] else "—",
                        f"{sc['marza']:.1f}".replace(".", ",") if sc["marza"] else "—",
                        f"{pe['kaucja_gir'] * ratio:.2f}".replace(".", ","),
                        f"{pe['kaucja_dw'] * ratio:.2f}".replace(".", ","),
                        f"{pe['koszt_budowy'] * ratio:.2f}".replace(".", ","),
                        f"{sc['zwolniony']:.2f}".replace(".", ","),
                        f"{sub_budzet:.2f}".replace(".", ","),
                        type_label,
                    ])
    table_data.append(["", "", "", "", "", "", "", "", "", "", "RAZEM:",
                       f"{total_budzet:.2f}".replace(".", ",") + " zł", ""])
    tbl_styles = [
        ("FONT", (0, 0), (-1, -1), base_font, 7),
        ("FONT", (0, 0), (-1, 0), bold_font, 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3F5235")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 1), (11, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 1), (3, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#AAAAAA")),
        ("FONT", (10, -1), (-1, -1), bold_font, 9),
        ("TEXTCOLOR", (11, -1), (11, -1), colors.HexColor("#B8860B")),
        # iter95av: padding + wyrównanie nagłówków
        ("TOPPADDING", (0, 0), (-1, 0), 5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    for (idx, kind) in row_styles:
        if kind == 'stage':
            tbl_styles.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#C8E4B5")))
            tbl_styles.append(("FONT", (0, idx), (-1, idx), bold_font, 8))
            tbl_styles.append(("SPAN", (0, idx), (-1, idx)))
        elif kind == 'pos':
            tbl_styles.append(("BACKGROUND", (11, idx), (11, idx), colors.HexColor("#FFF8DC")))
            tbl_styles.append(("FONT", (11, idx), (11, idx), bold_font, 7))
    # iter95bg: scal kolumny 0-10 dla "RAZEM:" zeby tekst sie zmiescil w landscape PDF
    last_row_idx = len(table_data) - 1
    tbl_styles.append(("SPAN", (0, last_row_idx), (10, last_row_idx)))
    tbl_styles.append(("ALIGN", (0, last_row_idx), (10, last_row_idx), "RIGHT"))
    tbl_styles.append(("RIGHTPADDING", (10, last_row_idx), (10, last_row_idx), 6))
    tbl_styles.append(("BACKGROUND", (0, last_row_idx), (11, last_row_idx), colors.HexColor("#FFF8DC")))
    # iter95av: szerokości zoptymalizowane - więcej miejsca dla kolumn pieniężnych, mniej dla Uwagi
    # Suma: 12+46+12+10+15+13+13+17+17+19+22+22+50 = 268mm (landscape A4 281mm użyt.)
    tbl = Table(table_data, colWidths=[12 * mm, 46 * mm, 12 * mm, 10 * mm, 15 * mm,
                                         13 * mm, 13 * mm, 17 * mm, 17 * mm, 19 * mm,
                                         22 * mm, 22 * mm, 50 * mm], repeatRows=1)
    tbl.setStyle(TableStyle(tbl_styles))
    elements.append(tbl)
    doc.build(elements)
    safe_name = (wycena_name or "wycena").replace("/", "_").replace(" ", "_")[:50]
    filename = f"Wycena_{safe_name}_{'pelna' if detail == 'full' else 'pozycje'}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buf.getvalue(), filename


@router.get("/wyceny/{wycena_id}/export.xlsx")
async def export_wycena_xlsx(
    wycena_id: str,
    detail: str = Query("positions", pattern="^(positions|full|client)$"),
    include_surface: bool = Query(True),
    include_wskazniki: bool = Query(True),
    include_notes: bool = Query(True),
    _user: dict = Depends(get_current_admin_export),
):
    data = await _build_wycena_export(wycena_id)
    if detail == "client":
        opts = {
            "include_surface": include_surface,
            "include_wskazniki": include_wskazniki,
            "include_notes": include_notes,
        }
        content, filename = _generate_wycena_client_xlsx_bytes(data, opts)
    else:
        content, filename = _generate_wycena_xlsx_bytes(data, detail)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# iter95ak/95ap: PDF dla klienta - tylko nazwa, ilosc, cena, wartosc + opcjonalne sekcje
# iter95bk: jedyny szablon "premium" - granat #152033 + akcent zielony FeGrro #9DBC85
# (poprzednio bylo 3 szablony: classic/branded/premium; uzytkownik wybral premium z zielonym)
_TEMPLATE_CONFIGS = {
    "premium": {
        "primary": "#152033",       # granat (header tabeli, ramki)
        "primary_text": "#152033",  # tytul, etapy
        "accent": "#9DBC85",        # zielony FeGrro (akcent — pasek pod headerem, tagline)
        "header_bg_alt": "#F1F4F9", # cool gray zebra
        "logo_mm": 42,
        "tagline": "PROFESJONALNE USŁUGI BUDOWLANE",
        "show_gold_bar": True,      # zielony pasek pod headerem
        "total_bg": "#152033",      # ciemne tlo total
        "total_text": "#C8E4B5",    # jasnozielony tekst total (kontrastowy na granacie)
    },
}


def _generate_wycena_client_pdf_bytes(data: dict, opts: Optional[dict] = None, template_style: str = "premium"):
    opts = opts or {}
    cfg = _TEMPLATE_CONFIGS.get(template_style) or _TEMPLATE_CONFIGS["premium"]
    include_surface = bool(opts.get("include_surface", True))
    include_wskazniki = bool(opts.get("include_wskazniki", True))
    include_notes = bool(opts.get("include_notes", True))
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os
    font_paths = ["/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"]
    font_bold_paths = ["/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"]
    base_font, bold_font = "Helvetica", "Helvetica-Bold"
    for fp in font_paths:
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("DejaVu", fp)); base_font = "DejaVu"; break
            except Exception: pass
    for fp in font_bold_paths:
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("DejaVuBold", fp)); bold_font = "DejaVuBold"; break
            except Exception: pass

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("title", parent=styles["Title"], fontName=bold_font, fontSize=18,
                              textColor=colors.HexColor(cfg["primary_text"]), alignment=0, spaceAfter=2)
    company_st = ParagraphStyle("company", parent=styles["Normal"], fontName=bold_font, fontSize=11,
                                textColor=colors.HexColor(cfg["primary_text"]), alignment=2)
    sub_st = ParagraphStyle("sub", parent=styles["Normal"], fontName=base_font, fontSize=9,
                            textColor=colors.grey)
    name_st = ParagraphStyle("name", parent=styles["Normal"], fontName=base_font, fontSize=9, leading=11)
    stage_st = ParagraphStyle("stage", parent=styles["Normal"], fontName=bold_font, fontSize=10,
                              textColor=colors.HexColor(cfg["primary_text"]))

    elements = []
    wycena_name = data["wycena"].get("name", "")

    # iter95w: Naglowek z wiekszym logo + NIP + telefon
    # iter95bh: rozmiar logo i tagline z configu szablonu
    logo_mm_val = cfg["logo_mm"]
    logo_paths = [
        "/app/frontend/public/icon-192x192.png",
        "/app/frontend/public/apple-touch-icon.png",
    ]
    logo_path = next((p for p in logo_paths if _os.path.exists(p)), None)
    header_cells = []
    if logo_path:
        try:
            img = Image(logo_path, width=logo_mm_val * mm, height=logo_mm_val * mm)
            header_cells.append(img)
        except Exception:
            header_cells.append("")
    else:
        header_cells.append("")
    tagline_html = ""
    if cfg["tagline"]:
        tagline_html = f"<br/><font size=7 color='{cfg['accent']}'><b>{cfg['tagline']}</b></font>"
    header_cells.append(Paragraph(
        "FeGrro"
        f"{tagline_html}<br/>"
        "<font size=8 color='#444444'>NIP: 589-206-61-74</font><br/>"
        "<font size=8 color='#444444'>Tel: 885 213 273</font><br/>"
        "<font size=8 color='#666666'>biuro@fegrro.pl</font>",
        company_st,
    ))
    header_tbl = Table([header_cells], colWidths=[(logo_mm_val + 8) * mm, (180 - logo_mm_val - 8) * mm])
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    elements.append(header_tbl)
    # iter95bh: zloty pasek pod naglowkiem (branded / premium)
    if cfg["show_gold_bar"]:
        bar = Table([[""]], colWidths=[180 * mm], rowHeights=[1.4 * mm])
        bar.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(cfg["accent"])),
            ("LINEABOVE", (0, 0), (-1, 0), 0, colors.white),
        ]))
        elements.append(Spacer(1, 2 * mm))
        elements.append(bar)
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph(f"Oferta: {wycena_name}", title_st))
    elements.append(Paragraph(
        f"Data wystawienia: {datetime.now().strftime('%Y-%m-%d')}",
        sub_st,
    ))
    elements.append(Spacer(1, 6 * mm))

    # iter95al: blok adresata (dane klienta) jezeli wypelnione
    w_doc = data["wycena"]
    client_name = (w_doc.get("client_name") or "").strip()
    client_nip = (w_doc.get("client_nip") or "").strip()
    client_address = (w_doc.get("client_address") or "").strip()
    if client_name or client_nip or client_address:
        addr_label = ParagraphStyle("addrlabel", parent=styles["Normal"], fontName=base_font, fontSize=7,
                                    textColor=colors.HexColor("#94A3B8"), spaceAfter=2)
        addr_body = ParagraphStyle("addrbody", parent=styles["Normal"], fontName=bold_font, fontSize=10,
                                   textColor=colors.HexColor(cfg["primary_text"]), leading=13)
        addr_body_sub = ParagraphStyle("addrbodysub", parent=styles["Normal"], fontName=base_font, fontSize=9,
                                       textColor=colors.HexColor("#444444"), leading=12)
        addr_inner = [Paragraph("ADRESAT", addr_label)]
        if client_name:
            addr_inner.append(Paragraph(client_name.replace("\n", "<br/>"), addr_body))
        if client_nip:
            addr_inner.append(Paragraph(f"NIP: {client_nip}", addr_body_sub))
        if client_address:
            addr_inner.append(Paragraph(client_address.replace("\n", "<br/>"), addr_body_sub))
        addr_box = Table([[addr_inner]], colWidths=[85 * mm])
        addr_box.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor(cfg["primary"])),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(cfg["header_bg_alt"])),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(addr_box)
        elements.append(Spacer(1, 6 * mm))

    # iter95ap: sekcja Powierzchnie (PC, PC↓, PC↑, PUM) - jezeli wybrana i sa dane
    w_full = data["wycena"]
    pc = w_full.get("pc_m2")
    pc_pod = w_full.get("pc_podziemie_m2")
    pc_nad = w_full.get("pc_nadziemie_m2")
    pum = w_full.get("pum_m2")
    # iter95bj: scalono "Powierzchnie budynku" + "Wskazniki kosztowe" w jedna tabele
    # na DOLE oferty (po total pozycji). Zostala tylko 1 tabela: [Powierzchnia | m² | zł/m²]

    # Tabela: L.p. | Nazwa | Ilosc | Jedn. | Cena netto | Wartosc netto
    headers = ["L.p.", "Nazwa pozycji", "Ilość", "Jedn.", "Cena netto", "Wartość netto"]
    table_data = [headers]
    total_netto = 0.0
    row_styles = []
    lp_counter = 0
    for st_idx, st_data in enumerate(data["stages"], start=1):
        st = st_data["stage"]
        if not st_data["positions"]:
            continue
        row_styles.append((len(table_data), 'stage'))
        table_data.append([Paragraph(f"<b>Etap {st_idx}: {st.get('name', '')}</b>", stage_st),
                           "", "", "", "", ""])
        for pe in st_data["positions"]:
            p = pe["position"]
            lp_counter += 1
            qty = pe["qty"]
            # iter95ak: dla klienta uzywamy budzet (cena koncowa zawiera marze + kaucje)
            wartosc = pe["budzet"]
            cena = wartosc / qty if qty > 0 else 0
            row_styles.append((len(table_data), 'pos'))
            table_data.append([
                str(lp_counter),
                Paragraph(p.get("name", ""), name_st),
                f"{qty:.2f}".replace(".", ","),
                p.get("unit") or "",
                f"{cena:,.2f}".replace(",", " ").replace(".", ",") + " zł",
                f"{wartosc:,.2f}".replace(",", " ").replace(".", ",") + " zł",
            ])
            total_netto += wartosc

    # wiersz sumy
    row_styles.append((len(table_data), 'total'))
    table_data.append(["", "", "", "", "RAZEM netto:",
                       f"{total_netto:,.2f}".replace(",", " ").replace(".", ",") + " zł"])

    tbl_styles = [
        ("FONT", (0, 0), (-1, -1), base_font, 9),
        ("FONT", (0, 0), (-1, 0), bold_font, 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(cfg["primary"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ("ALIGN", (3, 1), (3, -1), "CENTER"),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor(cfg["header_bg_alt"])]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    # iter95bh: dla 'stage' wiersza kolor tla zalezny od configu (lekko jasniejszy niz tlo zebry)
    stage_bg_hex = cfg["header_bg_alt"]
    for (idx, kind) in row_styles:
        if kind == 'stage':
            tbl_styles.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor(stage_bg_hex)))
            tbl_styles.append(("SPAN", (0, idx), (-1, idx)))
            tbl_styles.append(("LEFTPADDING", (0, idx), (0, idx), 6))
        elif kind == 'total':
            tbl_styles.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor(cfg["total_bg"])))
            # iter95bg: scal kolumny 0-4 dla "RAZEM netto:" zeby dlugi tekst sie zmiescil
            tbl_styles.append(("SPAN", (0, idx), (4, idx)))
            tbl_styles.append(("ALIGN", (0, idx), (4, idx), "RIGHT"))
            tbl_styles.append(("FONT", (0, idx), (-1, idx), bold_font, 11))
            tbl_styles.append(("TEXTCOLOR", (0, idx), (-1, idx), colors.HexColor(cfg["total_text"])))
            tbl_styles.append(("RIGHTPADDING", (4, idx), (4, idx), 6))
            tbl_styles.append(("TOPPADDING", (0, idx), (-1, idx), 8))
            tbl_styles.append(("BOTTOMPADDING", (0, idx), (-1, idx), 8))

    tbl = Table(
        table_data,
        colWidths=[10 * mm, 102 * mm, 16 * mm, 12 * mm, 25 * mm, 25 * mm],
        repeatRows=1,
    )
    tbl.setStyle(TableStyle(tbl_styles))
    elements.append(tbl)

    # iter95bj: scalona tabela "Powierzchnie i wskazniki kosztowe" na DOLE
    # 3 kolumny: [Powierzchnia | m² | zł/m²] - bez duplikacji
    any_surface = any(v not in (None, "", 0) for v in (pc, pc_pod, pc_nad, pum))
    show_surface_block = (include_surface or include_wskazniki) and any_surface
    if show_surface_block:
        def _fmt_num(v, suffix):
            return f"{float(v):,.2f}".replace(",", " ").replace(".", ",") + f" {suffix}"

        def _zlm2(area):
            if not area or float(area) <= 0:
                return "—"
            return _fmt_num(total_netto / float(area), "zł/m²")

        surf_rows = [["Powierzchnia", "m²", "zł/m²"]]
        if pc not in (None, "", 0):
            surf_rows.append(["PC — Powierzchnia całkowita", _fmt_num(pc, "m²"), _zlm2(pc)])
        if pc_pod not in (None, "", 0):
            surf_rows.append(["  ↓ w tym podziemie", _fmt_num(pc_pod, "m²"), _zlm2(pc_pod)])
        if pc_nad not in (None, "", 0):
            surf_rows.append(["  ↑ w tym nadziemie", _fmt_num(pc_nad, "m²"), _zlm2(pc_nad)])
        if pum not in (None, "", 0):
            surf_rows.append(["PUM — Pow. użytkowa mieszkalna", _fmt_num(pum, "m²"), _zlm2(pum)])

        if len(surf_rows) > 1:
            elements.append(Spacer(1, 5 * mm))
            block_label_st = ParagraphStyle("blocklabel", parent=styles["Normal"], fontName=bold_font,
                                            fontSize=9, textColor=colors.HexColor(cfg["primary_text"]), spaceAfter=2)
            elements.append(Paragraph("Powierzchnie i wskaźniki kosztowe", block_label_st))
            surf_tbl = Table(surf_rows, colWidths=[100 * mm, 40 * mm, 40 * mm])
            surf_tbl.setStyle(TableStyle([
                ("FONT", (0, 0), (-1, -1), base_font, 9),
                ("FONT", (0, 0), (-1, 0), bold_font, 9),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(cfg["primary"])),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("FONT", (2, 1), (2, -1), bold_font, 9),
                ("TEXTCOLOR", (2, 1), (2, -1), colors.HexColor(cfg["primary_text"])),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(cfg["header_bg_alt"])]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(surf_tbl)

    # iter95w: Sekcje "Oferta obejmuje" / "Oferta nie obejmuje"
    w_doc2 = data["wycena"]
    scope_inc = (w_doc2.get("scope_includes") or "").strip()
    scope_exc = (w_doc2.get("scope_excludes") or "").strip()
    if scope_inc or scope_exc:
        elements.append(Spacer(1, 6 * mm))
        scope_label_st = ParagraphStyle("scopelabel", parent=styles["Normal"], fontName=bold_font,
                                        fontSize=10, textColor=colors.HexColor(cfg["primary_text"]),
                                        spaceAfter=3)
        scope_inc_st = ParagraphStyle("scopeinc", parent=styles["Normal"], fontName=base_font,
                                       fontSize=9, textColor=colors.HexColor("#222222"),
                                       leading=12, leftIndent=4)
        scope_exc_st = ParagraphStyle("scopeexc", parent=styles["Normal"], fontName=base_font,
                                       fontSize=9, textColor=colors.HexColor("#7A2E2E"),
                                       leading=12, leftIndent=4)

        def _scope_paragraph(text: str, style: ParagraphStyle):
            lines = [ln.strip() for ln in text.replace("\r", "").split("\n") if ln.strip()]
            if not lines:
                return None
            html = "<br/>".join(
                (ln if ln.startswith(("&bull;", "\u2022", "-", "*")) else f"&bull; {ln}")
                for ln in lines
            )
            return Paragraph(html, style)

        if scope_inc:
            elements.append(Paragraph("Oferta obejmuje", scope_label_st))
            p = _scope_paragraph(scope_inc, scope_inc_st)
            if p is not None:
                box = Table([[p]], colWidths=[180 * mm])
                box.setStyle(TableStyle([
                    ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#5F7552")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F7EC")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]))
                elements.append(box)
                elements.append(Spacer(1, 3 * mm))

        if scope_exc:
            elements.append(Paragraph("Oferta nie obejmuje", scope_label_st))
            p = _scope_paragraph(scope_exc, scope_exc_st)
            if p is not None:
                box = Table([[p]], colWidths=[180 * mm])
                box.setStyle(TableStyle([
                    ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#9B2C2C")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FDF2F2")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]))
                elements.append(box)

    if include_notes:
        elements.append(Spacer(1, 8 * mm))
        notice_st = ParagraphStyle(
            "notice", parent=styles["Normal"], fontName=base_font, fontSize=8,
            textColor=colors.grey, leading=10,
        )
        notice = data["wycena"].get("notes") or (
            "Oferta ważna 30 dni od daty wystawienia. "
            "Podane ceny są cenami netto. Płatność wg ustaleń umowy. "
            "Zakres prac i warunki realizacji do uzgodnienia."
        )
        elements.append(Paragraph(f"<b>Uwagi:</b> {notice}", notice_st))

    doc.build(elements)
    safe_name = (wycena_name or "wycena").replace("/", "_").replace(" ", "_")[:50]
    filename = f"Oferta_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buf.getvalue(), filename


@router.get("/wyceny/{wycena_id}/export.pdf")
async def export_wycena_pdf(
    wycena_id: str,
    detail: str = Query("positions", pattern="^(positions|full|client)$"),
    inline: bool = Query(False),
    include_surface: bool = Query(True),
    include_wskazniki: bool = Query(True),
    include_notes: bool = Query(True),
    _user: dict = Depends(get_current_admin_export),
):
    data = await _build_wycena_export(wycena_id)
    if detail == "client":
        opts = {
            "include_surface": include_surface,
            "include_wskazniki": include_wskazniki,
            "include_notes": include_notes,
        }
        # iter95bk: zawsze szablon "premium" (granat + zielony akcent)
        content, filename = _generate_wycena_client_pdf_bytes(data, opts, template_style="premium")
    else:
        content, filename = _generate_wycena_pdf_bytes(data, detail)
    disposition = "inline" if inline else "attachment"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": f'{disposition}; filename="{filename}"'},
    )


# iter95ap: aktywny Excel dla klienta z formulami (ilosc*cena, SUM) - inwestor moze podmieniac wartosci
def _generate_wycena_client_xlsx_bytes(data: dict, opts: Optional[dict] = None):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    opts = opts or {}
    include_surface = bool(opts.get("include_surface", True))
    include_wskazniki = bool(opts.get("include_wskazniki", True))
    include_notes = bool(opts.get("include_notes", True))

    wb = Workbook()
    ws = wb.active
    ws.title = "Oferta"
    wycena_name = data["wycena"].get("name", "")
    w_full = data["wycena"]

    # iter95w: logo + naglowek firmowy z NIP/telefonem
    _xlsx_add_logo(ws, "A1", width=110, height=110)
    ws["B1"] = "FeGrro"
    ws["B1"].font = Font(bold=True, size=14, color="3F5235")
    ws["B2"] = "NIP: 589-206-61-74"
    ws["B2"].font = Font(size=10, color="444444")
    ws["B3"] = "Tel: 885 213 273"
    ws["B3"].font = Font(size=10, color="444444")
    ws["B4"] = "biuro@fegrro.pl"
    ws["B4"].font = Font(italic=True, size=9, color="666666")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    ws["A6"] = f"Oferta: {wycena_name}"
    ws["A6"].font = Font(bold=True, size=14, color="3F5235")
    ws.merge_cells("A6:F6")
    ws["A7"] = f"Data wystawienia: {datetime.now().strftime('%Y-%m-%d')}"
    ws["A7"].font = Font(italic=True, size=9, color="888888")
    ws.merge_cells("A7:F7")

    r = 9
    # adresat
    client_name = (w_full.get("client_name") or "").strip()
    client_nip = (w_full.get("client_nip") or "").strip()
    client_address = (w_full.get("client_address") or "").strip()
    if client_name or client_nip or client_address:
        ws.cell(row=r, column=1, value="ADRESAT").font = Font(bold=True, size=9, color="3F5235")
        r += 1
        if client_name:
            ws.cell(row=r, column=1, value=client_name).font = Font(bold=True, size=10, color="3F5235")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
        if client_nip:
            ws.cell(row=r, column=1, value=f"NIP: {client_nip}").font = Font(size=9, color="444444"); r += 1
        if client_address:
            for line in client_address.split("\n"):
                ws.cell(row=r, column=1, value=line).font = Font(size=9, color="444444"); r += 1
        r += 1

    # powierzchnie
    pc = w_full.get("pc_m2")
    pc_pod = w_full.get("pc_podziemie_m2")
    pc_nad = w_full.get("pc_nadziemie_m2")
    pum = w_full.get("pum_m2")
    surf_items = []
    if pc not in (None, "", 0): surf_items.append(("Powierzchnia całkowita (PC)", float(pc)))
    if pc_pod not in (None, "", 0): surf_items.append(("  ↓ w tym podziemie", float(pc_pod)))
    if pc_nad not in (None, "", 0): surf_items.append(("  ↑ w tym nadziemie", float(pc_nad)))
    if pum not in (None, "", 0): surf_items.append(("Powierzchnia użytkowa mieszkalna (PUM)", float(pum)))
    surface_row_map = {}  # nazwa -> wiersz dla wskaznikow
    if include_surface and surf_items:
        ws.cell(row=r, column=1, value="Powierzchnie budynku").font = Font(bold=True, size=10, color="3F5235")
        r += 1
        header_fill = PatternFill(start_color="3F5235", end_color="3F5235", fill_type="solid")
        for col, h in enumerate(["Powierzchnia", "Wartość (m²)"], start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
        r += 1
        for lbl, val in surf_items:
            ws.cell(row=r, column=1, value=lbl).font = Font(size=10)
            ws.cell(row=r, column=2, value=val).font = Font(size=10)
            ws.cell(row=r, column=2).number_format = '#,##0.00" m²"'
            surface_row_map[lbl] = r
            r += 1
        r += 1

    # tabela pozycji z formulami
    ws.cell(row=r, column=1, value="L.p.")
    ws.cell(row=r, column=2, value="Nazwa pozycji")
    ws.cell(row=r, column=3, value="Ilość")
    ws.cell(row=r, column=4, value="Jedn.")
    ws.cell(row=r, column=5, value="Cena netto")
    ws.cell(row=r, column=6, value="Wartość netto")
    header_row = r
    header_fill = PatternFill(start_color="3F5235", end_color="3F5235", fill_type="solid")
    for col in range(1, 7):
        c = ws.cell(row=header_row, column=col)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    stage_fill = PatternFill(start_color="E8F0E0", end_color="E8F0E0", fill_type="solid")

    lp = 0
    first_pos_row = None
    last_pos_row = None
    for st_idx, st_data in enumerate(data["stages"], start=1):
        st = st_data["stage"]
        if not st_data["positions"]:
            continue
        st_cell = ws.cell(row=r, column=1, value=f"Etap {st_idx}: {st.get('name', '')}")
        st_cell.font = Font(bold=True, color="3F5235"); st_cell.fill = stage_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        # tlo na zlaczonych
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = stage_fill
        r += 1
        for pe in st_data["positions"]:
            p = pe["position"]
            lp += 1
            qty = float(pe["qty"] or 0)
            wartosc = float(pe["budzet"] or 0)
            cena = wartosc / qty if qty > 0 else 0
            ws.cell(row=r, column=1, value=lp).alignment = Alignment(horizontal="center", vertical="center")
            name_cell = ws.cell(row=r, column=2, value=p.get("name", ""))
            name_cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row=r, column=3, value=qty).number_format = '#,##0.00'
            ws.cell(row=r, column=3).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=r, column=4, value=p.get("unit") or "").alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r, column=5, value=cena).number_format = '#,##0.00" zł"'
            ws.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center")
            # iter95ap: AKTYWNA FORMULA wartosc = ilosc * cena
            ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = '#,##0.00" zł"'
            ws.cell(row=r, column=6).font = Font(bold=True)
            ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="center")
            for col in range(1, 7):
                ws.cell(row=r, column=col).border = border
            # iter95w: auto-rozmiar wiersza dla dlugich nazw
            name_len = len(p.get("name", "") or "")
            if name_len > 50:
                ws.row_dimensions[r].height = min(60, 18 + (name_len // 40) * 12)
            if first_pos_row is None:
                first_pos_row = r
            last_pos_row = r
            r += 1

    # wiersz sumy z formula
    sum_row = r
    # iter95bg: scal kolumny A-E (1-5) zeby "RAZEM netto:" zmiescil sie bez wychodzenia poza komorke
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    razem_cell = ws.cell(row=r, column=1, value="RAZEM netto:")
    razem_cell.font = Font(bold=True, size=12)
    razem_cell.alignment = Alignment(horizontal="right", vertical="center")
    razem_cell.fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    if first_pos_row and last_pos_row:
        ws.cell(row=r, column=6, value=f"=SUM(F{first_pos_row}:F{last_pos_row})")
    else:
        ws.cell(row=r, column=6, value=0)
    ws.cell(row=r, column=6).font = Font(bold=True, size=12, color="B8860B")
    ws.cell(row=r, column=6).number_format = '#,##0.00" zł"'
    ws.cell(row=r, column=6).fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    # border na wszystkich scalonych komorkach + wartosci
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = border
    r += 2

    # wskazniki: dziele SUM przez powierzchnie - aktywne formuly!
    if include_wskazniki and (pc or pc_pod or pc_nad or pum) and last_pos_row:
        ws.cell(row=r, column=1, value="Wskaźniki kosztowe").font = Font(bold=True, size=10, color="3F5235")
        r += 1
        for col, h in enumerate(["Wskaźnik", "Wartość"], start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
        r += 1
        sum_cell = f"F{sum_row}"
        def _wsk(label, val):
            nonlocal r
            if val and float(val) > 0 and label in surface_row_map:
                surf_cell = f"B{surface_row_map[label]}"
                ws.cell(row=r, column=1, value=label.replace("  ", "") + " (zł/m²)")
                ws.cell(row=r, column=2, value=f"={sum_cell}/{surf_cell}")
                ws.cell(row=r, column=2).number_format = '#,##0.00" zł/m²"'
                ws.cell(row=r, column=2).font = Font(bold=True, color="3F5235")
                r += 1
        _wsk("Powierzchnia całkowita (PC)", pc)
        _wsk("  ↓ w tym podziemie", pc_pod)
        _wsk("  ↑ w tym nadziemie", pc_nad)
        _wsk("Powierzchnia użytkowa mieszkalna (PUM)", pum)
        r += 1

    if include_notes:
        notice = w_full.get("notes") or (
            "Oferta ważna 30 dni od daty wystawienia. "
            "Podane ceny są cenami netto. Płatność wg ustaleń umowy. "
            "Zakres prac i warunki realizacji do uzgodnienia."
        )
        ws.cell(row=r, column=1, value="Uwagi:").font = Font(bold=True, size=9, color="666666")
        r += 1
        ws.cell(row=r, column=1, value=notice).font = Font(size=8, color="666666")
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 40
        r += 2

    # iter95w: Oferta obejmuje / nie obejmuje
    scope_inc = (w_full.get("scope_includes") or "").strip()
    scope_exc = (w_full.get("scope_excludes") or "").strip()

    def _render_scope_xlsx(title: str, text: str, color_text: str, color_bg: str, color_border: str):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=11, color="3F5235")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        lines = [ln.strip() for ln in (text or "").replace("\r", "").split("\n") if ln.strip()]
        bullet_text = "\n".join(
            (ln if ln.startswith(("\u2022", "-", "*")) else f"\u2022 {ln}") for ln in lines
        )
        cell = ws.cell(row=r, column=1, value=bullet_text)
        cell.font = Font(size=10, color=color_text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = PatternFill(start_color=color_bg, end_color=color_bg, fill_type="solid")
        thin_border = Border(
            left=Side(border_style="thin", color=color_border),
            right=Side(border_style="thin", color=color_border),
            top=Side(border_style="thin", color=color_border),
            bottom=Side(border_style="thin", color=color_border),
        )
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = thin_border
            ws.cell(row=r, column=col).fill = PatternFill(start_color=color_bg, end_color=color_bg, fill_type="solid")
        # wysokosc wiersza wg ilosci linii
        ws.row_dimensions[r].height = max(28, len(lines) * 18 + 8)
        r += 2

    if scope_inc:
        _render_scope_xlsx("Oferta obejmuje", scope_inc, "222222", "F2F7EC", "5F7552")
    if scope_exc:
        _render_scope_xlsx("Oferta nie obejmuje", scope_exc, "7A2E2E", "FDF2F2", "9B2C2C")

    # szerokosci kolumn - szersza kolumna Nazwa pozycji dla dlugich opisow
    widths = {"A": 8, "B": 65, "C": 13, "D": 9, "E": 16, "F": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # info dla inwestora w tooltipie naglowka kolumny F
    from openpyxl.comments import Comment
    ws.cell(row=header_row, column=6).comment = Comment(
        "Formuła: ilość × cena netto. Możesz zmienić ilość lub cenę — wartość przeliczy się automatycznie.",
        "FeGrro"
    )

    buf = BytesIO()
    wb.save(buf)
    safe_name = (wycena_name or "wycena").replace("/", "_").replace(" ", "_")[:50]
    filename = f"Oferta_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename


# ============================================================
# iter95av: SNAPSHOTY + TRYB NEGOCJACJI
# ============================================================

async def _create_wycena_snapshot(wycena_id: str, label: str, user_id: str) -> str:
    """Skopiuj pelna strukture wyceny do wyceny_snapshots. Zwraca snapshot_id."""
    wycena = await db.wyceny.find_one({"id": wycena_id}, {"_id": 0})
    if not wycena:
        raise HTTPException(404, "Wycena nie istnieje")
    stages = await db.wyceny_stages.find({"wycena_id": wycena_id}, {"_id": 0}).to_list(length=None)
    positions = await db.wyceny_positions.find({"wycena_id": wycena_id}, {"_id": 0}).to_list(length=None)
    lines = await db.wyceny_lines.find({"wycena_id": wycena_id}, {"_id": 0}).to_list(length=None)
    snapshot_id = str(uuid.uuid4())
    await db.wyceny_snapshots.insert_one({
        "id": snapshot_id,
        "wycena_id": wycena_id,
        "label": label,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user_id,
        "data": {
            "wycena": wycena,
            "stages": stages,
            "positions": positions,
            "lines": lines,
        },
        "stats": {
            "stages": len(stages),
            "positions": len(positions),
            "lines": len(lines),
        },
    })
    return snapshot_id


@router.get("/wyceny/{wycena_id}/snapshots")
async def list_snapshots(wycena_id: str, _user: dict = Depends(get_current_admin)):
    rows = await db.wyceny_snapshots.find(
        {"wycena_id": wycena_id},
        {"_id": 0, "id": 1, "label": 1, "created_at": 1, "created_by": 1, "stats": 1},
    ).sort("created_at", -1).to_list(length=200)
    return {"rows": rows}


class SnapshotCreateRequest(BaseModel):
    label: str


@router.post("/wyceny/{wycena_id}/snapshots")
async def create_snapshot(
    wycena_id: str, payload: SnapshotCreateRequest,
    current_user: dict = Depends(get_current_admin),
):
    sid = await _create_wycena_snapshot(wycena_id, payload.label.strip() or "Wersja", current_user["sub"])
    return {"ok": True, "id": sid}


@router.post("/wyceny/{wycena_id}/snapshots/{snapshot_id}/restore")
async def restore_snapshot(
    wycena_id: str, snapshot_id: str,
    current_user: dict = Depends(get_current_admin),
):
    snap = await db.wyceny_snapshots.find_one(
        {"id": snapshot_id, "wycena_id": wycena_id}, {"_id": 0}
    )
    if not snap:
        raise HTTPException(404, "Snapshot nie istnieje")
    # Auto-snapshot biezacego stanu przed nadpisaniem
    await _create_wycena_snapshot(
        wycena_id,
        f"Auto: stan przed przywróceniem '{snap['label']}'",
        current_user["sub"],
    )
    # Wyczysc i wpisz dane ze snapshota
    data = snap["data"]
    await db.wyceny.update_one({"id": wycena_id}, {"$set": data["wycena"]})
    await db.wyceny_stages.delete_many({"wycena_id": wycena_id})
    if data["stages"]:
        await db.wyceny_stages.insert_many(data["stages"])
    await db.wyceny_positions.delete_many({"wycena_id": wycena_id})
    if data["positions"]:
        await db.wyceny_positions.insert_many(data["positions"])
    await db.wyceny_lines.delete_many({"wycena_id": wycena_id})
    if data["lines"]:
        await db.wyceny_lines.insert_many(data["lines"])
    return {"ok": True}


@router.delete("/wyceny/{wycena_id}/snapshots/{snapshot_id}")
async def delete_snapshot(
    wycena_id: str, snapshot_id: str, _user: dict = Depends(get_current_admin),
):
    r = await db.wyceny_snapshots.delete_one({"id": snapshot_id, "wycena_id": wycena_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Snapshot nie istnieje")
    return {"ok": True}


class NegotiationApplyRequest(BaseModel):
    """Trwale zastosowanie wynikow negocjacji."""
    labor_factor: float = 1.0  # mnoznik ceny robocizny (0.98 = -2%)
    material_factor: float = 1.0  # mnoznik ceny materialow
    equipment_factor: float = 1.0  # mnoznik ceny sprzetu
    narzut_pct: Optional[float] = None  # nowy default_narzut_pct (lub None = bez zmian)
    marza_pct: Optional[float] = None
    snapshot_label: Optional[str] = None  # etykieta auto-snapshota


# iter95bs: refresh ceny z cennika dla wszystkich linii w wycenie
@router.post("/wyceny/{wycena_id}/refresh-prices")
async def refresh_wycena_prices(
    wycena_id: str,
    current_user: dict = Depends(get_current_admin),
):
    """Aktualizuje pola (name, unit_price_netto, price_min/max) we wszystkich liniach
    wyceny ktore maja price_book_id - biorac aktualne wartosci z `wyceny_price_book`.
    Zwraca licznik zaktualizowanych linii.
    """
    wycena = await db.wyceny.find_one({"id": wycena_id}, {"_id": 0, "id": 1})
    if not wycena:
        raise HTTPException(404, "Wycena nie istnieje")
    lines_cursor = db.wyceny_lines.find(
        {"wycena_id": wycena_id, "price_book_id": {"$ne": None}},
        {"_id": 0},
    )
    updated_count = 0
    skipped_count = 0
    async for line in lines_cursor:
        pb = await db.wyceny_price_book.find_one(
            {"id": line["price_book_id"]}, {"_id": 0}
        )
        if not pb:
            skipped_count += 1
            continue
        ltype = line.get("type")  # materials | labor | equipment
        updates = {}
        if pb.get("name") and pb["name"] != line.get("name"):
            updates["name"] = pb["name"]
        pb_min = pb.get("price_min")
        pb_max = pb.get("price_max")
        if pb_min is not None:
            cur_min = line.get("price_min")
            if ltype in ("materials", "equipment"):
                new_min = max(float(cur_min), float(pb_min)) if cur_min is not None else float(pb_min)
            else:
                new_min = float(pb_min)
            if cur_min != new_min:
                updates["price_min"] = new_min
        if pb_max is not None:
            cur_max = line.get("price_max")
            if cur_max != pb_max:
                updates["price_max"] = float(pb_max)
        new_price = None
        if ltype == "labor":
            lu = (line.get("unit") or "").strip()
            if lu == "m²" and pb.get("price_m2") is not None:
                new_price = float(pb["price_m2"])
            elif lu == "m³" and pb.get("price_m3") is not None:
                new_price = float(pb["price_m3"])
            elif pb.get("unit_other") and pb["unit_other"] == lu and pb.get("price_other") is not None:
                new_price = float(pb["price_other"])
        else:
            if pb.get("unit_price_netto") is not None:
                new_price = float(pb["unit_price_netto"])
        if new_price is not None and abs((line.get("unit_price_netto") or 0) - new_price) > 1e-9:
            updates["unit_price_netto"] = new_price
        if not updates:
            skipped_count += 1
            continue
        updates["updated_at"] = datetime.now().isoformat()
        await db.wyceny_lines.update_one({"id": line["id"]}, {"$set": updates})
        updated_count += 1
    return {
        "updated": updated_count,
        "skipped": skipped_count,
        "total_linked": updated_count + skipped_count,
    }



@router.post("/wyceny/{wycena_id}/negotiation/apply")
async def apply_negotiation(
    wycena_id: str,
    payload: NegotiationApplyRequest,
    current_user: dict = Depends(get_current_admin),
):
    """Trwale zaakceptuj wyniki negocjacji.

    Kroki:
      1. Auto-snapshot biezacego stanu z labelka (default: "Przed negocjacja {data}")
      2. Aktualizuj defaults wyceny (narzut/marza)
      3. Bulk update unit_price_netto we wszystkich liniach wedlug category factor
    """
    # 1. Snapshot
    snap_label = payload.snapshot_label or f"Przed negocjacją ({datetime.now().strftime('%Y-%m-%d %H:%M')})"
    sid = await _create_wycena_snapshot(wycena_id, snap_label, current_user["sub"])

    # 2. Update defaults wyceny
    update = {}
    if payload.narzut_pct is not None:
        update["default_narzut_pct"] = float(payload.narzut_pct)
    if payload.marza_pct is not None:
        update["default_marza_pct"] = float(payload.marza_pct)
    if update:
        await db.wyceny.update_one({"id": wycena_id}, {"$set": update})

    # 3. Bulk update lines per type (materials/labor/equipment)
    factor_by_type = {
        "labor": payload.labor_factor,
        "materials": payload.material_factor,
        "equipment": payload.equipment_factor,
    }
    modified_total = 0
    for typ, factor in factor_by_type.items():
        if factor == 1.0:
            continue
        r = await db.wyceny_lines.update_many(
            {"wycena_id": wycena_id, "type": typ},
            {"$mul": {"unit_price_netto": float(factor)}},
        )
        modified_total += r.modified_count

    return {
        "ok": True,
        "snapshot_id": sid,
        "snapshot_label": snap_label,
        "lines_modified": modified_total,
    }


# iter95as: konwersja wyceny do budowy + budzetu (kopiuje pelna strukture)
class ConvertToBudgetRequest(BaseModel):
    budowa_name: str
    code: Optional[str] = None
    zamawiajacy: Optional[str] = None
    umowa_nr: Optional[str] = None
    umowa_data: Optional[str] = None


@router.post("/wyceny/{wycena_id}/convert-to-budget")
async def convert_wycena_to_budget(
    wycena_id: str,
    payload: ConvertToBudgetRequest,
    current_user: dict = Depends(get_current_admin),
):
    """Stworz nowa budowe (finance_budowy) + skopiowana pelna strukture wyceny do budzetu:
    - wyceny -> finance_budowy (z kaucjami/koszt budowy z defaults wyceny)
    - wyceny_stages -> budget_stages
    - wyceny_positions -> budget_positions
    - wyceny_lines -> budget_lines (z position_id)
    Zwraca {budowa_id, budowa_name, stats}.
    """
    wycena = await db.wyceny.find_one({"id": wycena_id}, {"_id": 0})
    if not wycena:
        raise HTTPException(404, "Wycena nie istnieje")

    # 1. Sprawdz czy budowa o tej nazwie juz istnieje
    exists = await db.finance_budowy.find_one({"name": payload.budowa_name}, {"_id": 0, "id": 1})
    if exists:
        raise HTTPException(400, f"Budowa o nazwie '{payload.budowa_name}' już istnieje")

    # 2. Stworz budowe
    budowa_id = str(uuid.uuid4())
    zamawiajacy = (payload.zamawiajacy or "").strip()
    # Pre-fill zamawiajacy z client_name wyceny jezeli puste
    if not zamawiajacy:
        cname = (wycena.get("client_name") or "").strip()
        cnip = (wycena.get("client_nip") or "").strip()
        if cname:
            zamawiajacy = f"{cname}" + (f" NIP: {cnip}" if cnip else "")
    budowa_doc = {
        "id": budowa_id,
        "name": payload.budowa_name,
        "code": payload.code or "",
        "show_in_hours": True,
        "has_budget": True,
        "is_gir": False,
        "is_dw": False,
        "kaucja_gir_pct": float(wycena.get("default_gir_pct") or 2.0),
        "kaucja_dw_pct": float(wycena.get("default_dw_pct") or 2.0),
        "koszt_budowy_pct": float(wycena.get("default_koszt_pct") or 0.0),
        "notes": wycena.get("notes") or "",
        "zamawiajacy": zamawiajacy,
        "umowa_nr": payload.umowa_nr or "",
        "umowa_data": payload.umowa_data or "",
        "wykonawca": "FEGRRO SP. Z O.O. NIP: 589-206-61-74",
        "archived": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["sub"],
        # link wsteczny do wyceny zrodlowej
        "source_wycena_id": wycena_id,
    }
    await db.finance_budowy.insert_one(budowa_doc)

    # 3. Skopiuj etapy
    src_stages = await db.wyceny_stages.find(
        {"wycena_id": wycena_id}, {"_id": 0}
    ).sort("order", 1).to_list(length=None)
    stage_map = {}  # old_stage_id -> new_stage_id
    for st in src_stages:
        new_sid = str(uuid.uuid4())
        stage_map[st["id"]] = new_sid
        await db.budget_stages.insert_one({
            "id": new_sid,
            "budowa_id": budowa_id,
            "name": st.get("name") or "Etap",
            "start_date": None,
            "end_date": None,
            "order": int(st.get("order") or 0),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })

    # 4. Skopiuj pozycje glowne + ich subpozycje (lines)
    src_positions = await db.wyceny_positions.find(
        {"wycena_id": wycena_id}, {"_id": 0}
    ).sort("order", 1).to_list(length=None)
    pos_count = 0
    line_count = 0
    for p in src_positions:
        new_pid = str(uuid.uuid4())
        new_sid = stage_map.get(p.get("stage_id"))
        if not new_sid:
            continue  # bez etapu nie da sie utworzyc pozycji
        await db.budget_positions.insert_one({
            "id": new_pid,
            "budowa_id": budowa_id,
            "stage_id": new_sid,
            "name": p.get("name") or "Pozycja",
            "notes": p.get("notes") or "",
            "order": int(p.get("order") or 0),
            "include_in_protocol": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        pos_count += 1
        # subpozycje (lines) tej pozycji
        src_lines = await db.wyceny_lines.find(
            {"wycena_id": wycena_id, "position_id": p["id"]}, {"_id": 0}
        ).sort("order", 1).to_list(length=None)
        # Mapowanie wyceny.category (materials/labor/equipment) -> budget.type
        # Plus wyznaczenie czytelnej kategorii (w budgecie to label, nie enum)
        type_to_cat = {"materials": "Materiały", "labor": "Robocizna", "equipment": "Sprzęt"}
        for ln in src_lines:
            qty = float(ln.get("quantity") or 0)
            price = float(ln.get("unit_price_netto") or 0)
            t = ln.get("category") or "materials"
            await db.budget_lines.insert_one({
                "id": str(uuid.uuid4()),
                "budowa_id": budowa_id,
                "category": type_to_cat.get(t, "Materiały"),
                "name": ln.get("name") or "—",
                "type": t,
                "unit": ln.get("unit") or "",
                "quantity": qty,
                "unit_price_netto": price,
                "plan_netto": round(qty * price, 2),
                "kaucja_gir_pct": None,  # dziedziczy z budowa
                "kaucja_dw_pct": None,
                "stage_id": new_sid,
                "position_id": new_pid,
                "parent_id": None,
                "is_income": False,
                "forecast_cost": None,
                "forecast_note": None,
                "notes": ln.get("notes") or "",
                "order": int(ln.get("order") or 0),
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            line_count += 1

    return {
        "ok": True,
        "budowa_id": budowa_id,
        "budowa_name": payload.budowa_name,
        "stats": {
            "stages": len(src_stages),
            "positions": pos_count,
            "lines": line_count,
        },
    }



# =========== iter95v: IMPORT Z EXCELA ===========

class ExcelImportPreviewBody(BaseModel):
    file_base64: str  # data URI or pure base64 content of XLSX
    sheet_name: Optional[str] = None


@router.post("/wyceny/import/preview")
async def wyceny_import_preview(payload: ExcelImportPreviewBody, _user: dict = Depends(get_current_admin)):
    """iter95v: Sparsuj XLSX i zwroc preview (sheety + wiersze) do mapowania w UI.

    Klient wysyla plik jako base64 (data URI lub czysty base64).
    Odpowiedz: { sheets: [{ name, rows: [[..],[..]], cols: int }] }
    """
    raw = payload.file_base64 or ""
    if "," in raw:
        raw = raw.split(",", 1)[1]
    try:
        data = base64.b64decode(raw)
    except Exception:
        raise HTTPException(400, "Nieprawidlowy plik base64")
    if not data:
        raise HTTPException(400, "Pusty plik")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl niedostepny")

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Nie udalo sie odczytac pliku XLSX: {e}")

    sheets = []
    target_sheets = [payload.sheet_name] if payload.sheet_name else wb.sheetnames
    for sname in target_sheets:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        rows_data = []
        max_cols = 0
        # Limit do 500 wierszy x 30 kolumn dla preview
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx >= 500:
                break
            cells = []
            for c_idx, val in enumerate(row[:30]):
                if val is None:
                    cells.append("")
                elif isinstance(val, (int, float)):
                    # Normalizuj liczby (usun trailing .0)
                    if isinstance(val, float) and val.is_integer():
                        cells.append(str(int(val)))
                    else:
                        cells.append(str(val))
                else:
                    cells.append(str(val).strip())
            # Usun trailing puste komorki
            while cells and not cells[-1]:
                cells.pop()
            max_cols = max(max_cols, len(cells))
            rows_data.append(cells)
        # Wyrownaj wszystkie wiersze do max_cols
        for i in range(len(rows_data)):
            while len(rows_data[i]) < max_cols:
                rows_data[i].append("")
        sheets.append({
            "name": sname,
            "rows": rows_data,
            "cols": max_cols,
            "row_count": len(rows_data),
        })

    return {"sheets": sheets}


class ImportRowMap(BaseModel):
    row_index: int  # 0-based
    role: str       # "stage" | "position" | "skip"


class ExcelImportApplyBody(BaseModel):
    file_base64: str
    sheet_name: str
    name_col: int           # kolumna z nazwa etapu/pozycji
    unit_col: Optional[int] = None
    quantity_col: Optional[int] = None
    notes_col: Optional[int] = None
    rows: List[ImportRowMap]
    default_stage_name: str = "Etap 1"  # gdy pozycja pojawi sie przed pierwszym etapem


@router.post("/wyceny/{wycena_id}/import/apply")
async def wyceny_import_apply(
    wycena_id: str,
    payload: ExcelImportApplyBody,
    _user: dict = Depends(get_current_admin),
):
    """iter95v: Importuj zmapowane wiersze jako etapy + pozycje do wyceny."""
    if not await db.wyceny.find_one({"id": wycena_id}, {"_id": 0, "id": 1}):
        raise HTTPException(404, "Wycena nie istnieje")

    raw = payload.file_base64 or ""
    if "," in raw:
        raw = raw.split(",", 1)[1]
    try:
        data = base64.b64decode(raw)
    except Exception:
        raise HTTPException(400, "Nieprawidlowy plik base64")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Blad XLSX: {e}")

    if payload.sheet_name not in wb.sheetnames:
        raise HTTPException(400, f"Sheet '{payload.sheet_name}' nie istnieje")
    ws = wb[payload.sheet_name]

    # Zaladuj wszystkie wiersze do listy dla indeksowania
    all_rows: list = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    def cell(r: int, c: Optional[int]) -> str:
        if c is None or c < 0 or r >= len(all_rows) or c >= len(all_rows[r]):
            return ""
        v = all_rows[r][c]
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip() if isinstance(v, str) else str(v)

    def to_float(v: str) -> Optional[float]:
        if not v:
            return None
        s = str(v).replace(" ", "").replace(",", ".")
        # Usun jednostki w nawiasach np. "12 (m2)"
        try:
            return float(s)
        except Exception:
            # sprobuj pierwszej liczby z napisu
            import re
            m = re.search(r"-?\d+(?:\.\d+)?", s)
            return float(m.group(0)) if m else None

    def normalize_unit(v: str) -> Optional[str]:
        """iter95v: znormalizuj jednostke do formatu uzywanego przez UI.
        m2 -> m², m3 -> m³, godz/h -> godz, dni -> dzien itp.
        """
        if not v:
            return None
        s = str(v).strip().lower().replace(" ", "")
        # Usun kropki ze skroten
        s = s.replace(".", "")
        # Mapowania wzajemnie zamienne
        aliases = {
            "m2": "m\u00b2", "m^2": "m\u00b2", "m\u00b2": "m\u00b2",
            "m3": "m\u00b3", "m^3": "m\u00b3", "m\u00b3": "m\u00b3",
            "mb": "mb", "mp": "mb",
            "szt": "szt", "sztuk": "szt", "pcs": "szt",
            "kg": "kg", "kilogram": "kg",
            "t": "t", "tona": "t", "ton": "t",
            "godz": "godz", "h": "godz", "rgodz": "godz", "rg": "godz",
            "dzien": "dzie\u0144", "dzień": "dzie\u0144", "dni": "dzie\u0144", "day": "dzie\u0144", "d": "dzie\u0144",
            "m-c": "m-c", "mc": "m-c", "miesiac": "m-c", "miesi\u0105c": "m-c", "month": "m-c",
            "kpl": "kpl", "komplet": "kpl",
            "l": "l", "litr": "l",
            "m": "mb",  # samo "m" to zwykle metr biezacy
        }
        return aliases.get(s, str(v).strip())  # zachowaj oryginal gdy nie znamy

    # Sortuj rows wedlug row_index zeby etapy/pozycje zachowaly kolejnosc
    mapped = sorted(payload.rows, key=lambda r: r.row_index)

    # Istniejace etapy - zeby kolejnosc nowych byla po nich
    existing_stages_count = await db.wyceny_stages.count_documents({"wycena_id": wycena_id})
    existing_positions_count = await db.wyceny_positions.count_documents({"wycena_id": wycena_id})

    current_stage_id: Optional[str] = None
    stage_order = existing_stages_count
    pos_order = existing_positions_count
    created_stages = 0
    created_positions = 0
    skipped = 0
    now = datetime.now().isoformat()

    for m in mapped:
        if m.role == "skip":
            skipped += 1
            continue
        name = cell(m.row_index, payload.name_col)
        if not name:
            skipped += 1
            continue

        if m.role == "stage":
            sid = str(uuid.uuid4())
            await db.wyceny_stages.insert_one({
                "id": sid,
                "wycena_id": wycena_id,
                "name": name,
                "order": stage_order,
                "created_at": now,
            })
            current_stage_id = sid
            stage_order += 1
            created_stages += 1
        elif m.role == "position":
            # Gdy nie mamy jeszcze etapu - stworz domyslny
            if not current_stage_id:
                sid = str(uuid.uuid4())
                await db.wyceny_stages.insert_one({
                    "id": sid,
                    "wycena_id": wycena_id,
                    "name": payload.default_stage_name,
                    "order": stage_order,
                    "created_at": now,
                })
                current_stage_id = sid
                stage_order += 1
                created_stages += 1

            pid = str(uuid.uuid4())
            doc = {
                "id": pid,
                "wycena_id": wycena_id,
                "stage_id": current_stage_id,
                "name": name,
                "order": pos_order,
                "quantity": to_float(cell(m.row_index, payload.quantity_col)) if payload.quantity_col is not None else None,
                "unit": normalize_unit(cell(m.row_index, payload.unit_col)) if payload.unit_col is not None else None,
                "kaucja_gir_pct": 2.0,
                "kaucja_dw_pct": 2.0,
                "koszt_budowy_pct": 2.0,
                "koszt_prognozowany": None,
                "created_at": now,
            }
            # uwagi -> dopisz do nazwy w nawiasie jezeli sa
            if payload.notes_col is not None:
                notes = cell(m.row_index, payload.notes_col)
                if notes:
                    doc["notes"] = notes
            await db.wyceny_positions.insert_one(doc)
            pos_order += 1
            created_positions += 1
        else:
            skipped += 1

    return {
        "ok": True,
        "stages_created": created_stages,
        "positions_created": created_positions,
        "skipped": skipped,
    }



# =========== iter95x: SZABLONY ZAKRESU OFERTY (Settings) ===========

class ScopeTemplate(BaseModel):
    id: Optional[str] = None
    name: str  # np. "Dom jednorodzinny", "Mieszkanie", "Komercja", "Default"
    scope_includes: str = ""
    scope_excludes: str = ""
    is_default: bool = False


class ScopeTemplatesPayload(BaseModel):
    templates: List[ScopeTemplate]


SCOPE_SETTINGS_KEY = "wyceny_scope_templates"


@router.get("/wyceny/scope-templates")
async def get_scope_templates(_user: dict = Depends(get_current_admin)):
    doc = await db.app_settings.find_one({"key": SCOPE_SETTINGS_KEY}, {"_id": 0})
    if not doc:
        return {"templates": []}
    return {"templates": doc.get("templates", [])}


@router.put("/wyceny/scope-templates")
async def save_scope_templates(
    payload: ScopeTemplatesPayload,
    _user: dict = Depends(get_current_admin),
):
    # Tylko jeden moze byc is_default
    default_seen = False
    out_tpl = []
    for t in payload.templates:
        td = t.model_dump()
        if not td.get("id"):
            td["id"] = str(uuid.uuid4())
        if td.get("is_default"):
            if default_seen:
                td["is_default"] = False
            else:
                default_seen = True
        out_tpl.append(td)
    await db.app_settings.update_one(
        {"key": SCOPE_SETTINGS_KEY},
        {"$set": {"key": SCOPE_SETTINGS_KEY, "templates": out_tpl,
                  "updated_at": datetime.now().isoformat()}},
        upsert=True,
    )
    return {"templates": out_tpl}

# iter95bq: zarzadzanie custom kategoriami robocizny
LABOR_CATS_KEY = "wyceny_labor_custom_cats"


@router.get("/wyceny/labor-categories")
async def get_labor_categories(_user: dict = Depends(get_current_admin)):
    """Zwraca {custom: [...]} z app_settings (default sa po stronie frontendu w LABOR_SUB_CATS)."""
    doc = await db.app_settings.find_one({"key": LABOR_CATS_KEY}, {"_id": 0})
    return {"custom": (doc or {}).get("categories", [])}


@router.post("/wyceny/labor-categories")
async def add_labor_category(payload: dict, _user: dict = Depends(get_current_admin)):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Nazwa kategorii nie moze byc pusta")
    if len(name) > 40:
        raise HTTPException(400, "Nazwa kategorii max 40 znakow")
    doc = await db.app_settings.find_one({"key": LABOR_CATS_KEY}, {"_id": 0}) or {}
    cats = list(doc.get("categories", []))
    if name in cats:
        raise HTTPException(400, f"Kategoria '{name}' juz istnieje")
    cats.append(name)
    await db.app_settings.update_one(
        {"key": LABOR_CATS_KEY},
        {"$set": {"key": LABOR_CATS_KEY, "categories": cats,
                  "updated_at": datetime.now().isoformat()}},
        upsert=True,
    )
    return {"custom": cats}


@router.delete("/wyceny/labor-categories/{name}")
async def delete_labor_category(name: str, _user: dict = Depends(get_current_admin)):
    # walidacja: czy nie ma pozycji w tej kategorii
    count = await db.wyceny_price_book.count_documents(
        {"category": "labor", "sub_category": name}
    )
    if count > 0:
        raise HTTPException(
            400,
            f"Kategoria '{name}' ma {count} pozycji. Usun je najpierw lub przenies do innej kategorii.",
        )
    doc = await db.app_settings.find_one({"key": LABOR_CATS_KEY}, {"_id": 0}) or {}
    cats = [c for c in doc.get("categories", []) if c != name]
    await db.app_settings.update_one(
        {"key": LABOR_CATS_KEY},
        {"$set": {"key": LABOR_CATS_KEY, "categories": cats,
                  "updated_at": datetime.now().isoformat()}},
        upsert=True,
    )
    return {"custom": cats}
