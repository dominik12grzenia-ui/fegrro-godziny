"""Wyceny BOM (Bill of Materials) — wydzielone z routes/wyceny.py (iter95dw).

Zawiera:
- `_build_bom(wycena_id)` - agregacja materiałów + cennik + obliczenia opakowań
- `GET /wyceny/{id}/bom` - JSON podgląd
- `GET /wyceny/{id}/bom.xlsx` - eksport XLSX
- `GET /wyceny/{id}/bom.pdf` - eksport PDF (zapytanie ofertowe)
- `_generate_bom_xlsx_bytes` / `_generate_bom_pdf_bytes` - generatory
- Suppliers CRUD (`/wyceny/suppliers/*`) + wysyłka emailem (`POST /wyceny/{id}/bom/send`)
- `GET /wyceny/{id}/bom/history`

Reszta wyceny.py pozostała na 2 router.include_router(bom_router) wynik.
"""
import os
import uuid
import io
import logging
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr

from database import db
from auth import get_current_admin
from routes.wyceny_helpers import (
    _get_logo_path,
    _register_pdf_fonts,
    _safe_content_disposition,
    _xlsx_add_logo,
    _filter_bom_rows,
)

router = APIRouter()
logger = logging.getLogger(__name__)


# ============================================================
# _build_bom — agregacja materialow z wyceny
# ============================================================
async def _build_bom(wycena_id: str):
    """Zestawienie: grupuj po nazwie, sumuj ilosci, dolicz liczbe opakowan (ceil)
    na bazie cennika (pkg_qty + zapotrzebowanie + zap_unit)."""
    import math
    wycena = await db.wyceny.find_one({"id": wycena_id}, {"_id": 0})
    if not wycena:
        raise HTTPException(404, "Wycena nie istnieje")
    lines = await db.wyceny_lines.find(
        {"wycena_id": wycena_id, "type": "materials"}, {"_id": 0}
    ).to_list(length=None)
    # iter95dl: pomin lines z pozycji wylaczonych z wyceny
    excluded_pos_ids = set()
    async for p in db.wyceny_positions.find(
        {"wycena_id": wycena_id, "excluded": True}, {"_id": 0, "id": 1}
    ):
        excluded_pos_ids.add(p["id"])
    if excluded_pos_ids:
        lines = [ln for ln in lines if ln.get("position_id") not in excluded_pos_ids]
    # Pobierz cennik materialow do dopasowania po nazwie
    cennik = await db.wyceny_price_book.find(
        {"category": "materials"}, {"_id": 0}
    ).to_list(length=None)
    cennik_by_name = {(c.get("name") or "").lower().strip(): c for c in cennik}

    grouped: dict = {}
    for ln in lines:
        name = (ln.get("name") or "").strip()
        if not name:
            continue
        unit = (ln.get("unit") or "").strip()
        qty = float(ln.get("quantity") or 0)
        if qty <= 0:
            continue
        key = (name.lower(), unit.lower())
        if key not in grouped:
            ce = cennik_by_name.get(name.lower(), {}) or {}
            grouped[key] = {
                "name": name,
                "unit": unit,
                "quantity": 0.0,
                "occurrences": 0,
                "opakowanie": ce.get("opakowanie") or "",
                "pkg_qty": ce.get("pkg_qty"),
                "pkg_unit": ce.get("pkg_unit") or "",
                "zapotrzebowanie": ce.get("zapotrzebowanie"),
                "zap_unit": ce.get("zap_unit") or "",
                "sub_category": ce.get("sub_category") or "",
            }
        grouped[key]["quantity"] += qty
        grouped[key]["occurrences"] += 1

    rows = []
    for r in grouped.values():
        qty_in_pkg_unit = None
        num_packages = None
        pkg_qty = r.get("pkg_qty")
        zap = r.get("zapotrzebowanie")
        zap_unit = r.get("zap_unit") or ""
        pkg_unit = r.get("pkg_unit") or ""
        if pkg_qty and pkg_qty > 0:
            if "/" in zap_unit and zap:
                _num, denom = zap_unit.split("/", 1)
                if denom.strip() == r["unit"]:
                    qty_in_pkg_unit = r["quantity"] * zap
            if qty_in_pkg_unit is None and r["unit"] and r["unit"] == pkg_unit:
                qty_in_pkg_unit = r["quantity"]
            if qty_in_pkg_unit is not None:
                num_packages = math.ceil(qty_in_pkg_unit / pkg_qty)
        rows.append({
            "name": r["name"],
            "unit": r["unit"],
            "quantity": round(r["quantity"], 3),
            "occurrences": r["occurrences"],
            "opakowanie": r["opakowanie"],
            "pkg_qty": r["pkg_qty"],
            "pkg_unit": pkg_unit,
            "qty_in_pkg_unit": round(qty_in_pkg_unit, 3) if qty_in_pkg_unit is not None else None,
            "num_packages": num_packages,
            "sub_category": r.get("sub_category") or "",
        })
    rows.sort(key=lambda r: r["name"].lower())
    return {"wycena_name": wycena.get("name", ""), "rows": rows}


# ============================================================
# BOM Endpointy: JSON / XLSX / PDF
# ============================================================
@router.get("/wyceny/{wycena_id}/bom")
async def get_materials_bom(wycena_id: str, _user: dict = Depends(get_current_admin)):
    """JSON: zestawienie materialow do podgladu we frontendzie."""
    return await _build_bom(wycena_id)


@router.get("/wyceny/{wycena_id}/bom.xlsx")
async def export_bom_xlsx(
    wycena_id: str,
    subcategories: Optional[str] = Query(None, description="Lista sub_categories oddzielona przecinkiem"),
    _user: dict = Depends(get_current_admin),
):
    data = await _build_bom(wycena_id)
    if subcategories:
        data = _filter_bom_rows(data, [s.strip() for s in subcategories.split(",") if s.strip()])
    content, filename = _generate_bom_xlsx_bytes(data)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _safe_content_disposition("attachment", filename)},
    )


def _generate_bom_xlsx_bytes(data: dict):
    """Refaktor iter95ah: generuj XLSX jako bytes - reuzywalne przez endpoint i wysylke maila."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Zestawienie materiałów"
    _xlsx_add_logo(ws, "A1", width=90, height=90)
    ws["B1"] = f"Zestawienie materiałów: {data['wycena_name']}"
    ws["B1"].font = Font(bold=True, size=14, color="D4AF37")
    ws.merge_cells("B1:E1")
    ws["B2"] = f"Data wygenerowania: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("B2:E2")
    ws.row_dimensions[1].height = 50
    headers = ["L.p.", "Nazwa materiału", "Ilość zużycia", "Jednostka",
               "Opakowanie", "Wielkość opak.", "Liczba opakowań",
               "Cena netto za opak. (PLN)", "Wartość netto (PLN)", "Uwagi"]
    header_fill = PatternFill(start_color="3F5235", end_color="3F5235", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    n_cols = len(headers)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[4].height = 28
    for idx, row in enumerate(data["rows"], start=1):
        r_excel = 4 + idx
        ws.cell(row=r_excel, column=1, value=idx).border = border
        ws.cell(row=r_excel, column=2, value=row["name"]).border = border
        if row.get("qty_in_pkg_unit") is not None:
            ws.cell(row=r_excel, column=3, value=round(row["qty_in_pkg_unit"], 3)).border = border
            ws.cell(row=r_excel, column=4, value=row.get("pkg_unit") or "").border = border
        else:
            ws.cell(row=r_excel, column=3, value=round(row["quantity"], 3)).border = border
            ws.cell(row=r_excel, column=4, value=row["unit"]).border = border
        ws.cell(row=r_excel, column=5, value=row.get("opakowanie") or "—").border = border
        ws.cell(row=r_excel, column=6,
                value=f"{row['pkg_qty']} {row.get('pkg_unit') or ''}" if row.get("pkg_qty") else "—").border = border
        if row.get("num_packages") is not None:
            cell_pkg = ws.cell(row=r_excel, column=7, value=row["num_packages"])
            cell_pkg.font = Font(bold=True, color="D4AF37")
        else:
            ws.cell(row=r_excel, column=7, value="—")
        ws.cell(row=r_excel, column=7).border = border
        for col in range(8, n_cols + 1):
            ws.cell(row=r_excel, column=col, value="").border = border
        ws.cell(row=r_excel, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r_excel, column=3).alignment = Alignment(horizontal="right")
        for col in (4, 5, 6, 7):
            ws.cell(row=r_excel, column=col).alignment = Alignment(horizontal="center")
    foot_row = 5 + len(data["rows"]) + 1
    ws.cell(row=foot_row, column=1,
            value="Prosimy o uzupełnienie kolumn: cena netto, wartość netto i uwagi.").font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=n_cols)
    widths = {"A": 14, "B": 40, "C": 12, "D": 11, "E": 14, "F": 14, "G": 12, "H": 18, "I": 16, "J": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf)
    safe_name = (data["wycena_name"] or "wycena").replace("/", "_").replace(" ", "_")[:50]
    filename = f"BOM_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename


@router.get("/wyceny/{wycena_id}/bom.pdf")
async def export_bom_pdf(
    wycena_id: str,
    subcategories: Optional[str] = Query(None),
    _user: dict = Depends(get_current_admin),
):
    data = await _build_bom(wycena_id)
    if subcategories:
        data = _filter_bom_rows(data, [s.strip() for s in subcategories.split(",") if s.strip()])
    content, filename = _generate_bom_pdf_bytes(data)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": _safe_content_disposition("attachment", filename)},
    )


def _generate_bom_pdf_bytes(data: dict):
    """Refaktor iter95ah: generuj PDF jako bytes."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    base_font, bold_font, _ = _register_pdf_fonts()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=12 * mm, leftMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    name_st = ParagraphStyle("name", parent=styles["Normal"], fontName=base_font, fontSize=7.5, leading=9)
    head_st = ParagraphStyle("head", parent=styles["Normal"], fontName=bold_font, fontSize=7.5,
                             leading=9, textColor=colors.white, alignment=1)
    elements = []
    from reportlab.platypus import Image as RLImage
    logo_path = _get_logo_path()
    title_cell = []
    title_cell.append(Paragraph(
        f"<b>Zapytanie ofertowe — Zestawienie materiałów</b><br/>"
        f"<font size=10>{data['wycena_name']}</font><br/>"
        f"<font size=8 color='#666666'>Data: {datetime.now().strftime('%Y-%m-%d %H:%M')}</font>",
        ParagraphStyle("htitle", parent=styles["Normal"], fontName=base_font, fontSize=13,
                       leading=15, textColor=colors.HexColor("#3F5235")),
    ))
    if logo_path:
        try:
            from PIL import Image as PILImage
            with PILImage.open(logo_path) as _li:
                _lw, _lh = _li.size
            _ratio = (_lh / _lw) if _lw else 1.0
            _w_mm = 18
            img = RLImage(logo_path, width=_w_mm * mm, height=_w_mm * _ratio * mm)
            head_tbl = Table([[img, title_cell[0]]], colWidths=[22 * mm, 164 * mm])
        except Exception:
            head_tbl = Table([[title_cell[0]]], colWidths=[186 * mm])
    else:
        head_tbl = Table([[title_cell[0]]], colWidths=[186 * mm])
    head_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(head_tbl)
    elements.append(Spacer(1, 6 * mm))
    table_data = [[
        Paragraph("L.p.", head_st),
        Paragraph("Nazwa materiału", head_st),
        Paragraph("Ilość", head_st),
        Paragraph("Jedn.", head_st),
        Paragraph("Opak.", head_st),
        Paragraph("Wielk.<br/>opak.", head_st),
        Paragraph("Liczba<br/>opak.", head_st),
        Paragraph("Cena netto<br/>za opak.", head_st),
        Paragraph("Uwagi", head_st),
    ]]
    for idx, row in enumerate(data["rows"], start=1):
        if row.get("qty_in_pkg_unit") is not None:
            qty_str = f"{row['qty_in_pkg_unit']:.3f}".replace(".", ",")
            unit_str = row.get("pkg_unit") or ""
        else:
            qty_str = f"{row['quantity']:.3f}".replace(".", ",")
            unit_str = row["unit"]
        pkg_size = f"{row['pkg_qty']} {row.get('pkg_unit') or ''}" if row.get("pkg_qty") else "—"
        num_pkg = str(row["num_packages"]) if row.get("num_packages") is not None else "—"
        table_data.append([
            str(idx), Paragraph(row["name"], name_st), qty_str, unit_str,
            row.get("opakowanie") or "—", pkg_size, num_pkg, "", "",
        ])
    tbl = Table(table_data, colWidths=[10 * mm, 52 * mm, 14 * mm, 11 * mm, 20 * mm,
                                        20 * mm, 18 * mm, 22 * mm, 19 * mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), base_font, 8),
        ("FONT", (0, 0), (-1, 0), bold_font, 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3F5235")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("ALIGN", (3, 0), (3, -1), "CENTER"),
        ("ALIGN", (4, 0), (6, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
        ("TEXTCOLOR", (6, 1), (6, -1), colors.HexColor("#B8860B")),
        ("FONT", (6, 1), (6, -1), bold_font, 8),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 8 * mm))
    elements.append(Paragraph(
        "Prosimy o uzupełnienie kolumn: <b>cena netto za opakowanie</b> oraz <b>uwagi</b>. "
        "Liczby opakowań zostały zaokrąglone w górę do pełnych jednostek (paleta / wiaderko / rolka).",
        ParagraphStyle("foot", parent=styles["Normal"], fontName=base_font, fontSize=8, textColor=colors.grey)
    ))
    doc.build(elements)
    safe_name = (data["wycena_name"] or "wycena").replace("/", "_").replace(" ", "_")[:50]
    filename = f"BOM_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buf.getvalue(), filename


# ============================================================
# Suppliers + wysylka BOM emailem (iter95ai)
# ============================================================
class SupplierCreate(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    branze: Optional[str] = None
    notes: Optional[str] = None


class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    branze: Optional[str] = None
    notes: Optional[str] = None


@router.get("/wyceny/suppliers")
async def list_suppliers(_user: dict = Depends(get_current_admin)):
    rows = await db.wyceny_suppliers.find({}, {"_id": 0}).sort("name", 1).to_list(length=None)
    return {"rows": rows}


@router.post("/wyceny/suppliers")
async def create_supplier(payload: SupplierCreate, _user: dict = Depends(get_current_admin)):
    sid = str(uuid.uuid4())
    doc = {
        "id": sid, "name": payload.name, "email": payload.email,
        "phone": payload.phone or "",
        "branze": payload.branze or "", "notes": payload.notes or "",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.wyceny_suppliers.insert_one(doc)
    return {"id": sid, "ok": True}


@router.patch("/wyceny/suppliers/{sid}")
async def update_supplier(sid: str, payload: SupplierUpdate, _user: dict = Depends(get_current_admin)):
    update = payload.dict(exclude_unset=True)
    if not update:
        return {"ok": True}
    r = await db.wyceny_suppliers.update_one({"id": sid}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(404, "Hurtownia nie istnieje")
    return {"ok": True}


@router.delete("/wyceny/suppliers/{sid}")
async def delete_supplier(sid: str, _user: dict = Depends(get_current_admin)):
    r = await db.wyceny_suppliers.delete_one({"id": sid})
    if r.deleted_count == 0:
        raise HTTPException(404, "Hurtownia nie istnieje")
    return {"ok": True}


class SendBomRequest(BaseModel):
    to_email: EmailStr
    subject: Optional[str] = None
    body: Optional[str] = None
    supplier_id: Optional[str] = None
    subcategories: Optional[List[str]] = None


@router.post("/wyceny/{wycena_id}/bom/send")
async def send_bom_email(wycena_id: str, payload: SendBomRequest, _user: dict = Depends(get_current_admin)):
    """Wyslij zapytanie ofertowe (BOM) na maila hurtownika z zalacznikami PDF + XLSX."""
    import base64
    api_key = os.environ.get("RESEND_API_KEY")
    if not api_key:
        raise HTTPException(503, "Resend nie skonfigurowany (brak RESEND_API_KEY)")
    from_addr = "FeGrro <biuro@fegrro.pl>"
    try:
        import httpx
    except ImportError:
        raise HTTPException(500, "httpx not installed")
    data = await _build_bom(wycena_id)
    if payload.subcategories:
        data = _filter_bom_rows(data, payload.subcategories)
    if not data.get("rows"):
        raise HTTPException(400, "Wycena nie zawiera materiałów do wysłania")
    xlsx_bytes, xlsx_name = _generate_bom_xlsx_bytes(data)
    pdf_bytes, pdf_name = _generate_bom_pdf_bytes(data)
    wycena_name = data["wycena_name"] or "—"
    subject = payload.subject or f"Zapytanie ofertowe — {wycena_name}"
    body_text = payload.body or (
        f"Dzień dobry,\n\n"
        f"W załączeniu przesyłam zestawienie materiałów do wyceny: \u201E{wycena_name}\u201D.\n"
        f"Proszę o przygotowanie oferty cenowej (cena netto za opakowanie, termin dostawy).\n\n"
        f"Termin oferty: 7 dni.\n\n"
        f"Pozdrawiam,\nFeGrro"
    )
    body_html = "<p>" + body_text.replace("\n", "<br>") + "</p>"
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            resp = await client.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={
                    "from": from_addr,
                    "to": [payload.to_email],
                    "reply_to": ["biuro@fegrro.pl"],
                    "subject": subject,
                    "html": body_html,
                    "text": body_text,
                    "attachments": [
                        {"filename": xlsx_name, "content": base64.b64encode(xlsx_bytes).decode("ascii")},
                        {"filename": pdf_name, "content": base64.b64encode(pdf_bytes).decode("ascii")},
                    ],
                },
            )
        if resp.status_code >= 300:
            logger.warning(f"Resend BOM email returned {resp.status_code}: {resp.text}")
            raise HTTPException(502, f"Resend: {resp.status_code} {resp.text}")
        result = resp.json()
    except httpx.HTTPError as e:
        raise HTTPException(502, f"Network error: {e}")
    history = {
        "id": str(uuid.uuid4()),
        "wycena_id": wycena_id,
        "to_email": payload.to_email,
        "supplier_id": payload.supplier_id,
        "subject": subject,
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "message_id": result.get("id"),
    }
    await db.wyceny_bom_history.insert_one(history)
    return {"ok": True, "message_id": result.get("id")}


@router.get("/wyceny/{wycena_id}/bom/history")
async def get_bom_history(wycena_id: str, _user: dict = Depends(get_current_admin)):
    rows = await db.wyceny_bom_history.find(
        {"wycena_id": wycena_id}, {"_id": 0}
    ).sort("sent_at", -1).to_list(length=100)
    return {"rows": rows}
