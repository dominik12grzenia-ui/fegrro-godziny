"""
Budzetowanie budow - moduł do planowania i kontroli realizacji budowy.

Trzy poziomy danych:
  1. budget_lines       - pozycje budzetu (kategoria, plan netto, kaucje, jednostka)
  2. budget_progress    - zaawansowanie % per pozycja per miesiac (protokol)
  3. budget_tasks       - zadania harmonogramu (data start, koniec, %)

Integracja z modulem finansowym:
  - finance_zapisy.budget_line_id (opcjonalny FK) - mozliwosc przypisania zapisu
    do konkretnej pozycji budzetowej. Wykonanie liczone z sumy netto/brutto
    zapisow z dopasowanym budget_line_id.
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from datetime import datetime
from typing import Optional, List
from pydantic import BaseModel, Field
from io import BytesIO
import uuid
import logging
from urllib.parse import quote

from database import db
from auth import get_current_admin

logger = logging.getLogger(__name__)
router = APIRouter()


# ============== MODELS ==============

class BudgetCategoryCreate(BaseModel):
    budowa_id: str
    name: str
    order: int = 0


class BudgetStageCreate(BaseModel):
    budowa_id: str
    name: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    order: int = 0


class BudgetPositionCreate(BaseModel):
    budowa_id: str
    stage_id: str = Field(..., description="ID etapu - WYMAGANE")
    name: str = Field(..., description="Nazwa pozycji np. 'Wykonanie chodnika'")
    notes: Optional[str] = None
    order: int = 0
    include_in_protocol: bool = True


class BudgetPositionUpdate(BaseModel):
    stage_id: Optional[str] = None
    name: Optional[str] = None
    notes: Optional[str] = None
    order: Optional[int] = None
    include_in_protocol: Optional[bool] = None


class BudgetLineCreate(BaseModel):
    budowa_id: str
    category: str = Field(..., description="Kategoria np. 'Beton', 'Stal', 'Robocizna'")
    name: str = Field(..., description="Konkretna pozycja np. 'Beton C8/10 chudziaki'")
    type: str = Field("materials", description="materials | labor | equipment")
    unit: Optional[str] = Field(None, description="Jednostka np. m3, t, mb")
    quantity: float = 0.0
    unit_price_netto: float = 0.0
    plan_netto: Optional[float] = None  # jezeli puste, liczymy quantity * unit_price_netto
    kaucja_gir_pct: Optional[float] = None  # gdy None - dziedziczy z finance_budowy
    kaucja_dw_pct: Optional[float] = None
    stage_id: Optional[str] = None  # FK do budget_stages
    position_id: Optional[str] = None  # FK do budget_positions - wymagane dla nie-przychodowych w nowym modelu
    parent_id: Optional[str] = None  # FK do innej pozycji - jezeli ustawione, to skladowa kosztowa
    is_income: bool = False  # True dla pozycji przychodowych
    forecast_cost: Optional[float] = None  # kolumna L w kosztorysie (Koszt prognozowany) - reczna wartosc
    forecast_note: Optional[str] = None  # notatka widoczna pod kursorem na komorce L
    notes: Optional[str] = None
    order: int = 0


class BudgetLineUpdate(BaseModel):
    category: Optional[str] = None
    name: Optional[str] = None
    type: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = None
    unit_price_netto: Optional[float] = None
    plan_netto: Optional[float] = None
    kaucja_gir_pct: Optional[float] = None
    kaucja_dw_pct: Optional[float] = None
    stage_id: Optional[str] = None
    position_id: Optional[str] = None
    parent_id: Optional[str] = None
    is_income: Optional[bool] = None
    forecast_cost: Optional[float] = None
    forecast_note: Optional[str] = None
    notes: Optional[str] = None
    order: Optional[int] = None


class BudgetProgressSet(BaseModel):
    year: int
    month: int = Field(..., ge=1, le=12)
    progress_pct: float = Field(..., ge=0, le=100)
    notes: Optional[str] = None


class BudgetTaskCreate(BaseModel):
    budowa_id: str
    name: str
    start_date: str  # YYYY-MM-DD
    end_date: str  # YYYY-MM-DD
    progress_pct: float = Field(0.0, ge=0, le=100)
    color: Optional[str] = None
    notes: Optional[str] = None
    dependencies: List[str] = []  # lista task_id zaleznosci
    order: int = 0


class BudgetTaskUpdate(BaseModel):
    name: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    progress_pct: Optional[float] = None
    color: Optional[str] = None
    notes: Optional[str] = None
    dependencies: Optional[List[str]] = None
    order: Optional[int] = None


# ============== HELPERS ==============

def _compute_plan(line: dict) -> float:
    """Zwraca plan_netto. Jezeli zdefiniowany jawnie - uzywa go, inaczej liczy."""
    plan = line.get("plan_netto")
    if plan is not None:
        return float(plan)
    return round(float(line.get("quantity") or 0) * float(line.get("unit_price_netto") or 0), 2)


# ============== ENDPOINTS: LINES ==============

@router.get("/budget/budowy")
async def list_budowy_budgets(_user: dict = Depends(get_current_admin)):
    """Lista budow z podsumowaniem budzetu (czy ma pozycje, suma planu i wykonania)."""
    budowy = await db.finance_budowy.find({}, {"_id": 0, "id": 1, "name": 1, "code": 1}).to_list(length=1000)
    result = []
    for b in budowy:
        bid = b["id"]
        lines = await db.budget_lines.find({"budowa_id": bid}, {"_id": 0}).to_list(length=2000)
        # Wykonanie - suma netto z finance_zapisy gdzie budget_line_id IN lines
        line_ids = [ln["id"] for ln in lines]
        execution_netto = 0.0
        execution_brutto = 0.0
        if line_ids:
            pipe = [
                {"$match": {"budget_line_id": {"$in": line_ids}}},
                {"$group": {"_id": None,
                             "netto": {"$sum": "$netto"},
                             "brutto": {"$sum": "$brutto"}}},
            ]
            async for r in db.finance_zapisy.aggregate(pipe):
                execution_netto = float(r.get("netto") or 0)
                execution_brutto = float(r.get("brutto") or 0)

        plan_netto = sum(_compute_plan(ln) for ln in lines if not ln.get("is_income"))
        plan_income = sum(_compute_plan(ln) for ln in lines if ln.get("is_income"))
        # Zadania harmonogramu
        tasks_count = await db.budget_tasks.count_documents({"budowa_id": bid})
        result.append({
            "budowa_id": bid,
            "name": b.get("name") or "",
            "code": b.get("code") or "",
            "lines_count": len(lines),
            "tasks_count": tasks_count,
            "plan_costs_netto": round(plan_netto, 2),
            "plan_income_netto": round(plan_income, 2),
            "execution_netto": round(execution_netto, 2),
            "execution_brutto": round(execution_brutto, 2),
        })
    return {"rows": result}


# ============== BUDOWA INFO (defaulty dla modali) ==============

FEGRRO_WYKONAWCA = "FEGRRO SP. Z O.O.\nNIP: 589-206-61-74"


def _resolve_wykonawca(budowa: dict) -> str:
    """Zwraca tekst wykonawcy: ten wpisany na budowie (jeśli jest) albo domyślny FEGRRO."""
    w = (budowa.get("wykonawca") or "").strip()
    return w if w else FEGRRO_WYKONAWCA


@router.get("/budget/{budowa_id}/budowa-info")
async def get_budowa_info(budowa_id: str, _user: dict = Depends(get_current_admin)):
    """Zwraca pelne defaultowe dane budowy z finance_budowy:
    nazwa, kaucje, nr umowy, kontrahent. Uzywane jako defaulty w modalach Budzetu.
    Wykonawca jest stalą firmową (FeGrro Sp. z o.o.) i NIE jest tu zwracany - generatorzy uzywaja FEGRRO_WYKONAWCA."""
    b = await db.finance_budowy.find_one({"id": budowa_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Budowa nie istnieje")
    return {
        "id": b["id"],
        "name": b.get("name", ""),
        "code": b.get("code", ""),
        "kaucja_gir_pct": float(b.get("kaucja_gir_pct") or 0),
        "kaucja_dw_pct": float(b.get("kaucja_dw_pct") or 0),
        "koszt_budowy_pct": float(b.get("koszt_budowy_pct") or 0),
        "umowa_nr": b.get("umowa_nr", "") or "",
        "umowa_data": b.get("umowa_data", "") or "",
        "zamawiajacy": b.get("zamawiajacy", "") or "",
    }


@router.get("/budget/{budowa_id}/lines")
async def get_lines(budowa_id: str, _user: dict = Depends(get_current_admin)):
    """Pozycje budzetowe danej budowy + wyliczone wykonanie.
    Kaucje GIR/DW: pierwszenstwo ma wartosc z linii (override), fallback z finance_budowy.
    """
    budowa = await db.finance_budowy.find_one({"id": budowa_id}, {"_id": 0, "kaucja_gir_pct": 1, "kaucja_dw_pct": 1, "koszt_budowy_pct": 1}) or {}
    budowa_gir = float(budowa.get("kaucja_gir_pct") or 0)
    budowa_dw = float(budowa.get("kaucja_dw_pct") or 0)
    budowa_kb = float(budowa.get("koszt_budowy_pct") or 0)

    lines = await db.budget_lines.find(
        {"budowa_id": budowa_id}, {"_id": 0}
    ).sort([("order", 1), ("created_at", 1)]).to_list(length=2000)

    # Wykonanie per linia (agregat z finance_zapisy)
    line_ids = [ln["id"] for ln in lines]
    exec_map: dict = {}
    if line_ids:
        pipe = [
            {"$match": {"budget_line_id": {"$in": line_ids}}},
            {"$group": {"_id": "$budget_line_id",
                         "netto": {"$sum": "$netto"},
                         "brutto": {"$sum": "$brutto"},
                         "count": {"$sum": 1}}},
        ]
        async for r in db.finance_zapisy.aggregate(pipe):
            exec_map[r["_id"]] = {
                "netto": round(float(r["netto"] or 0), 2),
                "brutto": round(float(r["brutto"] or 0), 2),
                "count": r["count"],
            }

    for ln in lines:
        # Migracja: stare pozycje bez 'type' = materials
        if not ln.get("type"):
            ln["type"] = "materials"
        plan = _compute_plan(ln)
        ln["plan_netto_computed"] = round(plan, 2)
        ex = exec_map.get(ln["id"], {"netto": 0.0, "brutto": 0.0, "count": 0})
        ln["execution_netto"] = ex["netto"]
        ln["execution_brutto"] = ex["brutto"]
        ln["execution_count"] = ex["count"]
        ln["remaining_netto"] = round(plan - ex["netto"], 2)
        ln["progress_pct"] = round((ex["netto"] / plan) * 100, 1) if plan > 0 else 0.0
        # Efektywne kaucje: line override → fallback budowa
        eff_gir = ln.get("kaucja_gir_pct")
        eff_dw = ln.get("kaucja_dw_pct")
        eff_gir = float(eff_gir) if eff_gir is not None else budowa_gir
        eff_dw = float(eff_dw) if eff_dw is not None else budowa_dw
        ln["effective_kaucja_gir_pct"] = eff_gir
        ln["effective_kaucja_dw_pct"] = eff_dw
        ln["kaucja_gir_amount"] = round(plan * eff_gir / 100, 2)
        ln["kaucja_dw_amount"] = round(plan * eff_dw / 100, 2)
        # Koszt budowy: per linia identycznie jak kaucje (zaciag z budowy)
        ln["effective_koszt_budowy_pct"] = budowa_kb
        ln["koszt_budowy_amount"] = round(plan * budowa_kb / 100, 2)

    return {"rows": lines}


@router.get("/budget/{budowa_id}/options-flat")
async def get_budget_options_flat(budowa_id: str, _user: dict = Depends(get_current_admin)):
    """Splaszczona hierarchia budzetu dla dropdownu w Zapisach (Finance).

    Zwraca opcje wyboru w kolejnosci: Etap -> Pozycja -> Slot R/M/S -> Skladowa.
    Kazda opcja ma `id` (= budget_lines.id), `code` (np. "101.R" lub "101.M.2"),
    `label` (czytelna etykieta), `stage_name`, `position_name`, `type`, `level`.
    """
    stages = await db.budget_stages.find({"budowa_id": budowa_id}, {"_id": 0}).sort([("order", 1), ("created_at", 1)]).to_list(length=500)
    positions = await db.budget_positions.find({"budowa_id": budowa_id}, {"_id": 0}).sort([("order", 1), ("created_at", 1)]).to_list(length=2000)
    lines = await db.budget_lines.find({"budowa_id": budowa_id}, {"_id": 0}).sort([("order", 1), ("created_at", 1)]).to_list(length=5000)

    pos_by_stage: dict = {}
    for p in positions:
        pos_by_stage.setdefault(p.get("stage_id"), []).append(p)

    slots_by_pos: dict = {}  # position_id -> {type: slot_line}
    subs_by_parent: dict = {}  # parent_id -> [sub_lines]
    for ln in lines:
        if ln.get("parent_id"):
            subs_by_parent.setdefault(ln["parent_id"], []).append(ln)
        elif ln.get("position_id"):
            slots_by_pos.setdefault(ln["position_id"], {})[ln.get("type") or "materials"] = ln

    type_letter = {"equipment": "S", "labor": "R", "materials": "M"}
    type_pl = {"equipment": "sprzęt", "labor": "robocizna", "materials": "Materiał"}
    type_order = ["equipment", "labor", "materials"]

    options: list = []
    for stage in stages:
        positions_for_stage = pos_by_stage.get(stage["id"], [])
        for p_idx, pos in enumerate(positions_for_stage):
            pos_code = f"{100 + p_idx + 1}"  # 101, 102, ...
            slots = slots_by_pos.get(pos["id"], {})
            for t in type_order:
                slot = slots.get(t)
                if not slot:
                    continue
                slot_code = f"{pos_code}.{type_letter[t]}"
                options.append({
                    "id": slot["id"],
                    "code": slot_code,
                    "label": f"{slot_code} · {pos['name']} ({type_pl[t]})",
                    "stage_name": stage["name"],
                    "position_name": pos["name"],
                    "type": t,
                    "level": "slot",
                })
                subs = subs_by_parent.get(slot["id"], [])
                for sub_idx, sub in enumerate(subs):
                    sub_code = f"{slot_code}.{sub_idx + 1}"
                    options.append({
                        "id": sub["id"],
                        "code": sub_code,
                        "label": f"    {sub_code} · {sub.get('name') or ''}",
                        "stage_name": stage["name"],
                        "position_name": pos["name"],
                        "parent_slot_id": slot["id"],
                        "type": t,
                        "level": "sub",
                    })
    return {"options": options}


# ============== CATEGORIES ==============

@router.get("/budget/{budowa_id}/categories")
async def list_categories(budowa_id: str, _user: dict = Depends(get_current_admin)):
    rows = await db.budget_categories.find(
        {"budowa_id": budowa_id}, {"_id": 0}
    ).sort([("order", 1), ("name", 1)]).to_list(length=500)
    return {"rows": rows}


@router.post("/budget/categories")
async def create_category(payload: BudgetCategoryCreate, current_user: dict = Depends(get_current_admin)):
    if not payload.name.strip():
        raise HTTPException(400, "Nazwa kategorii jest wymagana")
    doc = {
        "id": str(uuid.uuid4()),
        "budowa_id": payload.budowa_id,
        "name": payload.name.strip(),
        "order": payload.order,
        "created_at": datetime.now().isoformat(),
        "created_by": current_user["sub"],
    }
    await db.budget_categories.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.delete("/budget/categories/{category_id}")
async def delete_category(category_id: str, _user: dict = Depends(get_current_admin)):
    res = await db.budget_categories.delete_one({"id": category_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Kategoria nie istnieje")
    return {"ok": True}


# ============== STAGES (Etapy budowy) ==============

@router.get("/budget/{budowa_id}/stages")
async def list_stages(budowa_id: str, _user: dict = Depends(get_current_admin)):
    rows = await db.budget_stages.find(
        {"budowa_id": budowa_id}, {"_id": 0}
    ).sort([("order", 1), ("start_date", 1), ("name", 1)]).to_list(length=500)
    # Dolicz liczbe pozycji per etap
    for s in rows:
        s["lines_count"] = await db.budget_lines.count_documents({"stage_id": s["id"]})
    return {"rows": rows}


@router.post("/budget/stages")
async def create_stage(payload: BudgetStageCreate, current_user: dict = Depends(get_current_admin)):
    if not payload.name.strip():
        raise HTTPException(400, "Nazwa etapu jest wymagana")
    doc = {
        "id": str(uuid.uuid4()),
        "budowa_id": payload.budowa_id,
        "name": payload.name.strip(),
        "start_date": payload.start_date,
        "end_date": payload.end_date,
        "order": payload.order,
        "created_at": datetime.now().isoformat(),
        "created_by": current_user["sub"],
    }
    await db.budget_stages.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.patch("/budget/stages/{stage_id}")
async def update_stage(stage_id: str, payload: BudgetStageCreate,
                        current_user: dict = Depends(get_current_admin)):
    existing = await db.budget_stages.find_one({"id": stage_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Etap nie istnieje")
    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None and k != "budowa_id"}
    updates["updated_at"] = datetime.now().isoformat()
    updates["updated_by"] = current_user["sub"]
    await db.budget_stages.update_one({"id": stage_id}, {"$set": updates})
    new_doc = await db.budget_stages.find_one({"id": stage_id}, {"_id": 0})
    return new_doc


@router.delete("/budget/{budowa_id}/wipe")
async def wipe_budget(budowa_id: str, _user: dict = Depends(get_current_admin)):
    """Czysci CALY budzet dla danej budowy:
    - usuwa wszystkie budget_lines (lacznie z sierotami sprzed iter68)
    - usuwa wszystkie budget_positions
    - usuwa wszystkie budget_progress
    - usuwa wszystkie budget_stages
    Czysci tez powiazania w finance_zapisy (budget_line_id -> usuniete).
    Etapy zostaja - jezeli chcesz je usunac, kasuj per etap.
    """
    budowa = await db.finance_budowy.find_one({"id": budowa_id}, {"_id": 0, "id": 1})
    if not budowa:
        raise HTTPException(404, "Budowa nie istnieje")
    # Linia ids tej budowy
    lines = await db.budget_lines.find({"budowa_id": budowa_id}, {"_id": 0, "id": 1}).to_list(length=10000)
    line_ids = [line["id"] for line in lines]
    if line_ids:
        await db.budget_progress.delete_many({"budget_line_id": {"$in": line_ids}})
        await db.finance_zapisy.update_many({"budget_line_id": {"$in": line_ids}}, {"$unset": {"budget_line_id": ""}})
    # Pozycje + ich progress (po position_id)
    positions = await db.budget_positions.find({"budowa_id": budowa_id}, {"_id": 0, "id": 1}).to_list(length=2000)
    pos_ids = [p["id"] for p in positions]
    if pos_ids:
        await db.budget_progress.delete_many({"position_id": {"$in": pos_ids}})
    r1 = await db.budget_lines.delete_many({"budowa_id": budowa_id})
    r2 = await db.budget_positions.delete_many({"budowa_id": budowa_id})
    return {
        "ok": True,
        "deleted_lines": r1.deleted_count,
        "deleted_positions": r2.deleted_count,
    }


@router.delete("/budget/stages/{stage_id}")
async def delete_stage(stage_id: str, _user: dict = Depends(get_current_admin)):
    # Odlinkowanie pozycji od etapu (nie usuwa pozycji!)
    await db.budget_lines.update_many({"stage_id": stage_id}, {"$unset": {"stage_id": ""}})
    res = await db.budget_stages.delete_one({"id": stage_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Etap nie istnieje")
    return {"ok": True}


# ============== POZYCJE KOSZTORYSOWE (Position) ==============

@router.get("/budget/{budowa_id}/positions")
async def list_positions(budowa_id: str, _user: dict = Depends(get_current_admin)):
    rows = await db.budget_positions.find({"budowa_id": budowa_id}, {"_id": 0}).sort([("stage_id", 1), ("order", 1), ("created_at", 1)]).to_list(length=2000)
    return {"rows": rows}


@router.post("/budget/positions")
async def create_position(payload: BudgetPositionCreate, current_user: dict = Depends(get_current_admin)):
    """Tworzy tylko pozycje kosztorysowa (BEZ auto-slotow). Podpozycje (R/M/S) admin dodaje recznie przez POST /budget/lines z position_id+type."""
    stage = await db.budget_stages.find_one({"id": payload.stage_id}, {"_id": 0})
    if not stage:
        raise HTTPException(400, "Etap nie istnieje - pozycja musi byc przypisana do etapu")
    if stage.get("budowa_id") != payload.budowa_id:
        raise HTTPException(400, "Etap nalezy do innej budowy")
    pos_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    pos_doc = {
        "id": pos_id,
        "budowa_id": payload.budowa_id,
        "stage_id": payload.stage_id,
        "name": payload.name,
        "notes": payload.notes,
        "order": payload.order,
        "include_in_protocol": payload.include_in_protocol,
        "created_at": now,
        "created_by": current_user["sub"],
    }
    await db.budget_positions.insert_one(pos_doc)
    pos_doc.pop("_id", None)
    return pos_doc


@router.patch("/budget/positions/{position_id}")
async def update_position(position_id: str, payload: BudgetPositionUpdate, current_user: dict = Depends(get_current_admin)):
    existing = await db.budget_positions.find_one({"id": position_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Pozycja nie istnieje")
    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    if updates:
        updates["updated_at"] = datetime.now().isoformat()
        updates["updated_by"] = current_user["sub"]
        await db.budget_positions.update_one({"id": position_id}, {"$set": updates})
        # Jezeli zmienila sie nazwa lub etap - sync na slotach (3 wiersze)
        sync = {}
        if "name" in updates:
            sync["name"] = updates["name"]
        if "stage_id" in updates:
            sync["stage_id"] = updates["stage_id"]
        if sync:
            await db.budget_lines.update_many({"position_id": position_id, "parent_id": None}, {"$set": sync})
    return await db.budget_positions.find_one({"id": position_id}, {"_id": 0})


@router.delete("/budget/positions/{position_id}")
async def delete_position(position_id: str, _user: dict = Depends(get_current_admin)):
    """Usuwa pozycje i wszystkie powiazane linie (sloty + skladowe)."""
    # Wszystkie linie z position_id (sloty + skladowe slotow)
    lines = await db.budget_lines.find({"position_id": position_id}, {"_id": 0, "id": 1}).to_list(length=2000)
    line_ids = [line["id"] for line in lines]
    if line_ids:
        await db.budget_progress.delete_many({"budget_line_id": {"$in": line_ids}})
        await db.finance_zapisy.update_many({"budget_line_id": {"$in": line_ids}}, {"$unset": {"budget_line_id": ""}})
        await db.budget_lines.delete_many({"id": {"$in": line_ids}})
    res = await db.budget_positions.delete_one({"id": position_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Pozycja nie istnieje")
    return {"ok": True, "deleted_lines": len(line_ids)}


# ============== END POZYCJE KOSZTORYSOWE ==============


@router.post("/budget/lines")
async def create_line(payload: BudgetLineCreate, current_user: dict = Depends(get_current_admin)):
    # Walidacja parent_id: musi istniec, miec ta sama budowa+typ, nie moze sam byc skladowa (max 2 poziomy)
    inherited_position_id = payload.position_id
    inherited_stage_id = payload.stage_id
    if payload.parent_id:
        parent = await db.budget_lines.find_one({"id": payload.parent_id}, {"_id": 0})
        if not parent:
            raise HTTPException(404, "Pozycja nadrzedna nie istnieje")
        if parent.get("budowa_id") != payload.budowa_id:
            raise HTTPException(400, "Pozycja nadrzedna nalezy do innej budowy")
        if parent.get("parent_id"):
            raise HTTPException(400, "Skladowa nie moze byc dodana do innej skladowej (max 2 poziomy)")
        if (parent.get("type") or "materials") != (payload.type or "materials"):
            raise HTTPException(400, "Typ skladowej musi byc taki sam jak pozycji nadrzednej")
        # Skladowe dziedzicza position_id i stage_id z rodzica
        if not inherited_position_id:
            inherited_position_id = parent.get("position_id")
        if not inherited_stage_id:
            inherited_stage_id = parent.get("stage_id")
    line_id = str(uuid.uuid4())
    doc = {
        "id": line_id,
        "budowa_id": payload.budowa_id,
        "category": payload.category,
        "name": payload.name,
        "type": payload.type or "materials",
        "unit": payload.unit,
        "quantity": payload.quantity,
        "unit_price_netto": payload.unit_price_netto,
        "plan_netto": payload.plan_netto,
        "kaucja_gir_pct": payload.kaucja_gir_pct,  # None = dziedziczenie z budowy
        "kaucja_dw_pct": payload.kaucja_dw_pct,
        "stage_id": inherited_stage_id,
        "position_id": inherited_position_id,
        "parent_id": payload.parent_id,
        "is_income": payload.is_income,
        "forecast_cost": payload.forecast_cost,
        "forecast_note": payload.forecast_note,
        "notes": payload.notes,
        "order": payload.order,
        "created_at": datetime.now().isoformat(),
        "created_by": current_user["sub"],
    }
    await db.budget_lines.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.patch("/budget/lines/{line_id}")
async def update_line(line_id: str, payload: BudgetLineUpdate, current_user: dict = Depends(get_current_admin)):
    existing = await db.budget_lines.find_one({"id": line_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Pozycja nie istnieje")
    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    updates["updated_at"] = datetime.now().isoformat()
    updates["updated_by"] = current_user["sub"]
    await db.budget_lines.update_one({"id": line_id}, {"$set": updates})
    new_doc = await db.budget_lines.find_one({"id": line_id}, {"_id": 0})
    return new_doc


@router.delete("/budget/lines/{line_id}")
async def delete_line(line_id: str, _user: dict = Depends(get_current_admin)):
    # Znajdz dzieci (skladowe) - usun je rowniez wraz z ich powiazaniami
    children = await db.budget_lines.find({"parent_id": line_id}, {"_id": 0, "id": 1}).to_list(length=500)
    child_ids = [c["id"] for c in children]
    all_ids = [line_id] + child_ids
    # Usun progres przypisany do tej linii i dzieci
    await db.budget_progress.delete_many({"budget_line_id": {"$in": all_ids}})
    # Wyczysc budget_line_id w zapisach (nie usuwaj zapisow!)
    await db.finance_zapisy.update_many({"budget_line_id": {"$in": all_ids}}, {"$unset": {"budget_line_id": ""}})
    # Usun dzieci + glowny rekord
    if child_ids:
        await db.budget_lines.delete_many({"id": {"$in": child_ids}})
    res = await db.budget_lines.delete_one({"id": line_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Pozycja nie istnieje")
    return {"ok": True, "deleted_children": len(child_ids)}


# ============== ENDPOINTS: PROGRESS (PROTOKOL) ==============

@router.get("/budget/{budowa_id}/progress")
async def get_progress(budowa_id: str, year: Optional[int] = Query(None),
                       _user: dict = Depends(get_current_admin)):
    """Macierz zaawansowania: pozycja x miesiac.
    Zwraca slownik {line_id: {YYYY-MM: pct, ...}}."""
    line_ids_cursor = db.budget_lines.find({"budowa_id": budowa_id}, {"_id": 0, "id": 1})
    line_ids = [ln["id"] async for ln in line_ids_cursor]
    if not line_ids:
        return {"rows": [], "year": year}

    q: dict = {"budget_line_id": {"$in": line_ids}}
    if year:
        q["year"] = year
    items = await db.budget_progress.find(q, {"_id": 0}).to_list(length=20000)
    return {"rows": items, "year": year}


@router.post("/budget/lines/{line_id}/progress")
async def set_progress(line_id: str, payload: BudgetProgressSet,
                       current_user: dict = Depends(get_current_admin)):
    """Ustawia przerob miesieczny dla linii budzetowej.

    UWAGA: progress_pct to PRZEROB DANEGO MIESIACA (nie narastajaco!).
    Walidacja: suma wszystkich przerobow per pozycja nie moze przekroczyc 100%.
    """
    line = await db.budget_lines.find_one({"id": line_id}, {"_id": 0})
    if not line:
        raise HTTPException(404, "Pozycja nie istnieje")

    # Walidacja sumy - SUM(inne miesiace) + nowa <= 100
    existing_sum = 0.0
    async for p in db.budget_progress.find(
        {"budget_line_id": line_id,
         "$or": [{"year": {"$ne": payload.year}},
                 {"year": payload.year, "month": {"$ne": payload.month}}]},
        {"_id": 0, "progress_pct": 1},
    ):
        existing_sum += float(p.get("progress_pct") or 0)
    total_after = existing_sum + payload.progress_pct
    if total_after > 100.01:
        raise HTTPException(
            400,
            f"Przekroczono 100% realizacji pozycji. "
            f"Suma innych miesiecy: {existing_sum:.1f}%, "
            f"pozostalo do rozdysponowania: {max(0, 100 - existing_sum):.1f}%"
        )

    plan = _compute_plan(line)
    value_netto = round(plan * (payload.progress_pct / 100.0), 2)
    key = {"budget_line_id": line_id, "year": payload.year, "month": payload.month}
    update_set = {
        **key,
        "progress_pct": payload.progress_pct,
        "value_netto": value_netto,
        "notes": payload.notes,
        "updated_at": datetime.now().isoformat(),
        "updated_by": current_user["sub"],
    }
    existing = await db.budget_progress.find_one(key, {"_id": 0})
    if existing:
        await db.budget_progress.update_one(key, {"$set": update_set})
        update_set["id"] = existing["id"]
    else:
        update_set["id"] = str(uuid.uuid4())
        update_set["created_at"] = datetime.now().isoformat()
        await db.budget_progress.insert_one(update_set)
    update_set.pop("_id", None)
    return update_set


@router.post("/budget/positions/{position_id}/progress")
async def set_position_progress(position_id: str, payload: BudgetProgressSet,
                                current_user: dict = Depends(get_current_admin)):
    """Ustawia przerob miesieczny dla POZYCJI kosztorysowej (iter72+).
    Klucz: position_id (zamiast budget_line_id).
    Walidacja: suma wszystkich przerobow per pozycja nie moze przekroczyc 100%.
    """
    pos = await db.budget_positions.find_one({"id": position_id}, {"_id": 0})
    if not pos:
        raise HTTPException(404, "Pozycja nie istnieje")

    # Walidacja sumy
    existing_sum = 0.0
    async for p in db.budget_progress.find(
        {"position_id": position_id,
         "$or": [{"year": {"$ne": payload.year}},
                 {"year": payload.year, "month": {"$ne": payload.month}}]},
        {"_id": 0, "progress_pct": 1},
    ):
        existing_sum += float(p.get("progress_pct") or 0)
    total_after = existing_sum + payload.progress_pct
    if total_after > 100.01:
        raise HTTPException(
            400,
            f"Przekroczono 100% realizacji pozycji. "
            f"Suma innych miesiecy: {existing_sum:.1f}%, "
            f"pozostalo do rozdysponowania: {max(0, 100 - existing_sum):.1f}%"
        )

    # Plan pozycji = suma plan_netto wszystkich linii pod pozycja
    lines_pos = await db.budget_lines.find(
        {"budowa_id": pos["budowa_id"], "position_id": position_id, "is_income": {"$ne": True}},
        {"_id": 0},
    ).to_list(length=500)
    plan = sum(_compute_plan(line) for line in lines_pos)
    value_netto = round(plan * (payload.progress_pct / 100.0), 2)

    key = {"position_id": position_id, "year": payload.year, "month": payload.month}
    update_set = {
        **key,
        "progress_pct": payload.progress_pct,
        "value_netto": value_netto,
        "notes": payload.notes,
        "updated_at": datetime.now().isoformat(),
        "updated_by": current_user["sub"],
    }
    existing = await db.budget_progress.find_one(key, {"_id": 0})
    if existing:
        await db.budget_progress.update_one(key, {"$set": update_set})
        update_set["id"] = existing["id"]
    else:
        update_set["id"] = str(uuid.uuid4())
        update_set["created_at"] = datetime.now().isoformat()
        await db.budget_progress.insert_one(update_set)
    update_set.pop("_id", None)
    return update_set


# ============== ENDPOINTS: TASKS (HARMONOGRAM) ==============

@router.get("/budget/{budowa_id}/tasks")
async def get_tasks(budowa_id: str, _user: dict = Depends(get_current_admin)):
    rows = await db.budget_tasks.find(
        {"budowa_id": budowa_id}, {"_id": 0}
    ).sort([("order", 1), ("start_date", 1)]).to_list(length=1000)
    return {"rows": rows}


@router.post("/budget/tasks")
async def create_task(payload: BudgetTaskCreate, current_user: dict = Depends(get_current_admin)):
    task_id = str(uuid.uuid4())
    doc = {
        "id": task_id,
        "budowa_id": payload.budowa_id,
        "name": payload.name,
        "start_date": payload.start_date,
        "end_date": payload.end_date,
        "progress_pct": payload.progress_pct,
        "color": payload.color or "#D4AF37",
        "notes": payload.notes,
        "dependencies": payload.dependencies,
        "order": payload.order,
        "created_at": datetime.now().isoformat(),
        "created_by": current_user["sub"],
    }
    await db.budget_tasks.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.patch("/budget/tasks/{task_id}")
async def update_task(task_id: str, payload: BudgetTaskUpdate,
                       current_user: dict = Depends(get_current_admin)):
    existing = await db.budget_tasks.find_one({"id": task_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Zadanie nie istnieje")
    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    updates["updated_at"] = datetime.now().isoformat()
    updates["updated_by"] = current_user["sub"]
    await db.budget_tasks.update_one({"id": task_id}, {"$set": updates})
    new_doc = await db.budget_tasks.find_one({"id": task_id}, {"_id": 0})
    return new_doc


@router.delete("/budget/tasks/{task_id}")
async def delete_task(task_id: str, _user: dict = Depends(get_current_admin)):
    # Usun referencje z dependencies innych zadan
    await db.budget_tasks.update_many(
        {"dependencies": task_id},
        {"$pull": {"dependencies": task_id}},
    )
    res = await db.budget_tasks.delete_one({"id": task_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Zadanie nie istnieje")
    return {"ok": True}


# ============== PROTOKOL XLSX GENERATOR ==============

MONTH_NAMES_PL = ["Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec",
                  "Lipiec", "Sierpień", "Wrzesień", "Październik", "Listopad", "Grudzień"]


def _month_range(year: int, month: int):
    """Zwraca (data_od, data_do) w formacie YYYY-MM-DD dla danego miesiaca."""
    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"


async def _fetch_protokol_data(budowa_id: str, year: int, month: int):
    """Pobiera dane potrzebne do wygenerowania protokolu (xlsx lub pdf).

    Nowy model (iter72+): protokol operuje na POZYCJACH (BudgetPosition), nie na liniach.
    Pozycja jest zaciagana do protokolu tylko jezeli `include_in_protocol=True`.
    Wartosc planu pozycji = SUMA `plan_netto_computed` wszystkich podpozycji + skladowych.

    Konwencja:
      - budget_progress.progress_pct = PRZEROB DANEGO MIESIACA (nie narastajaco)
      - klucz: position_id (nie budget_line_id)
      - progress_curr[position_id] = przerob biezacego miesiaca
      - progress_prev[position_id] = SUMA przerobow PRZED biezacym miesiacem

    Zwraca tuple: (budowa, positions_with_plan, progress_curr, progress_prev, nr_protokolu, stages_map).
    `positions_with_plan` to lista dictow z polami: id, name, stage_id, plan_netto, quantity, unit, unit_price, category.
    stages_map: dict {stage_id: {id, name, order, start_date, end_date}}.
    """
    if month < 1 or month > 12:
        raise HTTPException(400, "Nieprawidlowy miesiac (1-12)")
    budowa = await db.finance_budowy.find_one({"id": budowa_id}, {"_id": 0})
    if not budowa:
        raise HTTPException(404, "Budowa nie istnieje")

    # Etapy budowy + mapa
    stages = await db.budget_stages.find({"budowa_id": budowa_id}, {"_id": 0}).sort([("order", 1), ("start_date", 1)]).to_list(length=500)
    stages_map = {s["id"]: s for s in stages}
    stage_order = {s["id"]: i for i, s in enumerate(stages)}
    stage_order[None] = len(stages) + 1

    # Pozycje (tylko include_in_protocol=True, domyslnie True)
    positions = await db.budget_positions.find(
        {"budowa_id": budowa_id, "include_in_protocol": {"$ne": False}},
        {"_id": 0},
    ).to_list(length=2000)

    # Wszystkie linie dla tej budowy zeby zsumowac plan per pozycja
    all_lines = await db.budget_lines.find(
        {"budowa_id": budowa_id, "is_income": {"$ne": True}}, {"_id": 0},
    ).to_list(length=10000)
    # Suma `plan_netto_computed` (lub fallback `quantity * unit_price_netto` lub `plan_netto`) per position_id
    plan_per_pos = {}
    qty_per_pos = {}  # przyklad jednostki - pierwsza spotkana
    for line in all_lines:
        pid = line.get("position_id")
        if not pid:
            continue
        plan = _compute_plan(line)
        plan_per_pos[pid] = plan_per_pos.get(pid, 0.0) + plan
        if pid not in qty_per_pos and line.get("quantity"):
            qty_per_pos[pid] = {
                "quantity": line.get("quantity"),
                "unit": line.get("unit"),
                "unit_price_netto": line.get("unit_price_netto"),
            }

    # Wzbogac pozycje o plan
    positions_with_plan = []
    for p in positions:
        pid = p["id"]
        plan = round(plan_per_pos.get(pid, 0.0), 2)
        qi = qty_per_pos.get(pid, {})
        positions_with_plan.append({
            "id": pid,
            "name": p.get("name", ""),
            "stage_id": p.get("stage_id"),
            "order": p.get("order") or 0,
            "created_at": p.get("created_at") or "",
            "plan_netto": plan,
            "quantity": qi.get("quantity") or 0,
            "unit": qi.get("unit") or "",
            "unit_price_netto": qi.get("unit_price_netto") or 0,
            "category": "",
        })

    # Sortuj: stage_order -> position.order -> created_at
    positions_with_plan.sort(key=lambda pp: (
        stage_order.get(pp.get("stage_id"), 999),
        pp.get("order") or 0,
        pp.get("created_at") or "",
    ))

    pos_ids = [p["id"] for p in positions_with_plan]
    progress_curr = {}
    progress_prev = {}
    if pos_ids:
        async for p in db.budget_progress.find(
            {"position_id": {"$in": pos_ids}, "year": year, "month": month},
            {"_id": 0, "position_id": 1, "progress_pct": 1},
        ):
            progress_curr[p["position_id"]] = float(p["progress_pct"])
        async for p in db.budget_progress.find(
            {"position_id": {"$in": pos_ids},
             "$or": [
                 {"year": {"$lt": year}},
                 {"year": year, "month": {"$lt": month}},
             ]},
            {"_id": 0, "position_id": 1, "progress_pct": 1},
        ):
            pid = p["position_id"]
            progress_prev[pid] = progress_prev.get(pid, 0.0) + float(p["progress_pct"])

    distinct_months = set()
    async for p in db.budget_progress.find(
        {"position_id": {"$in": pos_ids},
         "progress_pct": {"$gt": 0},
         "$or": [{"year": {"$lt": year}}, {"year": year, "month": {"$lt": month}}]},
        {"_id": 0, "year": 1, "month": 1},
    ):
        distinct_months.add((p["year"], p["month"]))
    nr = len(distinct_months) + 1
    return budowa, positions_with_plan, progress_curr, progress_prev, nr, stages_map


@router.get("/budget/{budowa_id}/protokol-view/{year}/{month}")
async def get_protokol_view(
    budowa_id: str,
    year: int,
    month: int,
    _user: dict = Depends(get_current_admin),
):
    """Dane do widoku Protokol w stylu Excel.
    Zwraca wiersze: ETAPY jako sekcje + pozycje (Pozycja Główna) z narastajaco / poprzedni / miesiac rozliczeniowy.
    UWAGA (iter72): wiersze sa pozycjami (BudgetPosition), NIE podpozycjami. ID wiersza = position_id.
    """
    budowa, positions, progress_curr, progress_prev, nr, stages_map = await _fetch_protokol_data(budowa_id, year, month)

    rows = []
    last_stage = "__init__"
    lp = 1
    sum_budzet = sum_narast = sum_prev = sum_miesiac = 0.0

    for pos in positions:
        sid = pos.get("stage_id")
        if sid != last_stage:
            stage = stages_map.get(sid)
            stage_name = (stage.get("name") if stage else "Bez etapu") or "Bez etapu"
            rows.append({"type": "section", "stage_id": sid, "stage_name": stage_name})
            last_stage = sid
        plan = pos["plan_netto"]
        miesiac_pct = progress_curr.get(pos["id"], 0.0)
        prev_pct = progress_prev.get(pos["id"], 0.0)
        narast_pct = min(100.0, prev_pct + miesiac_pct)
        narast_val = round(plan * narast_pct / 100, 2)
        prev_val = round(plan * prev_pct / 100, 2)
        miesiac_val = round(plan * miesiac_pct / 100, 2)
        rows.append({
            "type": "line",  # zachowujemy 'line' dla kompatybilnosci frontendu
            "id": pos["id"],
            "lp": lp,
            "category": pos.get("category") or "",
            "name": pos.get("name", ""),
            "unit": pos.get("unit") or "",
            "quantity": pos.get("quantity") or 0,
            "unit_price_netto": pos.get("unit_price_netto") or 0,
            "plan_netto": plan,
            "narast_val": narast_val,
            "narast_pct": narast_pct,
            "prev_val": prev_val,
            "prev_pct": prev_pct,
            "miesiac_val": miesiac_val,
            "miesiac_pct": miesiac_pct,
        })
        sum_budzet += plan
        sum_narast += narast_val
        sum_prev += prev_val
        sum_miesiac += miesiac_val
        lp += 1

    return {
        "nr": nr,
        "year": year,
        "month": month,
        "budowa_name": budowa.get("name", ""),
        "rows": rows,
        "totals": {
            "plan_netto": round(sum_budzet, 2),
            "narast_val": round(sum_narast, 2),
            "narast_pct": round((sum_narast / sum_budzet * 100) if sum_budzet else 0, 2),
            "prev_val": round(sum_prev, 2),
            "prev_pct": round((sum_prev / sum_budzet * 100) if sum_budzet else 0, 2),
            "miesiac_val": round(sum_miesiac, 2),
            "miesiac_pct": round((sum_miesiac / sum_budzet * 100) if sum_budzet else 0, 2),
        },
    }


@router.get("/budget/{budowa_id}/protokol/{year}/{month}/pdf")
async def generate_protokol_pdf(
    budowa_id: str,
    year: int,
    month: int,
    _user: dict = Depends(get_current_admin),
):
    """Generuje PDF z protokolem miesiecznym (uzywa reportlab)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from calendar import monthrange
    import os

    budowa, lines, progress_curr, progress_prev, nr, stages_map = await _fetch_protokol_data(budowa_id, year, month)

    # Rejestruj font wspierajacy polskie znaki (DejaVuSans z systemu)
    try:
        font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        font_bold = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont("PL", font_path))
            pdfmetrics.registerFont(TTFont("PL-Bold", font_bold))
            base_font = "PL"
            bold_font = "PL-Bold"
        else:
            base_font = "Helvetica"
            bold_font = "Helvetica-Bold"
    except Exception:
        base_font = "Helvetica"
        bold_font = "Helvetica-Bold"

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)

    olive = colors.HexColor("#4F6343")
    gray_light = colors.HexColor("#D9D9D9")
    white = colors.white

    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontName=bold_font,
                                  fontSize=13, leading=15, textColor=colors.black,
                                  alignment=0, spaceAfter=2)
    label_style = ParagraphStyle("label", fontName=bold_font, fontSize=9, textColor=colors.black, leading=11)
    val_style = ParagraphStyle("val", fontName=base_font, fontSize=9, textColor=colors.black, leading=11)

    # === HEADER (logo + tytul) ===
    logo_path = "/app/frontend/public/icon-512x512.png"
    header_cell_logo = ""
    if os.path.exists(logo_path):
        try:
            header_cell_logo = RLImage(logo_path, width=28 * mm, height=28 * mm)
        except Exception:
            pass

    header_data = [
        [header_cell_logo,
         Paragraph(f"<b>PROTOKÓŁ STANU ZAAWANSOWANIA ROBÓT NR &nbsp;&nbsp;&nbsp;{nr}</b>", title_style)],
        ["",
         Paragraph(f"<b>DO UMOWY</b> &nbsp;&nbsp; {budowa.get('umowa_nr', '')}", val_style)],
    ]
    header_table = Table(header_data, colWidths=[35 * mm, 240 * mm])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("SPAN", (0, 0), (0, 1)),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 6 * mm))

    # === POLA ===
    last_day = monthrange(year, month)[1]
    d_from = f"{year:04d}-{month:02d}-01"
    d_to = f"{last_day:02d}.{month:02d}.{year:04d}"
    info_data = [
        [Paragraph("<b>OKRES ROZLICZENIOWY:</b>", label_style),
         Paragraph(d_from, val_style),
         Paragraph("<b>DO</b>", label_style),
         Paragraph(d_to, val_style)],
        [Paragraph("<b>ZAMAWIAJĄCY:</b>", label_style),
         Paragraph(budowa.get("zamawiajacy", "") or "", val_style),
         "", ""],
        [Paragraph("<b>WYKONAWCA:</b>", label_style),
         Paragraph(_resolve_wykonawca(budowa).replace("\n", "<br/>"), val_style),
         "", ""],
    ]
    info_table = Table(info_data, colWidths=[50 * mm, 90 * mm, 15 * mm, 120 * mm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("SPAN", (1, 1), (3, 1)),
        ("SPAN", (1, 2), (3, 2)),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 5 * mm))

    # === TABELA ===
    def fmt_pl(v):
        return f"{v:,.2f}".replace(",", " ").replace(".", ",") if v else "0,00"

    # Dwa rzedy naglowka
    head_top = ["", "", "", "", "", "",
                "NARASTAJĄCO", "",
                "POPRZEDNI MIESIĄC", "",
                "MIESIĄC ROZLICZENIOWY", ""]
    head_bot = ["LP", "Robocizna", "Jd.", "Ilość", "Cena", "Wartość",
                "WARTOŚĆ", "%", "WARTOŚĆ", "%", "WARTOŚĆ", "%"]

    data = [head_top, head_bot]
    last_stage = "__init__"
    lp = 1
    sum_budzet = sum_narast = sum_prev = sum_miesiac = 0.0
    section_rows = []  # indeksy wierszy sekcji (etapow)

    for ln in lines:
        sid = ln.get("stage_id")
        if sid != last_stage:
            stage = stages_map.get(sid)
            stage_name = (stage.get("name") if stage else "Bez etapu") or "Bez etapu"
            section_rows.append(len(data))
            row = ["", stage_name.upper(), "", "", "", "", "", "", "", "", "", ""]
            data.append(row)
            last_stage = sid
        plan = ln.get("plan_netto", 0)  # iter72: pozycje maja juz policzony plan
        miesiac_pct = progress_curr.get(ln["id"], 0.0)
        prev_pct = progress_prev.get(ln["id"], 0.0)
        narast_pct = min(100.0, prev_pct + miesiac_pct)
        narast_val = round(plan * narast_pct / 100, 2)
        prev_val = round(plan * prev_pct / 100, 2)
        miesiac_val = round(plan * miesiac_pct / 100, 2)
        data.append([
            str(lp),
            ln.get("name", ""),
            ln.get("unit") or "",
            fmt_pl(ln.get("quantity") or 0),
            fmt_pl(ln.get("unit_price_netto") or 0),
            fmt_pl(plan),
            fmt_pl(narast_val),
            f"{narast_pct:.1f}%",
            fmt_pl(prev_val),
            f"{prev_pct:.1f}%",
            fmt_pl(miesiac_val),
            f"{miesiac_pct:.1f}%",
        ])
        sum_budzet += plan
        sum_narast += narast_val
        sum_prev += prev_val
        sum_miesiac += miesiac_val
        lp += 1

    # Wiersz RAZEM
    razem_idx = len(data)
    data.append([
        "", "RAZEM", "", "", "", fmt_pl(sum_budzet),
        fmt_pl(sum_narast), f"{(sum_narast/sum_budzet*100 if sum_budzet else 0):.0f}%",
        fmt_pl(sum_prev), f"{(sum_prev/sum_budzet*100 if sum_budzet else 0):.0f}%",
        fmt_pl(sum_miesiac), f"{(sum_miesiac/sum_budzet*100 if sum_budzet else 0):.0f}%",
    ])

    col_widths = [10 * mm, 80 * mm, 12 * mm, 18 * mm, 18 * mm, 22 * mm,
                  22 * mm, 12 * mm, 22 * mm, 12 * mm, 22 * mm, 12 * mm]
    table = Table(data, colWidths=col_widths, repeatRows=2)
    table_style = [
        ("FONTNAME", (0, 0), (-1, -1), base_font),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Naglowek
        ("BACKGROUND", (0, 0), (-1, 1), olive),
        ("TEXTCOLOR", (0, 0), (-1, 1), white),
        ("FONTNAME", (0, 0), (-1, 1), bold_font),
        ("FONTSIZE", (0, 0), (-1, 1), 7.5),
        ("ALIGN", (0, 0), (-1, 1), "CENTER"),
        # Span grup
        ("SPAN", (6, 0), (7, 0)),
        ("SPAN", (8, 0), (9, 0)),
        ("SPAN", (10, 0), (11, 0)),
        # Pierwsze 6 kolumn naglowka top puste, bez wypelnienia
        ("BACKGROUND", (0, 0), (5, 0), white),
        # Wiersz RAZEM
        ("BACKGROUND", (0, razem_idx), (-1, razem_idx), olive),
        ("TEXTCOLOR", (0, razem_idx), (-1, razem_idx), white),
        ("FONTNAME", (0, razem_idx), (-1, razem_idx), bold_font),
        # Wyrownania w body
        ("ALIGN", (0, 2), (0, -1), "CENTER"),  # LP
        ("ALIGN", (1, 2), (1, -1), "LEFT"),    # Robocizna
        ("ALIGN", (2, 2), (2, -1), "CENTER"),  # Jd
        ("ALIGN", (3, 2), (-1, -1), "RIGHT"),  # liczby
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    # Sekcje - szare tlo + tylko nazwa kategorii
    for sr in section_rows:
        table_style.append(("BACKGROUND", (0, sr), (-1, sr), gray_light))
        table_style.append(("FONTNAME", (0, sr), (-1, sr), bold_font))
        table_style.append(("SPAN", (1, sr), (11, sr)))
        table_style.append(("ALIGN", (1, sr), (1, sr), "LEFT"))

    table.setStyle(TableStyle(table_style))
    elements.append(table)
    elements.append(Spacer(1, 6 * mm))

    # === STOPKA ===
    today = datetime.now().strftime("%d.%m.%Y")
    foot_data = [[
        Paragraph(f"<b>DATA I MIEJSCE SPORZĄDZENIA PROTOKOŁU:</b> &nbsp;&nbsp; {today}", val_style),
        Paragraph(f"<b>DO ZAFAKTUROWANIA:</b> &nbsp;&nbsp; {fmt_pl(sum_miesiac)} zł &nbsp;&nbsp; <b>NETTO</b>", val_style),
    ]]
    foot_table = Table(foot_data, colWidths=[150 * mm, 125 * mm])
    foot_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    elements.append(foot_table)
    elements.append(Spacer(1, 18 * mm))

    sig_data = [
        ["..................................", "", "..................................."],
        [Paragraph("<b>Zamawiający</b><br/>(podpis i pieczęć)", val_style),
         "",
         Paragraph("<b>Wykonawca</b><br/>(podpis i pieczęć)", val_style)],
    ]
    sig_table = Table(sig_data, colWidths=[100 * mm, 75 * mm, 100 * mm])
    sig_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    buf.seek(0)

    safe_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in (budowa.get("name", "budowa")))
    filename = f"Protokol_{safe_name}_{year}-{month:02d}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="protokol.pdf"; filename*=UTF-8\'\'{quote(filename)}'
        },
    )


class BudowaContractData(BaseModel):
    umowa_nr: Optional[str] = None
    umowa_data: Optional[str] = None
    zamawiajacy: Optional[str] = None
    wykonawca: Optional[str] = None


@router.patch("/budget/{budowa_id}/contract")
async def update_contract_data(
    budowa_id: str, payload: BudowaContractData,
    current_user: dict = Depends(get_current_admin),
):
    """Aktualizuje pola umowy/zamawiajacego/wykonawcy w finance_budowy.
    Wywolywane z modulu Budzetowanie gdy uzytkownik uzupelnia brakujace dane przed protokolem."""
    existing = await db.finance_budowy.find_one({"id": budowa_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Budowa nie istnieje")
    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    updates["updated_at"] = datetime.now().isoformat()
    updates["updated_by"] = current_user["sub"]
    await db.finance_budowy.update_one({"id": budowa_id}, {"$set": updates})
    return {"ok": True}


@router.get("/budget/{budowa_id}/protokol-check")
async def protokol_check(budowa_id: str, _user: dict = Depends(get_current_admin)):
    """Sprawdza czy budowa ma kompletne dane do generowania protokolu.
    Zwraca {ready: bool, missing: [...], budowa: {...wybrane pola...}}"""
    budowa = await db.finance_budowy.find_one({"id": budowa_id}, {"_id": 0})
    if not budowa:
        raise HTTPException(404, "Budowa nie istnieje")
    missing = []
    if not (budowa.get("umowa_nr") or "").strip():
        missing.append("umowa_nr")
    if not (budowa.get("zamawiajacy") or "").strip():
        missing.append("zamawiajacy")
    return {
        "ready": len(missing) == 0,
        "missing": missing,
        "budowa": {
            "id": budowa["id"],
            "name": budowa.get("name", ""),
            "umowa_nr": budowa.get("umowa_nr", ""),
            "umowa_data": budowa.get("umowa_data", ""),
            "zamawiajacy": budowa.get("zamawiajacy", ""),
            "wykonawca": _resolve_wykonawca(budowa),
        },
    }


@router.get("/budget/{budowa_id}/protokol/{year}/{month}")
async def generate_protokol_xlsx(
    budowa_id: str,
    year: int,
    month: int,
    _user: dict = Depends(get_current_admin),
):
    """Generuje xlsx z protokolem miesiecznym w stylu firmowym FEGRRO:

    Naglowek (lewa strona - logo, prawa - tytul + nr umowy):
      PROTOKOL STANU ZAAWANSOWANIA ROBOT NR X    DO UMOWY    {nr_umowy}

    Pola:
      OKRES ROZLICZENIOWY:  {data_od}   DO   {data_do}
      ZAMAWIAJACY:          {dane_zamawiajacego}
      WYKONAWCA:            {dane_wykonawcy}

    Tabela 10 kolumn z 3 zielonymi naglowkami zlaczonymi:
      LP | Robocizna | Jd. | Ilosc | Cena | Wartosc | NARASTAJACO [WARTOSC %] | POPRZEDNI MIESIAC [WARTOSC %] | MIESIAC ROZLICZENIOWY [WARTOSC %]

    Sekcje (kategorie) jako szare wiersze grupowe bez wartosci.
    Wiersz RAZEM zielony z sumami.

    Stopka:
      DATA I MIEJSCE SPORZADZENIA PROTOKOLU: ...   DO ZAFAKTUROWANIA: {suma_miesiaca} zl NETTO
      Linie podpisow ZAMAWIAJACY / WYKONAWCA
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    import os

    budowa, lines, progress_curr, progress_prev, nr, stages_map = await _fetch_protokol_data(budowa_id, year, month)

    # ===== WORKBOOK =====
    wb = Workbook()
    ws = wb.active
    ws.title = f"Protokol {month:02d}-{year}"

    # === Style ===
    bold = Font(bold=True, size=10, name="Calibri")
    bold_big = Font(bold=True, size=14, name="Calibri")
    normal = Font(size=10, name="Calibri")
    green_fill = PatternFill("solid", fgColor="4F6343")  # Olive Green jak w aplikacji
    green_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")  # bialy tekst na olive
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(border_style="thin", color="000000")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # === LOGO + NAGLOWEK ===
    logo_path = "/app/frontend/public/icon-512x512.png"
    if os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.width = 110
            img.height = 110
            ws.add_image(img, "B2")
        except Exception:
            pass

    # Tytul + nr (kolumna F-K w gornej czesci)
    ws.merge_cells("F2:K2")
    ws["F2"] = f"PROTOKÓŁ STANU ZAAWANSOWANIA ROBÓT NR  {nr}"
    ws["F2"].font = bold_big
    ws["F2"].alignment = center
    ws.merge_cells("F3:K3")
    ws["F3"] = f"DO UMOWY  {budowa.get('umowa_nr', '')}"
    ws["F3"].font = bold
    ws["F3"].alignment = center

    # === POLA OKRES / ZAMAWIAJACY / WYKONAWCA ===
    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    d_from = f"{year:04d}-{month:02d}-01"
    d_to = f"{last_day:02d}.{month:02d}.{year:04d}"

    ws.merge_cells("B9:E9")
    ws["B9"] = "OKRES ROZLICZENIOWY:"
    ws["B9"].font = bold
    ws["B9"].alignment = left
    ws["F9"] = d_from
    ws["F9"].alignment = left
    ws["G9"] = "DO"
    ws["G9"].font = bold
    ws["G9"].alignment = center
    ws["H9"] = d_to
    ws["H9"].alignment = left

    ws.merge_cells("B11:E12")
    ws["B11"] = "ZAMAWIAJĄCY:"
    ws["B11"].font = bold
    ws["B11"].alignment = Alignment(horizontal="left", vertical="top")
    ws["F11"] = budowa.get("zamawiajacy", "")
    ws["F11"].alignment = left
    ws.merge_cells("F11:K12")
    ws.row_dimensions[11].height = 18
    ws.row_dimensions[12].height = 18

    ws.merge_cells("B14:E14")
    ws["B14"] = "WYKONAWCA:"
    ws["B14"].font = bold
    ws["B14"].alignment = left
    ws["F14"] = _resolve_wykonawca(budowa)
    ws["F14"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells("F14:K14")
    ws.row_dimensions[14].height = 32

    # === TABELA - PODWOJNE NAGLOWKI ===
    HEAD_TOP = 17  # wiersz z grupami NARASTAJACO / POPRZEDNI / MIESIAC ROZLICZ.
    HEAD_BOT = 18  # wiersz z LP/Robocizna/Jd./Ilosc/Cena/Wartosc + WARTOSC/%

    # Gorne komorki LP-Wartosc puste, dolne wypelnione
    ws.cell(row=HEAD_TOP, column=1, value="").fill = green_fill
    for c_idx in range(1, 7):
        ws.cell(row=HEAD_TOP, column=c_idx).fill = green_fill
        ws.cell(row=HEAD_TOP, column=c_idx).border = box
    # Trzy grupy: kol 7-8, 9-10, 11-12
    ws.merge_cells(start_row=HEAD_TOP, start_column=7, end_row=HEAD_TOP, end_column=8)
    g1 = ws.cell(row=HEAD_TOP, column=7, value="NARASTAJĄCO")
    g1.font = green_font
    g1.fill = green_fill
    g1.alignment = center
    g1.border = box
    ws.merge_cells(start_row=HEAD_TOP, start_column=9, end_row=HEAD_TOP, end_column=10)
    g2 = ws.cell(row=HEAD_TOP, column=9, value="POPRZEDNI MIESIĄC")
    g2.font = green_font
    g2.fill = green_fill
    g2.alignment = center
    g2.border = box
    ws.merge_cells(start_row=HEAD_TOP, start_column=11, end_row=HEAD_TOP, end_column=12)
    g3 = ws.cell(row=HEAD_TOP, column=11, value="MIESIĄC ROZLICZENIOWY")
    g3.font = green_font
    g3.fill = green_fill
    g3.alignment = center
    g3.border = box

    headers_bot = ["LP", "Robocizna", "Jd.", "Ilość", "Cena", "Wartość",
                   "WARTOŚĆ", "%", "WARTOŚĆ", "%", "WARTOŚĆ", "%"]
    for i, h in enumerate(headers_bot, start=1):
        cell = ws.cell(row=HEAD_BOT, column=i, value=h)
        cell.font = green_font
        cell.fill = green_fill
        cell.alignment = center
        cell.border = box
    ws.row_dimensions[HEAD_TOP].height = 22
    ws.row_dimensions[HEAD_BOT].height = 22

    # === WIERSZE TABELI z grupowaniem po ETAPIE ===
    r = HEAD_BOT + 1
    sum_budzet = 0.0
    sum_narast = 0.0
    sum_prev = 0.0
    sum_miesiac = 0.0
    lp = 1
    last_stage = "__init__"

    for ln in lines:
        sid = ln.get("stage_id")
        if sid != last_stage:
            stage = stages_map.get(sid)
            stage_name = (stage.get("name") if stage else "Bez etapu") or "Bez etapu"
            # Wiersz sekcji - tylko nazwa etapu, szare tlo
            for ci in range(1, 13):
                cell = ws.cell(row=r, column=ci, value="")
                cell.fill = gray_fill
                cell.border = box
            ws.cell(row=r, column=2, value=stage_name.upper()).font = bold
            ws.cell(row=r, column=2).alignment = left
            r += 1
            last_stage = sid

        plan = ln.get("plan_netto", 0)  # iter72: pozycje maja juz policzony plan
        # Nowa konwencja: progress_curr to przerob miesiaca, progress_prev to suma narast. do poprz.
        miesiac_pct = progress_curr.get(ln["id"], 0.0)
        prev_pct = progress_prev.get(ln["id"], 0.0)
        narast_pct = min(100.0, prev_pct + miesiac_pct)
        narast_val = round(plan * narast_pct / 100, 2)
        prev_val = round(plan * prev_pct / 100, 2)
        miesiac_val = round(plan * miesiac_pct / 100, 2)

        row_data = [
            lp, ln.get("name", ""), ln.get("unit") or "",
            ln.get("quantity", 0) or 0,
            ln.get("unit_price_netto", 0) or 0,
            plan,
            narast_val, narast_pct / 100.0,
            prev_val, prev_pct / 100.0,
            miesiac_val, miesiac_pct / 100.0,
        ]
        for i, v in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = box
            cell.font = normal
            if i == 1:
                cell.alignment = center
            elif i == 2:
                cell.alignment = left
            elif i == 3:
                cell.alignment = center
            else:
                cell.alignment = right
            # Formatowanie
            if i in (4,):
                cell.number_format = "#,##0.0"
            elif i in (5, 6, 7, 9, 11):
                cell.number_format = "#,##0.00 zł"
            elif i in (8, 10, 12):
                cell.number_format = "0.00%"

        sum_budzet += plan
        sum_narast += narast_val
        sum_prev += prev_val
        sum_miesiac += miesiac_val
        r += 1
        lp += 1

    # === WIERSZ RAZEM (zielony) ===
    razem_data = [
        "", "RAZEM", "", "", "", sum_budzet,
        sum_narast, (sum_narast / sum_budzet) if sum_budzet else 0.0,
        sum_prev, (sum_prev / sum_budzet) if sum_budzet else 0.0,
        sum_miesiac, (sum_miesiac / sum_budzet) if sum_budzet else 0.0,
    ]
    for i, v in enumerate(razem_data, start=1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = green_font
        cell.fill = green_fill
        cell.border = box
        if i == 2:
            cell.alignment = center
        elif isinstance(v, (int, float)):
            cell.alignment = right
            if i in (8, 10, 12):
                cell.number_format = "0%"
            else:
                cell.number_format = "#,##0.00 zł"
        else:
            cell.alignment = center

    # === SZEROKOSCI KOLUMN ===
    # B = nazwa pozycji (Robocizna), powinno byc dosc szerokie ale nie ekstremalne
    widths = [5, 38, 7, 10, 12, 14, 14, 8, 14, 8, 14, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === STOPKA ===
    foot = r + 3
    ws.cell(row=foot, column=2, value="DATA I MIEJSCE SPORZĄDZENIA PROTOKOŁU:").font = bold
    ws.cell(row=foot, column=2).alignment = left
    today = datetime.now().strftime("%d.%m.%Y")
    ws.cell(row=foot, column=6, value=today).alignment = left

    ws.cell(row=foot, column=10, value="DO ZAFAKTUROWANIA:").font = bold
    ws.cell(row=foot, column=10).alignment = right
    cell = ws.cell(row=foot, column=11, value=sum_miesiac)
    cell.font = bold
    cell.number_format = "#,##0.00 zł"
    cell.alignment = right
    ws.cell(row=foot, column=12, value="NETTO").font = bold
    ws.cell(row=foot, column=12).alignment = left

    # Linie podpisow
    sig = foot + 4
    ws.cell(row=sig, column=6, value="..................................").alignment = center
    ws.cell(row=sig, column=11, value="..................................").alignment = center
    ws.cell(row=sig + 1, column=6, value="Zamawiający\n(podpis i pieczęć)").alignment = Alignment(horizontal="center", wrap_text=True)
    ws.cell(row=sig + 1, column=11, value="Wykonawca\n(podpis i pieczęć)").alignment = Alignment(horizontal="center", wrap_text=True)
    ws.cell(row=sig + 1, column=6).font = bold
    ws.cell(row=sig + 1, column=11).font = bold
    ws.row_dimensions[sig + 1].height = 32

    # Wysokosci wierszy header
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    for rh in (5, 6, 7, 8):
        ws.row_dimensions[rh].height = 14  # spacer pod logo

    # === EXPORT ===
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in (budowa.get("name", "budowa")))
    filename = f"Protokol_{safe_name}_{year}-{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="protokol.xlsx"; filename*=UTF-8\'\'{quote(filename)}'
        },
    )

