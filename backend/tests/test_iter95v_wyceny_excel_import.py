"""iter95v: Backend tests for Wyceny Excel import (preview + apply).

Endpoints under test (admin-only):
  POST /api/wyceny/import/preview
  POST /api/wyceny/{wycena_id}/import/apply
"""
import base64
import io
import os
import uuid

import pytest
import requests
from openpyxl import Workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
assert BASE_URL, "REACT_APP_BACKEND_URL is required"

ADMIN_EMAIL = "admin@fegrro.pl"
ADMIN_PASSWORD = "Admin123!"


# ---------- fixtures ----------
@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(
        f"{BASE_URL}/api/auth/admin/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
        timeout=20,
    )
    if r.status_code != 200:
        pytest.skip(f"Admin login failed: {r.status_code} {r.text}")
    return r.json()["access_token"]


@pytest.fixture(scope="module")
def auth_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}", "Content-Type": "application/json"}


@pytest.fixture()
def wycena_id(auth_headers):
    """Create a TEST wycena, yield id, delete after."""
    name = f"TEST_iter95v_{uuid.uuid4().hex[:8]}"
    r = requests.post(f"{BASE_URL}/api/wyceny", json={"name": name}, headers=auth_headers, timeout=20)
    assert r.status_code in (200, 201), f"Create wycena failed: {r.status_code} {r.text}"
    wid = r.json().get("id") or r.json().get("wycena", {}).get("id")
    assert wid, f"No id in response: {r.json()}"
    yield wid
    # cleanup
    requests.delete(f"{BASE_URL}/api/wyceny/{wid}", headers=auth_headers, timeout=20)


# ---------- helpers ----------
def make_xlsx_b64(sheet_name: str = "Wycena"):
    """Return base64 XLSX with structure used in spec:
    Row0: header (Lp, Nazwa, Jednostka, Ilosc, Uwagi)
    Row1: stage 'Etap A'
    Row2: position 'Poz 1' m2 12,5
    Row3: position 'Poz 2' szt 3
    Row4: stage 'Etap B'
    Row5: position 'Poz 3' m 1 000,75 + uwaga
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["Lp", "Nazwa", "Jednostka", "Ilosc", "Uwagi"])
    ws.append(["", "Etap A", "", "", ""])
    ws.append(["1", "Poz 1", "m2", "12,5", ""])
    ws.append(["2", "Poz 2", "szt", "3", "test"])
    ws.append(["", "Etap B", "", "", ""])
    ws.append(["3", "Poz 3", "m", "1 000,75", "uw"])
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode("ascii")


# ---------- PREVIEW tests ----------
class TestPreview:
    def test_admin_only(self):
        r = requests.post(
            f"{BASE_URL}/api/wyceny/import/preview",
            json={"file_base64": make_xlsx_b64()},
            timeout=20,
        )
        assert r.status_code in (401, 403), f"expected 401/403, got {r.status_code}"

    def test_preview_ok(self, auth_headers):
        b64 = make_xlsx_b64("Wycena")
        r = requests.post(
            f"{BASE_URL}/api/wyceny/import/preview",
            json={"file_base64": b64},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert "sheets" in data and isinstance(data["sheets"], list)
        assert len(data["sheets"]) >= 1
        sh = data["sheets"][0]
        assert sh["name"] == "Wycena"
        assert sh["cols"] == 5
        assert sh["row_count"] == 6
        # row 2 has "Poz 1" in col index 1
        assert sh["rows"][2][1] == "Poz 1"
        assert sh["rows"][2][2] == "m2"
        # quantity preserved as string
        assert sh["rows"][2][3] in ("12,5", "12.5")

    def test_preview_with_data_uri_prefix(self, auth_headers):
        b64 = make_xlsx_b64()
        data_uri = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"
        r = requests.post(
            f"{BASE_URL}/api/wyceny/import/preview",
            json={"file_base64": data_uri},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        assert len(r.json()["sheets"]) >= 1

    def test_preview_specific_sheet(self, auth_headers):
        # Build XLSX with 2 sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "S1"
        ws1.append(["a"])
        ws2 = wb.create_sheet("S2")
        ws2.append(["b"])
        buf = io.BytesIO()
        wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
        r = requests.post(
            f"{BASE_URL}/api/wyceny/import/preview",
            json={"file_base64": b64, "sheet_name": "S2"},
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        sheets = r.json()["sheets"]
        assert len(sheets) == 1
        assert sheets[0]["name"] == "S2"
        assert sheets[0]["rows"][0][0] == "b"

    def test_preview_empty_file(self, auth_headers):
        r = requests.post(
            f"{BASE_URL}/api/wyceny/import/preview",
            json={"file_base64": ""},
            headers=auth_headers,
            timeout=20,
        )
        assert r.status_code == 400

    def test_preview_invalid_base64(self, auth_headers):
        # Valid base64 but not an xlsx -> should be 400
        r = requests.post(
            f"{BASE_URL}/api/wyceny/import/preview",
            json={"file_base64": "bm90LWFuLXhsc3g="},  # "not-an-xlsx"
            headers=auth_headers,
            timeout=20,
        )
        assert r.status_code == 400


# ---------- APPLY tests ----------
class TestApply:
    def test_admin_only(self, wycena_id):
        r = requests.post(
            f"{BASE_URL}/api/wyceny/{wycena_id}/import/apply",
            json={
                "file_base64": make_xlsx_b64(),
                "sheet_name": "Wycena",
                "name_col": 1,
                "rows": [{"row_index": 1, "role": "stage"}],
            },
            timeout=20,
        )
        assert r.status_code in (401, 403)

    def test_apply_404_unknown_wycena(self, auth_headers):
        r = requests.post(
            f"{BASE_URL}/api/wyceny/no-such-id-xyz/import/apply",
            json={
                "file_base64": make_xlsx_b64(),
                "sheet_name": "Wycena",
                "name_col": 1,
                "rows": [{"row_index": 1, "role": "stage"}],
            },
            headers=auth_headers,
            timeout=20,
        )
        assert r.status_code == 404

    def test_apply_full_flow(self, auth_headers, wycena_id):
        b64 = make_xlsx_b64("Wycena")
        body = {
            "file_base64": b64,
            "sheet_name": "Wycena",
            "name_col": 1,
            "unit_col": 2,
            "quantity_col": 3,
            "notes_col": 4,
            "rows": [
                {"row_index": 0, "role": "skip"},   # header
                {"row_index": 1, "role": "stage"},  # Etap A
                {"row_index": 2, "role": "position"},
                {"row_index": 3, "role": "position"},
                {"row_index": 4, "role": "stage"},  # Etap B
                {"row_index": 5, "role": "position"},
            ],
        }
        r = requests.post(
            f"{BASE_URL}/api/wyceny/{wycena_id}/import/apply",
            json=body,
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        out = r.json()
        assert out["ok"] is True
        assert out["stages_created"] == 2
        assert out["positions_created"] == 3
        assert out["skipped"] == 1

        # Verify structure via GET full
        r2 = requests.get(f"{BASE_URL}/api/wyceny/{wycena_id}/template", headers=auth_headers, timeout=20)
        assert r2.status_code == 200, r2.text
        data = r2.json()
        stages = data["stages"]
        assert len(stages) == 2
        assert stages[0]["name"] == "Etap A"
        assert stages[1]["name"] == "Etap B"
        # ordering preserved
        assert stages[0]["order"] < stages[1]["order"]
        # positions
        s1_pos = stages[0]["positions"]
        s2_pos = stages[1]["positions"]
        assert len(s1_pos) == 2
        assert len(s2_pos) == 1
        assert s1_pos[0]["name"] == "Poz 1"
        assert s1_pos[0]["unit"] == "m\u00b2"  # iter95v: normalized m2 -> m²
        assert s1_pos[0]["quantity"] == 12.5
        assert s1_pos[1]["name"] == "Poz 2"
        assert s1_pos[1]["quantity"] == 3.0
        assert s2_pos[0]["name"] == "Poz 3"
        # Polish thousand sep + comma decimal -> 1000.75
        assert s2_pos[0]["quantity"] == 1000.75

    def test_apply_auto_default_stage(self, auth_headers, wycena_id):
        """Position before any stage -> create default 'Etap 1'."""
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws.append(["Orphan Pos", "szt", "1"])
        ws.append(["Etap X", "", ""])
        ws.append(["P2", "m", "2"])
        buf = io.BytesIO()
        wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")

        body = {
            "file_base64": b64,
            "sheet_name": "S",
            "name_col": 0,
            "unit_col": 1,
            "quantity_col": 2,
            "rows": [
                {"row_index": 0, "role": "position"},
                {"row_index": 1, "role": "stage"},
                {"row_index": 2, "role": "position"},
            ],
        }
        r = requests.post(
            f"{BASE_URL}/api/wyceny/{wycena_id}/import/apply",
            json=body,
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        out = r.json()
        # default Etap 1 + Etap X = 2 stages, 2 positions
        assert out["stages_created"] == 2
        assert out["positions_created"] == 2

        # First stage created should be 'Etap 1' (default)
        r2 = requests.get(f"{BASE_URL}/api/wyceny/{wycena_id}/template", headers=auth_headers, timeout=20)
        stages = r2.json()["stages"]
        names = [s["name"] for s in stages]
        assert "Etap 1" in names
        assert "Etap X" in names

    def test_apply_skip_and_empty_name(self, auth_headers, wycena_id):
        """skip role + empty name in name_col should both increment skipped, no creates."""
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws.append(["", "", ""])  # empty name row
        ws.append(["X", "", ""])
        buf = io.BytesIO()
        wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")

        body = {
            "file_base64": b64,
            "sheet_name": "S",
            "name_col": 0,
            "rows": [
                {"row_index": 0, "role": "stage"},   # empty -> skipped
                {"row_index": 1, "role": "skip"},    # explicit skip
            ],
        }
        r = requests.post(
            f"{BASE_URL}/api/wyceny/{wycena_id}/import/apply",
            json=body,
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        out = r.json()
        assert out["stages_created"] == 0
        assert out["positions_created"] == 0
        assert out["skipped"] == 2

    def test_apply_optional_cols_none_no_crash(self, auth_headers, wycena_id):
        """unit_col, quantity_col, notes_col all None should not crash."""
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws.append(["Etap"])
        ws.append(["Poz"])
        buf = io.BytesIO()
        wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")

        body = {
            "file_base64": b64,
            "sheet_name": "S",
            "name_col": 0,
            "unit_col": None,
            "quantity_col": None,
            "notes_col": None,
            "rows": [
                {"row_index": 0, "role": "stage"},
                {"row_index": 1, "role": "position"},
            ],
        }
        r = requests.post(
            f"{BASE_URL}/api/wyceny/{wycena_id}/import/apply",
            json=body,
            headers=auth_headers,
            timeout=30,
        )
        assert r.status_code == 200, r.text
        out = r.json()
        assert out["stages_created"] == 1
        assert out["positions_created"] == 1

        # GET and verify position has quantity=None, unit=None
        r2 = requests.get(f"{BASE_URL}/api/wyceny/{wycena_id}/template", headers=auth_headers, timeout=20)
        st = r2.json()["stages"]
        assert len(st) == 1
        assert st[0]["positions"][0]["quantity"] is None
        assert st[0]["positions"][0]["unit"] is None

    def test_apply_append_to_existing_stages(self, auth_headers, wycena_id):
        """If wycena already has stages, new ones should be appended (higher order)."""
        # First import 1 stage
        b64_a = make_xlsx_b64()
        r1 = requests.post(
            f"{BASE_URL}/api/wyceny/{wycena_id}/import/apply",
            json={
                "file_base64": b64_a,
                "sheet_name": "Wycena",
                "name_col": 1,
                "rows": [{"row_index": 1, "role": "stage"}],  # Etap A
            },
            headers=auth_headers,
            timeout=30,
        )
        assert r1.status_code == 200, r1.text
        # Second import - another stage
        r2 = requests.post(
            f"{BASE_URL}/api/wyceny/{wycena_id}/import/apply",
            json={
                "file_base64": b64_a,
                "sheet_name": "Wycena",
                "name_col": 1,
                "rows": [{"row_index": 4, "role": "stage"}],  # Etap B
            },
            headers=auth_headers,
            timeout=30,
        )
        assert r2.status_code == 200, r2.text

        rf = requests.get(f"{BASE_URL}/api/wyceny/{wycena_id}/template", headers=auth_headers, timeout=20)
        stages = rf.json()["stages"]
        assert len(stages) == 2
        # Order should be ascending; Etap A first (order=0), Etap B second (order=1)
        assert stages[0]["name"] == "Etap A"
        assert stages[1]["name"] == "Etap B"
        assert stages[0]["order"] < stages[1]["order"]
