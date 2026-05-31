"""
iter95w - Eksport wyceny:
  * Klient PDF/XLSX naglowek z NIP + telefon + email FeGrro (+ logo)
  * Sekcje 'Oferta obejmuje' / 'Oferta nie obejmuje' z bulletami
  * Szersza kolumna B w XLSX (Nazwa pozycji), wrap_text, szerokosci A..F
  * Backend: WycenaUpdate przyjmuje scope_includes / scope_excludes
"""

import os
import io
import re
import uuid
import pytest
import requests
from openpyxl import load_workbook

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
ADMIN_EMAIL = "admin@fegrro.pl"
ADMIN_PASSWORD = "Admin123!"


# ------------------------------- fixtures -------------------------------

@pytest.fixture(scope="module")
def admin_client():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/admin/login",
               json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, f"admin login failed: {r.status_code} {r.text}"
    tok = r.json().get("token") or r.json().get("access_token")
    assert tok, f"no token in login response: {r.text}"
    s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


@pytest.fixture()
def wycena(admin_client):
    """Tworzy minimalna wycene z jednym etapem i jedna pozycja - dla eksportu."""
    name = f"TEST_iter95w_{uuid.uuid4().hex[:6]}"
    r = admin_client.post(f"{BASE_URL}/api/wyceny", json={
        "name": name,
        "client_name": "TEST Klient Sp. z o.o.",
        "client_nip": "111-222-33-44",
        "client_address": "ul. Testowa 1\n00-001 Warszawa",
    })
    assert r.status_code in (200, 201), r.text
    w = r.json()
    wid = w["id"]

    # add a stage + position (so position table renders)
    rs = admin_client.post(f"{BASE_URL}/api/wyceny/stages",
                           json={"wycena_id": wid, "name": "Etap 1", "order": 0})
    assert rs.status_code in (200, 201), rs.text
    stage_id = rs.json()["id"]
    rp = admin_client.post(f"{BASE_URL}/api/wyceny/positions", json={
        "wycena_id": wid,
        "stage_id": stage_id,
        "name": "Bardzo dluga nazwa pozycji testowej dla weryfikacji szerokosci kolumny B oraz dzialania wrap_text w XLSX",
        "order": 0,
        "quantity": 2,
        "unit": "szt",
    })
    assert rp.status_code in (200, 201), rp.text

    yield wid

    admin_client.delete(f"{BASE_URL}/api/wyceny/{wid}")


# ---------------- 1) WycenaUpdate przyjmuje nowe pola ----------------

class TestWycenaUpdateScopeFields:
    def test_patch_scope_includes_excludes_persists(self, admin_client, wycena):
        scope_inc = "Roboty murowe\nIzolacja przeciwwilgociowa\nWylewki anhydrytowe"
        scope_exc = "Wyposazenie meblowe\nUrzadzenia AGD"
        r = admin_client.patch(f"{BASE_URL}/api/wyceny/{wycena}", json={
            "scope_includes": scope_inc,
            "scope_excludes": scope_exc,
        })
        assert r.status_code == 200, r.text

        # GET to verify persistence
        g = admin_client.get(f"{BASE_URL}/api/wyceny/{wycena}/template")
        assert g.status_code == 200, g.text
        body = g.json()
        w = body.get("wycena") if isinstance(body, dict) else None
        assert w, f"unexpected template response: {body}"
        assert w.get("scope_includes") == scope_inc
        assert w.get("scope_excludes") == scope_exc

    def test_patch_accepts_empty_scope(self, admin_client, wycena):
        # set, then clear by sending empty string (Optional[str] -> "" allowed)
        admin_client.patch(f"{BASE_URL}/api/wyceny/{wycena}",
                           json={"scope_includes": "Cos"})
        r = admin_client.patch(f"{BASE_URL}/api/wyceny/{wycena}",
                               json={"scope_includes": ""})
        assert r.status_code == 200, r.text
        g = admin_client.get(f"{BASE_URL}/api/wyceny/{wycena}/template").json()
        assert (g["wycena"].get("scope_includes") or "") == ""


# ---------------- 2) XLSX (client variant) -------------------------------

class TestClientXlsxExport:
    def _fetch_xlsx(self, admin_client, wid):
        r = admin_client.get(
            f"{BASE_URL}/api/wyceny/{wid}/export.xlsx",
            params={"detail": "client"},
        )
        assert r.status_code == 200, f"xlsx export status {r.status_code}: {r.text[:300]}"
        ct = r.headers.get("Content-Type", "")
        assert "spreadsheetml" in ct or "xlsx" in ct.lower() or "officedocument" in ct, ct
        return load_workbook(io.BytesIO(r.content), data_only=False)

    def test_xlsx_header_rows_have_company_data(self, admin_client, wycena):
        wb = self._fetch_xlsx(admin_client, wycena)
        ws = wb.active
        # B1..B4 - naglowek firmowy
        assert (ws["B1"].value or "").strip() == "FeGrro"
        assert "589-206-61-74" in (ws["B2"].value or "")
        assert "885 213 273" in (ws["B3"].value or "")
        assert (ws["B4"].value or "").strip().lower() == "biuro@fegrro.pl"

    def test_xlsx_column_widths(self, admin_client, wycena):
        wb = self._fetch_xlsx(admin_client, wycena)
        ws = wb.active
        expected = {"A": 8, "B": 65, "C": 13, "D": 9, "E": 16, "F": 18}
        for col, w in expected.items():
            got = ws.column_dimensions[col].width
            assert got == pytest.approx(w, rel=1e-3), (
                f"column {col} width expected {w}, got {got}"
            )

    def test_xlsx_position_name_cell_wraps_text(self, admin_client, wycena):
        """Komorka pozycji w kol. B powinna miec wrap_text=True."""
        wb = self._fetch_xlsx(admin_client, wycena)
        ws = wb.active
        long_name_marker = "Bardzo dluga nazwa pozycji testowej"
        found = False
        for row in ws.iter_rows(min_col=2, max_col=2):
            cell = row[0]
            val = cell.value or ""
            if isinstance(val, str) and long_name_marker in val:
                assert cell.alignment is not None
                assert cell.alignment.wrap_text is True, (
                    "wrap_text powinno byc True dla nazwy pozycji"
                )
                found = True
                break
        assert found, "Nie znaleziono pozycji testowej w kolumnie B"

    def test_xlsx_scope_sections_render_with_bullets(self, admin_client, wycena):
        # set scope first
        admin_client.patch(f"{BASE_URL}/api/wyceny/{wycena}", json={
            "scope_includes": "Pierwsza pozycja\nDruga pozycja",
            "scope_excludes": "Wykluczenie A\nWykluczenie B",
        })
        wb = self._fetch_xlsx(admin_client, wycena)
        ws = wb.active
        all_text = "\n".join(
            str(c.value) for c in ws[1] + ws[2]  # dummy; we'll iterate full sheet below
            if c.value
        )
        # collect every cell as text
        cells = []
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    cells.append(str(c.value))
        combined = "\n".join(cells)
        assert "Oferta obejmuje" in combined
        assert "Oferta nie obejmuje" in combined
        # bullet prefix
        assert "\u2022 Pierwsza pozycja" in combined
        assert "\u2022 Druga pozycja" in combined
        assert "\u2022 Wykluczenie A" in combined
        assert "\u2022 Wykluczenie B" in combined

    def test_xlsx_no_scope_sections_when_empty(self, admin_client, wycena):
        # ensure scope cleared
        admin_client.patch(f"{BASE_URL}/api/wyceny/{wycena}",
                           json={"scope_includes": "", "scope_excludes": ""})
        wb = self._fetch_xlsx(admin_client, wycena)
        ws = wb.active
        combined = "\n".join(
            str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
        )
        assert "Oferta obejmuje" not in combined
        assert "Oferta nie obejmuje" not in combined


# ---------------- 3) PDF (client variant) -------------------------------

class TestClientPdfExport:
    def _fetch_pdf(self, admin_client, wid):
        r = admin_client.get(
            f"{BASE_URL}/api/wyceny/{wid}/export.pdf",
            params={"detail": "client"},
        )
        assert r.status_code == 200, f"pdf export status {r.status_code}: {r.text[:300]}"
        assert r.headers.get("Content-Type", "").lower().startswith("application/pdf")
        assert r.content[:4] == b"%PDF", "no %PDF magic"
        return r.content

    def _extract_text(self, pdf_bytes: bytes) -> str:
        if PdfReader is None:
            return ""
        try:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            return "\n".join((p.extract_text() or "") for p in reader.pages)
        except Exception:
            return ""

    def test_pdf_header_contains_company_data(self, admin_client, wycena):
        pdf = self._fetch_pdf(admin_client, wycena)
        text = self._extract_text(pdf)
        # treat extracted text as primary; fallback to raw bytes scan
        haystack = text if text else pdf.decode("latin-1", errors="ignore")
        assert "FeGrro" in haystack, "FeGrro missing in PDF header"
        # NIP - may render with or without separators, accept both forms
        assert ("589-206-61-74" in haystack) or ("5892066174" in haystack), \
            "NIP 589-206-61-74 missing in PDF"
        assert ("885 213 273" in haystack) or ("885213273" in haystack), \
            "Tel 885 213 273 missing in PDF"
        assert "biuro@fegrro.pl" in haystack, "email missing in PDF"

    def test_pdf_scope_sections_with_bullets(self, admin_client, wycena):
        admin_client.patch(f"{BASE_URL}/api/wyceny/{wycena}", json={
            "scope_includes": "Murowanie scian\nIzolacje pionowe",
            "scope_excludes": "Meble kuchenne\nLampy sufitowe",
        })
        pdf = self._fetch_pdf(admin_client, wycena)
        text = self._extract_text(pdf)
        haystack = text if text else pdf.decode("latin-1", errors="ignore")

        assert "Oferta obejmuje" in haystack
        assert "Oferta nie obejmuje" in haystack
        # bullet glyph (UTF-8 in raw stream) or ASCII fallback
        bullet_present = (
            "\u2022" in haystack
            or "•" in haystack
            or b"\xe2\x80\xa2" in pdf
        )
        assert bullet_present, "bullet character (\u2022) not found in PDF"

    def test_pdf_no_scope_when_empty(self, admin_client, wycena):
        admin_client.patch(f"{BASE_URL}/api/wyceny/{wycena}",
                           json={"scope_includes": "", "scope_excludes": ""})
        pdf = self._fetch_pdf(admin_client, wycena)
        text = self._extract_text(pdf)
        haystack = text if text else pdf.decode("latin-1", errors="ignore")
        assert "Oferta obejmuje" not in haystack
        assert "Oferta nie obejmuje" not in haystack

    def test_pdf_contains_image_object(self, admin_client, wycena):
        """logo 32mm jest zalaczone do PDF jako image (XObject)."""
        pdf = self._fetch_pdf(admin_client, wycena)
        # ReportLab zawsze produkuje XObject typu Image gdy zalaczone jest Image()
        assert b"/Subtype /Image" in pdf or b"/Subtype/Image" in pdf, \
            "PDF nie zawiera zadnego obrazka (logo nie wstawione?)"
